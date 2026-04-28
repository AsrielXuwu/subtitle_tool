import difflib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from tkinter import font as tkfont
from tkinter import scrolledtext
import pandas as pd
import os
import re
import json
import zipfile
import math
import platform
import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openai import AzureOpenAI, NOT_GIVEN
import base64
import sys
import threading
# ---------------- 新增：多线程预备 ----------------
from concurrent.futures import ThreadPoolExecutor, as_completed

# ================= 核心配置区 =================
DEFAULT_API_VERSION = "2025-04-01-preview"

# 动态获取当前脚本/软件所在的绝对目录（兼容直接运行 .py 和打包后的 .exe）
if getattr(sys, 'frozen', False):
    CONFIG_DIR = os.path.dirname(sys.executable)
else:
    CONFIG_DIR = os.path.dirname(os.path.abspath(__file__))

# 强制将配置文件路径绑定到软件所在目录
LQA_CONFIG_FILE = os.path.join(CONFIG_DIR, "lqa_api_config.enc")
SECRET_KEY = b"LQA_TOOL_SECURE_XOR_KEY_2026_!@"

def encrypt_data(data_dict):
    """字节级异或 + Base64 加密"""
    text = json.dumps(data_dict).encode('utf-8')
    encrypted = bytearray(b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(text))
    return base64.b64encode(encrypted).decode('utf-8')

def decrypt_data(b64_text):
    """Base64 解密 + 字节级异或还原"""
    encrypted = base64.b64decode(b64_text)
    decrypted = bytearray(b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(encrypted))
    return json.loads(decrypted.decode('utf-8'))
# ==============================================

ENGINES_MAP = {
    'GPT-5.3-chat': 'gpt-5.3-chat-2026-03-03',
    'GPT-5.2-chat': 'gpt-5.2-chat-2025-12-11',
    'GPT-o3-mini': 'o3-mini-2025-01-31'
}

# ======= LQA 目标语言映射表 (包含所有要求及常见语种) =======
LANGUAGES_MAP = {
    "English (United States)": "en-US",
    "English (United Kingdom)": "en-GB",
    "Indonesian (Bahasa Indonesia)": "id-ID",
    "Portuguese (Brazil)": "pt-BR",
    "Portuguese (Portugal)": "pt-PT",
    "Polish (Polski)": "pl-PL",
    "Turkish (Türkçe)": "tr-TR",
    "Italian (Italiano)": "it-IT",
    "Russian (Русский)": "ru-RU",
    "Romanian (Română)": "ro-RO",
    "German (Deutsch)": "de-DE",
    "French (Français)": "fr-FR",
    "Chinese (Simplified)": "zh-CN",
    "Chinese (Traditional, Taiwan)": "zh-TW",
    "Chinese (Traditional, Hong Kong)": "zh-HK",
    "Bulgarian (Български)": "bg-BG",
    "Filipino (Tagalog)": "tl-PH",
    "Czech (Čeština)": "cs-CZ",
    "Vietnamese (Tiếng Việt)": "vi-VN",
    "Hindi (हिन्दी)": "hi-IN",
    "Spanish (Spain)": "es-ES",
    "Spanish (Latin America)": "es-419",
    "Japanese (日本語)": "ja-JP",
    "Korean (한국어)": "ko-KR",
    "Thai (ไทย)": "th-TH",
    "Arabic (العربية)": "ar-SA",
    "Dutch (Nederlands)": "nl-NL",
    "Greek (Ελληνικά)": "el-GR",
    "Hungarian (Magyar)": "hu-HU",
    "Swedish (Svenska)": "sv-SE",
    "Danish (Dansk)": "da-DK",
    "Finnish (Suomi)": "fi-FI",
    "Norwegian (Norsk)": "no-NO",
    "Malay (Bahasa Melayu)": "ms-MY",
    "Ukrainian (Українська)": "uk-UA"
}

# ======= 连续断句处理：无空格语言集合 =======
# 这里的语言在跨行拼接时，AI不会强行插入空格 (加入了台繁、港繁、日文、泰文等)
NO_SPACE_LANGS = {
    "zh-CN", "zh-TW", "zh-HK", "ja-JP", "th-TH", 
    "Chinese (Simplified)", "Chinese (Traditional, Taiwan)", "Chinese (Traditional, Hong Kong)", 
    "Japanese (日本語)", "Thai (ไทย)"
}

class LQA_App:
    def browse_tb_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("CSV/Excel Files", "*.csv;*.xlsx")])
        if filepath:
            self.tb_file_path.set(filepath)

    def browse_term_file(self):
        filepath = filedialog.askopenfilename(
            title="选择术语表文件",
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("CSV Files", "*.csv")]
        )
        if filepath:
            self.term_file_path.set(filepath)

    def scan_type_column(self):
        term_file = self.term_file_path.get()
        if not term_file or not os.path.exists(term_file):
            messagebox.showwarning("警告", "请先选择术语表文件。")
            return
            
        type_col_letter = self.term_type_col.get().strip().upper()
        if not type_col_letter:
            messagebox.showwarning("警告", "请填写术语表中 Type 列所在的字母（如 C）。")
            return
            
        try:
            from openpyxl.utils import column_index_from_string
            type_idx = column_index_from_string(type_col_letter) - 1 # 转换为 pandas 的 0 索引
            
            df = pd.read_excel(term_file, sheet_name=0)
            if type_idx >= len(df.columns):
                messagebox.showwarning("警告", "填写的 Type 列字母超出了表格实际列数。")
                return
                
            unique_types = df.iloc[:, type_idx].dropna().astype(str).unique().tolist()
            
            self.type_listbox.delete(0, tk.END)
            for t in sorted(unique_types):
                self.type_listbox.insert(tk.END, t)
                
            self.log(f"成功扫描术语表，提取 {len(unique_types)} 个唯一类型。")
        except Exception as e:
            messagebox.showerror("错误", f"扫描术语表时发生异常: {e}")

    def apply_term_rich_text(self, cell_rich_text, matched_terms, ignore_case=True):
        """处理富文本覆盖逻辑：彻底解决 WPS/Excel 绿色溢出问题。"""
        if not matched_terms:
            return cell_rich_text

        # 将纯文本转为列表，无缝进入下方的组装逻辑
        if isinstance(cell_rich_text, str):
            cell_rich_text = [cell_rich_text]
        elif not isinstance(cell_rich_text, CellRichText):
            return cell_rich_text

        # 术语的绿色加粗标准格式
        green_bold = InlineFont(color="FF00B050", b=True)
        # 标记纯文本的防溢出标识符
        PLAIN_TEXT_FLAG = "PLAIN_TEXT"

        new_blocks = []
        
        sorted_terms = sorted(matched_terms, key=len, reverse=True)
        # 防御性清理空字符
        sorted_terms = [t for t in sorted_terms if t.strip()] 
        if not sorted_terms:
            return CellRichText(*cell_rich_text)
            
        flags = re.IGNORECASE if ignore_case else 0

        for block in cell_rich_text:
            if isinstance(block, str):
                text = block
                current_font = PLAIN_TEXT_FLAG
            else:
                text = block.text
                current_font = block.font
                # 【核心修复】：如果富文本块原先没有指定颜色，强制打上纯文本标记
                if not current_font or not current_font.color:
                    current_font = PLAIN_TEXT_FLAG
                
            segments = [(text, current_font)]
            for term in sorted_terms:
                temp_segments = []
                for seg_text, seg_font in segments:
                    if seg_font == green_bold:
                        temp_segments.append((seg_text, seg_font))
                        continue
                        
                    parts = re.split(f'({re.escape(term)})', seg_text, flags=flags)
                    for part in parts:
                        if not part:
                            continue
                        if (ignore_case and part.lower() == term.lower()) or (not ignore_case and part == term):
                            temp_segments.append((part, green_bold))
                        else:
                            temp_segments.append((part, seg_font))
                segments = temp_segments
                
            for seg_text, seg_font in segments:
                if seg_text:
                    # 【核心修复】：如果是纯文本，直接作为字符串插入！
                    # 不生成任何带属性的空 TextBlock，WPS 找不到空标签，自然无法传染颜色！
                    if seg_font == PLAIN_TEXT_FLAG:
                        new_blocks.append(seg_text)
                    else:
                        new_blocks.append(TextBlock(font=seg_font, text=seg_text))
                
        return CellRichText(*new_blocks)
    
    def __init__(self, parent_frame):
        self.parent = parent_frame  # 保存父容器的作用域引用
        
        
        self.file_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.stop_flag = False
        self.api_endpoint = tk.StringVar()  # 新增：绑定 UI 的 Endpoint
        self.api_key = tk.StringVar()       # 新增：绑定 UI 的 Key
        
        self.setup_ui()

    def setup_ui(self):
        # 直接使用外部传进来的、已经带有滚动条的父容器
        self.scrollable_frame = self.parent

        # --- 术语检查配置区 ---
        frame_term = ttk.LabelFrame(self.scrollable_frame, text="术语标记 (选中的Type为部分匹配)", padding=10)
        frame_term.pack(fill="x", padx=10, pady=5)
        frame_term.columnconfigure(1, weight=1)

        self.term_file_path = tk.StringVar()
        self.term_output_col = tk.StringVar(value="G")

        # --- Row 0: 术语表文件与输出列 ---
        ttk.Label(frame_term, text="术语表文件:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_term, textvariable=self.term_file_path, width=35).grid(row=0, column=1, columnspan=4, padx=5, sticky="we")
        ttk.Button(frame_term, text="浏览...", command=self.browse_term_file).grid(row=0, column=5, padx=5)

        ttk.Label(frame_term, text="输出展示列(如G):").grid(row=0, column=6, sticky="w", padx=(5, 2))
        ttk.Entry(frame_term, textvariable=self.term_output_col, width=4).grid(row=0, column=7, sticky="w")

        # --- Row 1: 选项与列字母指定区 (重新排版) ---
        self.enable_term_check_var = tk.BooleanVar(value=True) # 术语标记总开关
        self.ignore_case_var = tk.BooleanVar(value=True)
        self.term_src_col = tk.StringVar(value="A")
        self.term_tgt_col = tk.StringVar(value="B")
        self.term_type_col = tk.StringVar(value="C")

        # 开关与选项前置
        ttk.Checkbutton(frame_term, text="启用术语标记", variable=self.enable_term_check_var).grid(row=1, column=0, sticky="w", pady=5)
        ttk.Checkbutton(frame_term, text="忽略大小写", variable=self.ignore_case_var).grid(row=1, column=1, sticky="w", pady=5)

        ttk.Label(frame_term, text="原文列:").grid(row=1, column=2, sticky="e", padx=(5,2))
        ttk.Entry(frame_term, textvariable=self.term_src_col, width=4).grid(row=1, column=3, sticky="w")

        ttk.Label(frame_term, text="译文列:").grid(row=1, column=4, sticky="e", padx=(5,2))
        ttk.Entry(frame_term, textvariable=self.term_tgt_col, width=4).grid(row=1, column=5, sticky="w")

        ttk.Label(frame_term, text="Type列:").grid(row=1, column=6, sticky="e", padx=(5,2))
        ttk.Entry(frame_term, textvariable=self.term_type_col, width=4).grid(row=1, column=7, sticky="w")

        # --- Row 2: 扫描与多选区 ---
        ttk.Button(frame_term, text="扫描术语表 Type", command=self.scan_type_column).grid(row=2, column=0, pady=10, sticky="nw")

        type_list_frame = ttk.Frame(frame_term)
        type_list_frame.grid(row=2, column=1, columnspan=7, sticky="we", pady=5)
        self.type_listbox = tk.Listbox(type_list_frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        self.type_listbox.pack(side=tk.LEFT, fill="both", expand=True)
        type_scrollbar = ttk.Scrollbar(type_list_frame, orient="vertical", command=self.type_listbox.yview)
        type_scrollbar.pack(side=tk.RIGHT, fill="y")
        self.type_listbox.config(yscrollcommand=type_scrollbar.set)

        # --- 0. API 接口配置区 (新增) ---
        frame_api = ttk.LabelFrame(self.scrollable_frame, text="0. API 接口配置 (自动加密存储本地)", padding=10)
        frame_api.pack(fill="x", padx=10, pady=5)
        frame_api.columnconfigure(1, weight=1)

        ttk.Label(frame_api, text="API Azure Site:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_api, textvariable=self.api_endpoint, width=60).grid(row=0, column=1, padx=5, pady=2, sticky="we")
        #ttk.Button(api_btn_frame, text="💾 保存配置", command=self.save_config).pack(side="top", fill="x", pady=(0, 2))
        ttk.Button(frame_api, text="💾 保存配置", command=self.save_config).grid(row=0, column=2, pady=2)
        ttk.Label(frame_api, text="API Key:").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame_api, textvariable=self.api_key, width=60, show="*").grid(row=1, column=1, padx=5, pady=2, sticky="we")
        #ttk.Button(api_btn_frame, text="📂 加载配置", command=self.manual_load_config).pack(side="top", fill="x", pady=(2, 0))
        ttk.Button(frame_api, text="📂 加载配置", command=self.manual_load_config).grid(row=1, column=2, pady=2)
        # 增加一个 Frame 用来纵向堆叠这两个按钮
        api_btn_frame = ttk.Frame(frame_api)
        api_btn_frame.grid(row=0, column=2, rowspan=2, padx=10, pady=2)
        
        # --- 1. 文件设置区 ---
        frame_file = ttk.LabelFrame(self.scrollable_frame, text="1. 文件设置", padding=10)
        frame_file.pack(fill="x", padx=10, pady=5)
        frame_file.columnconfigure(1, weight=1)
        
        ttk.Label(frame_file, text="输入 Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_file, textvariable=self.file_path, width=60).grid(row=0, column=1, padx=5, pady=2, sticky="we")
        ttk.Button(frame_file, text="浏览...", command=self.browse_file).grid(row=0, column=2, pady=2)

        ttk.Label(frame_file, text="输出位置(前缀):").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame_file, textvariable=self.output_path, width=60).grid(row=1, column=1, padx=5, pady=2, sticky="we")
        ttk.Button(frame_file, text="另存为...", command=self.browse_output_file).grid(row=1, column=2, pady=2)
        
        # --- 性能设置区 (新增) ---
        frame_perf = ttk.LabelFrame(self.scrollable_frame, text="性能与并发设置", padding=10)
        frame_perf.pack(fill="x", padx=10, pady=5)
        
        self.use_multithread_var = tk.BooleanVar(value=False)
        self.thread_count_var = tk.IntVar(value=5) # 默认5个并发
        self.retry_count_var = tk.IntVar(value=3)  # 新增：默认失败重试3次
        
        ttk.Checkbutton(frame_perf, text="开启多线程并发", variable=self.use_multithread_var).grid(row=0, column=0, sticky="w")
        
        ttk.Label(frame_perf, text="并发线程数:").grid(row=0, column=1, sticky="w", padx=(15, 2))
        ttk.Spinbox(frame_perf, from_=1, to=20, textvariable=self.thread_count_var, width=4).grid(row=0, column=2, sticky="w")
        
        # 新增：重试次数的 UI 组件
        ttk.Label(frame_perf, text="失败重试次数:").grid(row=0, column=3, sticky="w", padx=(15, 2))
        ttk.Spinbox(frame_perf, from_=0, to=10, textvariable=self.retry_count_var, width=4).grid(row=0, column=4, sticky="w")

        # === 新增：输出模式选择 ===
        ttk.Label(frame_file, text="输出模式:").grid(row=2, column=0, sticky="w", pady=5)
        self.var_output_mode = tk.StringVar(value="split")
        mode_frame = ttk.Frame(frame_file)
        mode_frame.grid(row=2, column=1, columnspan=2, sticky="w", pady=2)
        ttk.Radiobutton(mode_frame, text="分开输出 (按Sheet和集数独立文件)", variable=self.var_output_mode, value="split").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(mode_frame, text="合并输出 (全部结果保存在单一文件中)", variable=self.var_output_mode, value="merged").pack(side="left")
        # ==========================

        # --- 2. 列配置区 ---
        frame_col = ttk.LabelFrame(self.scrollable_frame, text="2. 表格列号配置 (A=1, B=2...)", padding=10)
        frame_col.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_col, text="原文列号:").grid(row=0, column=0, sticky="w")
        self.col_src = ttk.Entry(frame_col, width=5)
        self.col_src.insert(0, "1")
        self.col_src.grid(row=0, column=1, padx=5, sticky="w")
        
        ttk.Label(frame_col, text="译文列号:").grid(row=0, column=2, sticky="w", padx=(10, 0))
        self.col_tgt = ttk.Entry(frame_col, width=5)
        self.col_tgt.insert(0, "2")
        self.col_tgt.grid(row=0, column=3, padx=5, sticky="w")
        
        ttk.Label(frame_col, text="集数列号:").grid(row=0, column=4, sticky="w", padx=(10, 0))
        self.col_ep = ttk.Entry(frame_col, width=5)
        self.col_ep.insert(0, "3")
        self.col_ep.grid(row=0, column=5, padx=5, sticky="w")

        ttk.Label(frame_col, text="输出列号:").grid(row=1, column=0, sticky="w", pady=5)
        self.col_res = ttk.Entry(frame_col, width=5)
        self.col_res.insert(0, "4")
        self.col_res.grid(row=1, column=1, padx=5, sticky="w", pady=5)

        ttk.Label(frame_col, text="起始行号:").grid(row=1, column=2, sticky="w", padx=(10, 0), pady=5)
        self.row_start = ttk.Entry(frame_col, width=5)
        self.row_start.insert(0, "2")
        self.row_start.grid(row=1, column=3, padx=5, sticky="w", pady=5)

        self.var_with_source = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_col, text="双语模式 (参考原文)", variable=self.var_with_source).grid(row=1, column=4, columnspan=2, sticky="w", padx=(10, 0))

        # --- 3. 工作表 (Sheet) 过滤 ---
        frame_sheets = ttk.LabelFrame(self.scrollable_frame, text="3. 工作表 (Sheet) 选择 (不选默认仅检查第1个Sheet)", padding=10)
        frame_sheets.pack(fill="x", padx=10, pady=5)
        
        self.btn_load_sheets = ttk.Button(frame_sheets, text="📑 扫描 Sheet", command=self.load_sheets)
        self.btn_load_sheets.pack(side="left", padx=5)

        self.sheet_listbox = tk.Listbox(frame_sheets, selectmode=tk.MULTIPLE, height=3, exportselection=False)
        self.sheet_listbox.pack(side="left", fill="both", expand=True, padx=5)
        
        sheet_scrollbar = ttk.Scrollbar(frame_sheets, orient="vertical", command=self.sheet_listbox.yview)
        sheet_scrollbar.pack(side="left", fill="y")
        self.sheet_listbox.config(yscrollcommand=sheet_scrollbar.set)

        # --- 4. 集数过滤区 ---
        frame_ep = ttk.LabelFrame(self.scrollable_frame, text="4. 集数过滤 (可选)", padding=10)
        frame_ep.pack(fill="x", padx=10, pady=5)
        
        # 【修复：恢复被误删的扫描集数按钮】
        btn_frame = ttk.Frame(frame_ep)
        btn_frame.pack(fill="x", pady=(0, 5))
        self.btn_load_eps = ttk.Button(btn_frame, text="扫描工作表与集数", command=self.load_episodes)
        self.btn_load_eps.pack(side=tk.LEFT)

        # 【新增：手动输入集数】
        ttk.Label(frame_ep, text="手动输入集数 (多集以半角逗号分隔, 如 23_0001,25_0003):").pack(anchor="w", pady=(0, 2))
        self.manual_ep_var = tk.StringVar()
        self.manual_ep_entry = ttk.Entry(frame_ep, textvariable=self.manual_ep_var, width=60)
        self.manual_ep_entry.pack(fill="x", pady=(0, 5))

        ttk.Label(frame_ep, text="或从下方列表中选择 (与手动输入互斥):").pack(anchor="w")
        
        ep_list_frame = ttk.Frame(frame_ep)
        ep_list_frame.pack(fill="both", expand=True)
        self.ep_listbox = tk.Listbox(ep_list_frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        self.ep_listbox.pack(side=tk.LEFT, fill="both", expand=True)
        ep_scrollbar = ttk.Scrollbar(ep_list_frame, orient="vertical", command=self.ep_listbox.yview)
        ep_scrollbar.pack(side=tk.RIGHT, fill="y")
        self.ep_listbox.config(yscrollcommand=ep_scrollbar.set)
        
        # 【输入与列表互斥的事件绑定】
        def on_manual_ep_change(*args):
            if self.manual_ep_var.get().strip():
                self.ep_listbox.selection_clear(0, tk.END)
                self.ep_listbox.config(state="disabled")
            else:
                self.ep_listbox.config(state="normal")
                
        self.manual_ep_var.trace_add("write", on_manual_ep_change)
        
        def on_listbox_select(event):
            if self.ep_listbox.curselection():
                self.manual_ep_var.set("")
                
        self.ep_listbox.bind("<<ListboxSelect>>", on_listbox_select)

        # --- 5. AI 参数配置区 ---
        frame_ai = ttk.LabelFrame(self.scrollable_frame, text="5. AI 引擎及要求设置", padding=10)
        frame_ai.pack(fill="x", padx=10, pady=5)
        frame_ai.columnconfigure(3, weight=1)
        
        # Row 0: 任务模式 & 系统角色 (新增)
        ttk.Label(frame_ai, text="任务模式:").grid(row=0, column=0, sticky="w", pady=5)
        self.task_mode_box = ttk.Combobox(frame_ai, values=["拼写检查 (LQA)", "纯翻译 (Translation)"], width=20, state="readonly")
        self.task_mode_box.current(0)
        self.task_mode_box.grid(row=0, column=1, padx=5, sticky="w", pady=5)
        
        ttk.Label(frame_ai, text="系统角色(支持手填):").grid(row=0, column=2, sticky="w", padx=(10,0), pady=5)
        roles = [
            "You are an expert in subtitle localization and Language Quality Assurance (LQA).",
            "You are a professional native translator with extensive experience in subtitle translation.",
            "You are a creative subtitle translator, skilled at adapting slang, idioms, and cultural nuances.",
            "You are a technical document translator, focusing on precision, strict terminology, and consistency.",
            "You are a colloquial dialogue translator, making translations sound natural, spoken, and character-driven."
        ]
        self.role_box = ttk.Combobox(frame_ai, values=roles, width=50)
        self.role_box.current(0)
        self.role_box.grid(row=0, column=3, padx=5, sticky="we", pady=5)

        # Row 1: 模型 & 目标语言
        ttk.Label(frame_ai, text="选择模型:").grid(row=1, column=0, sticky="w")
        self.model_box = ttk.Combobox(frame_ai, values=list(ENGINES_MAP.keys()), width=20)
        self.model_box.current(0)
        self.model_box.grid(row=1, column=1, padx=5, sticky="w")

        ttk.Label(frame_ai, text="目标语言(支持手填):").grid(row=1, column=2, sticky="w", padx=(10,0))
        self.lang_box = ttk.Combobox(frame_ai, values=list(LANGUAGES_MAP.keys()), width=25)
        self.lang_box.set("English (United States)")
        self.lang_box.grid(row=1, column=3, padx=5, sticky="w")

        # Row 2: Token 上限
        ttk.Label(frame_ai, text="Token 上限/次:").grid(row=2, column=0, sticky="w", pady=5)
        self.token_limit = ttk.Entry(frame_ai, width=10)
        self.token_limit.insert(0, "2000")
        self.token_limit.grid(row=2, column=1, padx=5, sticky="w", pady=5)

        # Row 3: 额外背景
        ttk.Label(frame_ai, text="要求/背景:").grid(row=3, column=0, sticky="nw", pady=5)
        self.context_text = tk.Text(frame_ai, width=65, height=3)
        self.context_text.insert("1.0", "这通常是短剧的字幕文件，xxx背景。请确保称呼和语气符合设定，按要求检查并修改，使其符合本地化要求，需要保留原始译文使用的人名和地名，数字、货币格式，专有名词（以及术语）。")
        self.context_text.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="we")

        # --- 新增：6. 术语表约束配置 (可选) ---
        frame_tb = ttk.LabelFrame(self.scrollable_frame, text="6. 术语表约束配置 (向AI发送术语表，防止错误修改)", padding=10)
        frame_tb.pack(fill="x", padx=10, pady=5)
        
        self.var_use_tb = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_tb, text="启用术语约束 (勾选后会将下列文件中的术语发给AI)", variable=self.var_use_tb).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 5))
        
        ttk.Label(frame_tb, text="术语表文件:").grid(row=1, column=0, sticky="w")
        self.tb_file_path = tk.StringVar()
        ttk.Entry(frame_tb, textvariable=self.tb_file_path, width=50).grid(row=1, column=1, padx=5, sticky="w")
        ttk.Button(frame_tb, text="浏览...", command=self.browse_tb_file).grid(row=1, column=2)

        tb_col_frame = ttk.Frame(frame_tb)
        tb_col_frame.grid(row=2, column=0, columnspan=3, sticky="w", pady=5)
        
        ttk.Label(tb_col_frame, text="原文列名:").pack(side="left")
        self.tb_col_src = tk.StringVar(value="Source")
        ttk.Entry(tb_col_frame, textvariable=self.tb_col_src, width=10).pack(side="left", padx=(5, 15))
        
        ttk.Label(tb_col_frame, text="译文列名:").pack(side="left")
        self.tb_col_tgt = tk.StringVar(value="Target")
        ttk.Entry(tb_col_frame, textvariable=self.tb_col_tgt, width=10).pack(side="left", padx=(5, 15))

        ttk.Label(tb_col_frame, text="Type列名:").pack(side="left")
        self.tb_col_type = tk.StringVar(value="Type")
        ttk.Entry(tb_col_frame, textvariable=self.tb_col_type, width=10).pack(side="left", padx=(5, 0))
        # --------------------------------------

        # --- 6. 操作与日志区 ---
        frame_action = tk.Frame(self.scrollable_frame)
        frame_action.pack(fill="both", expand=True, padx=10, pady=5)

        # === 新增：实时统计信息 ===
        self.var_stats = tk.StringVar(value="准备就绪")
        lbl_stats = ttk.Label(frame_action, textvariable=self.var_stats, font=("Arial", 10, "bold"))
        lbl_stats.pack(pady=(5, 0))
        # ==========================

        btn_container = tk.Frame(frame_action)
        btn_container.pack(pady=5)
        
        self.btn_start = ttk.Button(btn_container, text="🚀 开始检查", command=self.start_processing)
        self.btn_start.pack(side="left", padx=10)

        self.btn_stop = ttk.Button(btn_container, text="🛑 停止检查", command=self.stop_processing, state="disabled")
        self.btn_stop.pack(side="left", padx=10)

        # 将 height=13 修改为你想要的高度，比如 20 或 25
        self.log_area = scrolledtext.ScrolledText(frame_action, width=85, height=30, state='disabled')
        self.log_area.pack(fill="both", expand=True)
        self.load_config()
    
    def save_config(self):
        """加密并保存 API 配置到软件所在目录"""
        data = {
            "endpoint": self.api_endpoint.get().strip(),
            "api_key": self.api_key.get().strip()
        }
        if not data["endpoint"] or not data["api_key"]:
            messagebox.showwarning("警告", "Endpoint 和 API Key 不能为空！")
            return
            
        try:
            enc_data = encrypt_data(data)
            # 使用新的全局统一路径
            with open(LQA_CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(enc_data)
            messagebox.showinfo("成功", f"API 配置已加密保存至:\n{LQA_CONFIG_FILE}")
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败 (请检查目录是否有写入权限):\n{e}")

    def manual_load_config(self):
        filepath = filedialog.askopenfilename(
            title="选择 API 配置文件",
            filetypes=[("Encrypted Config", "*.enc"), ("All Files", "*.*")],
            initialdir=CONFIG_DIR  # 修改为全局统一配置目录
        )
        if filepath:
            self.load_config(filepath)

    def load_config(self, filepath=LQA_CONFIG_FILE):
        if os.path.exists(filepath):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    enc_data = f.read()
                data = decrypt_data(enc_data)
                self.api_endpoint.set(data.get("endpoint", ""))
                self.api_key.set(data.get("api_key", ""))
                
                # 如果是手动加载的，给个成功提示
                if filepath != LQA_CONFIG_FILE:
                    messagebox.showinfo("成功", "配置文件加载成功！")
            except Exception as e:
                if filepath != LQA_CONFIG_FILE:
                    messagebox.showerror("错误", f"配置文件解析失败或已损坏:\n{e}")
                else:
                    print(f"自动加载配置失败: {e}")

    def browse_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.file_path.set(filepath)
            default_output = filepath.replace(".xlsx", "_CheckedResult.xlsx")
            self.output_path.set(default_output)

    def browse_output_file(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="选择保存前缀位置"
        )
        if filepath:
            self.output_path.set(filepath)

    def log(self, message):
        self.parent.after(0, self._append_log, message)

    def _append_log(self, message):
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')

    def load_sheets(self):
        """扫描 Excel 中的所有 Sheet 并加载到列表"""
        input_file = self.file_path.get()
        if not input_file:
            messagebox.showerror("错误", "请先在上方选择输入 Excel 文件！")
            return
        try:
            self.log("正在扫描工作表 (Sheet)...")
            self.parent.update()  # 替换 self.root.update()
            
            wb = openpyxl.load_workbook(input_file, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            
            self.sheet_listbox.delete(0, tk.END)
            for s in sheets:
                self.sheet_listbox.insert(tk.END, s)
                
            self.log(f"✅ 成功扫描到 {len(sheets)} 个工作表。")
        except Exception as e:
            self.log(f"❌ 扫描工作表失败: {e}")
            messagebox.showerror("错误", f"读取工作表失败：\n{e}")

    def load_episodes(self):
        """扫描选中的 Sheet 中的集数（带去重与自然排序）"""
        input_file = self.file_path.get()
        if not input_file:
            messagebox.showerror("错误", "请先在上方选择输入 Excel 文件！")
            return
            
        try:
            c_ep = int(self.col_ep.get())
            r_start = int(self.row_start.get())
            
            self.log("正在扫描集数，请稍候...")
            self.parent.update()  # 替换 self.root.update()
            
            wb = openpyxl.load_workbook(input_file, data_only=True)
            
            # --- 核心：只在选中的 Sheet 中扫描集数 ---
            selected_sheet_indices = self.sheet_listbox.curselection()
            if selected_sheet_indices:
                target_sheets = [self.sheet_listbox.get(i) for i in selected_sheet_indices]
            else:
                target_sheets = [wb.sheetnames[0]] # 默认第一个
            
            episodes = set()
            for sheet_name in target_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(r_start, ws.max_row + 1):
                        ep_val = str(ws.cell(row=row, column=c_ep).value or "未分类集数").strip()
                        if ep_val:
                            episodes.add(ep_val)
            wb.close()
            
            def natural_sort_key(s):
                return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]
                
            sorted_eps = sorted(list(episodes), key=natural_sort_key)
            
            self.ep_listbox.delete(0, tk.END)
            for ep in sorted_eps:
                self.ep_listbox.insert(tk.END, ep)
                
            self.log(f"✅ 成功从目标 Sheet 中扫描了 {len(sorted_eps)} 个独立集数。")
            
        except Exception as e:
            self.log(f"❌ 扫描集数失败: {e}")
            messagebox.showerror("错误", f"读取集数失败，请检查格式或列号是否正确：\n{e}")

    def start_processing(self):
        if not self.file_path.get() or not self.output_path.get():
            messagebox.showerror("错误", "请先选择输入文件并设置输出路径！")
            return
        
        self.stop_flag = False
        self.total_tokens_used = 0 
        self.var_stats.set("正在初始化任务...") # <--- 新增状态更新
        self.btn_start.config(state="disabled")
        self.btn_load_eps.config(state="disabled")
        self.btn_load_sheets.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.log_area.config(state='normal')
        self.log_area.delete(1.0, tk.END)
        self.log_area.config(state='disabled')
        
        threading.Thread(target=self.process_excel_worker, daemon=True).start()

    def stop_processing(self):
        self.stop_flag = True
        self.btn_stop.config(state="disabled")
        self.log("⚠️ 收到停止指令！当前批次请求完成后将安全退出...")

    def estimate_tokens(self, text):
        try:
            import tiktoken
            encoding = tiktoken.get_encoding('cl100k_base')
            return len(encoding.encode(text))
        except ImportError:
            return int(len(str(text)) * 1.5)

    def build_prompt(self, target_lang_name, target_lang_code, additional_context, with_src, task_mode, system_role):
        if target_lang_code in NO_SPACE_LANGS:
            concat_rule = f"Since {target_lang_name} does not use spaces between words, if a sentence spans across multiple lines (IDs), treat them as directly connected without any spaces when evaluating context and grammar."
        else:
            concat_rule = f"Since {target_lang_name} uses spaces between words, if a sentence spans across multiple lines (IDs), treat them as connected with a space when evaluating context and grammar."

        if "翻译" in task_mode or "Translation" in task_mode:
            sys_prompt = f"""# System Role:
{system_role}
# Context:
- Target Language: {target_lang_name} ({target_lang_code})
- Continues Rule: {concat_rule}
- Reqs: {additional_context}
# JSON Fields:
i: id
s: Source text
r: Translated text (Final output)
# Format:
In: [{{"i":"1","s":"Hello"}}]
Out: {{"result":[{{"i":"1","r":"你好"}}]}}
# Task:
Translate `s` to the Target Language accurately and naturally based on the Context. Return ONLY valid minified JSON.
"""
        else:
            if with_src:
                sys_prompt = f"""# System Role:
{system_role}
# Context:
- Target Language: {target_lang_name} ({target_lang_code})
- Continues Rule: {concat_rule}
- Reqs: {additional_context}
# JSON Fields:
i: id
s: Source text (Original)
t: Translation (To be reviewed)
r: Revised translation (Final localized output)
# Format:
In: [{{"i":"1","s":"Hello","t":"哈楼"}}]
Out: {{"result":[{{"i":"1","r":"你好"}}]}}
# Task:
STRICT BILINGUAL REVIEW: Deeply compare `t` against `s`. Fix ALL mistranslations, omissions, unidiomatic expressions, spelling, and grammar errors in `t` to ensure accurate localization. Return ONLY valid minified JSON.
"""
            else:
                sys_prompt = f"""# System Role:
{system_role}
# Context:
- Target Language: {target_lang_name} ({target_lang_code})
- Continues Rule: {concat_rule}
- Reqs: {additional_context}
# JSON Fields:
i: id
t: Translation (To be reviewed)
r: Revised translation (Fixed output)
# Format:
In: [{{"i":"1","t":"哈楼"}}]
Out: {{"result":[{{"i":"1","r":"哈喽"}}]}}
# Task:
Fix spelling, grammar, unidiomatic expressions in `t` to ensure accurate localization. Return ONLY valid minified JSON.
"""
        return sys_prompt
    
    def get_rich_text_diff(self, old_text, new_text):
        # 如果没有变动，直接返回两个纯文本
        if old_text == new_text:
            return old_text, new_text

        blue_font = InlineFont(color="0000FF") # 原始译文的变动用蓝色
        red_font = InlineFont(color="FF0000")  # 修改后的变动用红色
        
        # 【核心修复1】：显式定义黑色无加粗字体，彻底切断 Excel 颜色溢出
        default_font = InlineFont(color="000000", b=False) 
        
        rich_old = CellRichText()
        rich_new = CellRichText()
        
        matcher = difflib.SequenceMatcher(None, old_text, new_text)
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                # 不再直接 append 字符串，而是用 default_font 锁死黑色
                rich_old.append(TextBlock(font=default_font, text=old_text[i1:i2]))
                rich_new.append(TextBlock(font=default_font, text=new_text[j1:j2]))
            elif tag == 'insert':
                rich_new.append(TextBlock(font=red_font, text=new_text[j1:j2]))
            elif tag == 'delete':
                rich_old.append(TextBlock(font=blue_font, text=old_text[i1:i2]))
            elif tag == 'replace':
                rich_old.append(TextBlock(font=blue_font, text=old_text[i1:i2]))
                rich_new.append(TextBlock(font=red_font, text=new_text[j1:j2]))
                
        return rich_old, rich_new
    
    def process_excel_worker(self):
        try:
            import time
            import threading
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            input_file = self.file_path.get()
            output_file_base = self.output_path.get()
            
            # 安全提取列号，允许留空（翻译模式下 c_tgt 可为空）
            c_src = int(self.col_src.get()) if self.col_src.get().strip() else 0
            c_tgt = int(self.col_tgt.get()) if self.col_tgt.get().strip() else 0
            c_ep  = int(self.col_ep.get()) if self.col_ep.get().strip() else 0
            c_res = int(self.col_res.get()) if self.col_res.get().strip() else 0
            r_start = int(self.row_start.get())
            
            with_src = self.var_with_source.get()
            t_limit = int(self.token_limit.get())
            additional_context = self.context_text.get("1.0", tk.END).strip()
            output_mode = self.var_output_mode.get()

            ui_lang_name = self.lang_box.get()
            target_lang_code = LANGUAGES_MAP.get(ui_lang_name, ui_lang_name)
            
            ui_model_name = self.model_box.get()
            actual_deployment_name = ENGINES_MAP.get(ui_model_name, ui_model_name)

            # --- 新增：获取任务模式与角色 ---
            task_mode = self.task_mode_box.get()
            system_role = self.role_box.get().strip()
            if not system_role:
                system_role = "You are an expert in subtitle localization."
            is_translation_mode = "翻译" in task_mode or "Translation" in task_mode

            # 【核心修改：集数提取优先判断手动输入框】
            manual_eps_str = getattr(self, 'manual_ep_var', tk.StringVar()).get().strip()
            if manual_eps_str:
                selected_episodes = [ep.strip() for ep in manual_eps_str.split(",") if ep.strip()]
            else:
                selected_indices = self.ep_listbox.curselection()
                selected_episodes = [self.ep_listbox.get(i) for i in selected_indices]
            
            selected_sheet_indices = self.sheet_listbox.curselection()
            
            self.log(f"正在加载原始 Excel 文件: {input_file}")
            wb = openpyxl.load_workbook(input_file)
            
            if selected_sheet_indices:
                target_sheets = [self.sheet_listbox.get(i) for i in selected_sheet_indices]
            else:
                target_sheets = [wb.sheetnames[0]]
                
            # === 拦截校验：检查是否配置了 API ===
            api_endpoint_val = self.api_endpoint.get().strip()
            api_key_val = self.api_key.get().strip()
            
            if not api_endpoint_val or not api_key_val:
                self.log("❌ 错误：请求被拦截。请先在界面最上方配置并保存 API 接口信息！")
                self.parent.after(0, lambda: messagebox.showerror("错误", "API Endpoint 和 Key 不能为空，请先配置！"))
                return

            client = AzureOpenAI(
                azure_endpoint=api_endpoint_val,
                api_key=api_key_val,
                api_version=DEFAULT_API_VERSION
            )

            # ================= 新增：解析术语表文件 =================
            glossary_str = ""
            if self.var_use_tb.get():
                tb_path = self.tb_file_path.get().strip()
                tb_scol = self.tb_col_src.get().strip()
                tb_tcol = self.tb_col_tgt.get().strip()
                tb_typecol = self.tb_col_type.get().strip()
                
                if not tb_path or not os.path.exists(tb_path):
                    self.log("❌ 错误：启用了术语表约束，但文件路径无效！")
                    self.root.after(0, lambda: messagebox.showerror("错误", "术语表文件不存在！"))
                    return
                    
                self.log(f"正在加载并提取术语表: {os.path.basename(tb_path)}")
                term_list = []
                try:
                    import csv
                    if tb_path.lower().endswith('.csv'):
                        with open(tb_path, 'r', encoding='utf-8-sig') as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                s = row.get(tb_scol, "").strip() if row.get(tb_scol) else ""
                                t = row.get(tb_tcol, "").strip() if row.get(tb_tcol) else ""
                                typ = row.get(tb_typecol, "").strip() if row.get(tb_typecol) else ""
                                if s and t:
                                    term_list.append(f"- {s}: {t} ({typ})" if typ else f"- {s}: {t}")
                    else:
                        wb_tb = openpyxl.load_workbook(tb_path, data_only=True)
                        ws_tb = wb_tb.active
                        headers = [str(cell.value).strip() if cell.value else "" for cell in ws_tb[1]]
                        if tb_scol in headers and tb_tcol in headers:
                            s_idx = headers.index(tb_scol)
                            t_idx = headers.index(tb_tcol)
                            type_idx = headers.index(tb_typecol) if tb_typecol in headers else -1
                            for row in ws_tb.iter_rows(min_row=2, values_only=True):
                                s = str(row[s_idx]).strip() if row[s_idx] else ""
                                t = str(row[t_idx]).strip() if row[t_idx] else ""
                                typ = str(row[type_idx]).strip() if type_idx != -1 and row[type_idx] else ""
                                if s and t and s != 'None' and t != 'None':
                                    term_list.append(f"- {s}: {t} ({typ})" if typ else f"- {s}: {t}")
                        wb_tb.close()
                        
                    if term_list:
                        glossary_str = "\n".join(term_list)
                        self.log(f"✅ 成功提取 {len(term_list)} 条术语用于 AI 强制约束。")
                    else:
                        self.log("⚠️ 术语表提取为空，请检查 CSV/Excel 的列名是否输入正确。")
                except Exception as e:
                    self.log(f"❌ 读取术语表失败: {e}")
                    self.root.after(0, lambda: messagebox.showerror("错误", f"读取术语表失败:\n{e}"))
                    return
            # ========================================================

            # --- 传入新的模式和角色参数 ---
            sys_prompt = self.build_prompt(ui_lang_name, target_lang_code, additional_context, with_src, task_mode, system_role)

            # ---------------- 注入：术语检查前置准备 ----------------
            enable_term_check = self.enable_term_check_var.get()
            selected_indices = self.type_listbox.curselection()
            active_types = [self.type_listbox.get(i) for i in selected_indices]
            ignore_case = self.ignore_case_var.get()
            
            term_list = []
            
            # 只有当总开关勾选时，才执行术语表的读取和加载
            if enable_term_check:
                term_file = self.term_file_path.get()
                if term_file and os.path.exists(term_file):
                    try:
                        from openpyxl.utils import column_index_from_string
                        src_idx = column_index_from_string(self.term_src_col.get().strip().upper()) - 1
                        tgt_idx = column_index_from_string(self.term_tgt_col.get().strip().upper()) - 1
                        type_idx = column_index_from_string(self.term_type_col.get().strip().upper()) - 1
                        
                        term_df = pd.read_excel(term_file)
                        for _, tr in term_df.iterrows():
                            row_type = str(tr.iloc[type_idx]).strip() if type_idx < len(tr) else ""
                            is_partial = (row_type in active_types) if active_types else False
                                
                            t_src = str(tr.iloc[src_idx]).strip() if src_idx < len(tr) else ""
                            t_tgt = str(tr.iloc[tgt_idx]).strip() if tgt_idx < len(tr) else ""
                            if t_src and t_tgt and t_src != 'nan' and t_tgt != 'nan':
                                term_list.append({
                                    'src': t_src, 'tgt': t_tgt, 'is_partial': is_partial
                                })
                                
                        self.log(f"成功加载 {len(term_list)} 条术语用于检查。")
                    except Exception as e:
                        self.log(f"术语表加载异常，将跳过术语检查: {e}")

            from openpyxl.utils import column_index_from_string
            term_out_col_letter = self.term_output_col.get().strip().upper()
            try:
                term_out_col_idx = column_index_from_string(term_out_col_letter)
            except:
                term_out_col_idx = c_res + 1
            # --------------------------------------------------------

            # === 1. 预扫描：构建任务池并计算总集数 ===
            all_tasks = []
            
            if output_mode == "merged":
                for s_name in target_sheets:
                    if s_name in wb.sheetnames:
                        wb[s_name].cell(row=r_start-1, column=c_res, value="Spell Check Result (AI)")

            for sheet_name in target_sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                
                headers = []
                for r in range(1, r_start):
                    headers.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

                episodes_data = {}
                for row in range(r_start, ws.max_row + 1):
                    # 如果列号大于 0 且填了，才去读取，否则直接当做空字符串
                    src_text_raw = ws.cell(row=row, column=c_src).value if c_src > 0 else ""
                    tgt_text_raw = ws.cell(row=row, column=c_tgt).value if c_tgt > 0 else ""
                    
                    src_text_str = str(src_text_raw).strip() if src_text_raw else ""
                    tgt_text_str = str(tgt_text_raw).strip() if tgt_text_raw else ""
                    
                    # --- 核心修改：翻译模式根据原文判断，拼写模式根据译文判断 ---
                    if is_translation_mode:
                        if not src_text_str:
                            continue
                    else:
                        if not tgt_text_str:
                            continue
                    
                    ep_val = str(ws.cell(row=row, column=c_ep).value or "未分类集数").strip()
                    original_row_values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
                    
                    # === 调整 JSON 字段拼装顺序 ===
                    item = {"i": str(row)}
                    if is_translation_mode:
                        item["s"] = src_text_str
                    else:
                        if with_src:
                            item["s"] = src_text_str
                        item["t"] = tgt_text_str
                    
                    if ep_val not in episodes_data:
                        episodes_data[ep_val] = []
                    
                    episodes_data[ep_val].append({
                        "id": str(row),
                        "item": item,
                        "original_row_values": original_row_values
                    })
                    
                for ep_name, rows_in_ep in episodes_data.items():
                    if selected_episodes and ep_name not in selected_episodes:
                        continue
                    all_tasks.append({
                        "sheet_name": sheet_name,
                        "ep_name": ep_name,
                        "rows_in_ep": rows_in_ep,
                        "headers": headers,
                        "ws": ws
                    })

            total_eps = len(all_tasks)
            success_count = 0
            fail_count = 0
            error_logs = []
            
            if total_eps == 0:
                self.log("⚠️ 没有找到需要处理的集数！")
                self.parent.after(0, lambda: self.btn_start.config(state="normal"))
                return
            
            # ================= 核心修改：多线程调度设置 =================
            excel_lock = threading.Lock()
            enable_mt = self.use_multithread_var.get()
            max_threads = self.thread_count_var.get() if enable_mt else 1
            max_retries = self.retry_count_var.get() # <--- 新增：读取设置的重试次数

            def process_episode(current_idx, task):
                nonlocal success_count, fail_count, error_logs

                # 这里原来是 break，改为 return，结束当前集数的处理
                if self.stop_flag:
                    return
                    
                sheet_name = task["sheet_name"]
                ep_name = task["ep_name"]
                rows_in_ep = task["rows_in_ep"]
                headers = task["headers"]
                ws = task["ws"]

                self.parent.after(0, lambda c=current_idx, t=total_eps, s=success_count, f=fail_count: 
                                self.var_stats.set(f"正在处理第{c}集，共{t}集，成功{s}集，失败{f}集"))
                self.log(f"\n====== 处理中: Sheet [{sheet_name}] -> 集数 [{ep_name}] (共 {len(rows_in_ep)} 行) ======")

                current_batch = []
                current_tokens = 0
                episode_results = []
                has_error = False
                error_msg = ""

                # 发送请求 (不写入，多线程并发执行)
                for idx, item_data in enumerate(rows_in_ep):
                    row_item = item_data["item"]
                    item_json_str = json.dumps(row_item, ensure_ascii=False)
                    item_tokens = self.estimate_tokens(item_json_str)
                    
                    # 触发条件 1：Token超限，需要发送当前批次
                    if current_tokens + item_tokens > t_limit and current_batch:
                        for attempt in range(max_retries + 1):
                            try:
                                res = self._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                                episode_results.extend(res)
                                break  # 成功则跳出重试循环
                            except Exception as e:
                                if attempt < max_retries:
                                    self.log(f"  ⚠️ 请求失败，等待2秒后进行第 {attempt + 1}/{max_retries} 次重试... ({str(e)})")
                                    time.sleep(2) # 稍微暂停，缓解 API 频率限制
                                else:
                                    has_error = True
                                    error_msg = str(e)
                                    break
                                    
                        if has_error:
                            break  # 彻底失败，中止内部循环跳到失败拦截
                            
                        current_batch = []
                        current_tokens = 0
                        
                    current_batch.append(row_item)
                    current_tokens += item_tokens
                    
                    # 触发条件 2：到达最后一行的收尾批次
                    if idx == len(rows_in_ep) - 1 and current_batch:
                        for attempt in range(max_retries + 1):
                            try:
                                res = self._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                                episode_results.extend(res)
                                break  # 成功则跳出重试循环
                            except Exception as e:
                                if attempt < max_retries:
                                    self.log(f"  ⚠️ 请求失败，等待2秒后进行第 {attempt + 1}/{max_retries} 次重试... ({str(e)})")
                                    time.sleep(2)
                                else:
                                    has_error = True
                                    error_msg = str(e)
                                    break
                                    
                        if has_error:
                            break  # 彻底失败，中止内部循环跳到失败拦截

                # 失败拦截与降级处理
                if has_error:
                    with excel_lock:
                        fail_count += 1
                        error_logs.append([os.path.basename(input_file), f"[{sheet_name}] {ep_name}", error_msg])
                    self.log(f"  ❌ 【检查失败拦截】AI 请求失败，将执行降级保存(仅原样输出+术语标记)。错误原因: {error_msg}")
                    # 【核心修复】：取消了 return。让它继续往下走，走到 Excel 写入逻辑里去！

                # ================= 无论成功失败，排队获取写入锁，处理 Excel 写入 =================
                with excel_lock:
                    if not has_error:
                        success_count += 1  # 只有真正没报错的，才计入成功集数
                        
                    self.parent.after(0, lambda c=current_idx, t=total_eps, s=success_count, f=fail_count: 
                                    self.var_stats.set(f"正在处理第{c}集，共{t}集，成功{s}集，失败{f}集"))

                    res_mapping = {str(r.get("i", r.get("id"))): r.get("r", r.get("revisedTranslation", "")) for r in episode_results}

                    try:
                        if output_mode == "split":
                            ep_wb = openpyxl.Workbook()
                            ep_ws = ep_wb.active
                            ep_ws.title = "LQA Result"
                            
                            for h_row_idx, h_row_vals in enumerate(headers, start=1):
                                for col_idx, val in enumerate(h_row_vals, start=1):
                                    ep_ws.cell(row=h_row_idx, column=col_idx, value=val)
                                ep_ws.cell(row=h_row_idx, column=c_res, value="Spell Check Result (AI)")

                            current_new_row_idx = r_start
                            for item_data in rows_in_ep:
                                # 1. 复制原表该行的所有基础数据（纯文本）
                                for col_idx, val in enumerate(item_data["original_row_values"], start=1):
                                    ep_ws.cell(row=current_new_row_idx, column=col_idx, value=val)

                                row_id = item_data["id"]
                                
                                # 2. 提取原文、原译文、AI修改后译文
                                try:
                                    src_text = str(item_data["original_row_values"][c_src-1])
                                except:
                                    src_text = " ".join([str(x) for x in item_data["original_row_values"] if x])
                                    
                                # --- 新增：兼容“纯翻译”与“拼写检查”模式 ---
                                if is_translation_mode:
                                    # 严谨判断 c_tgt > 0，防止出现 -1 获取到列表最后一个元素的 Bug
                                    old_text = str(item_data["original_row_values"][c_tgt-1]) if c_tgt > 0 and c_tgt-1 < len(item_data["original_row_values"]) else ""
                                    new_text = res_mapping.get(row_id, "") # 没翻译出来就为空
                                    rich_old = str(old_text)
                                    rich_new = str(new_text)
                                else:
                                    old_text = item_data["item"].get("t", "")
                                    new_text = res_mapping.get(row_id, old_text)
                                    rich_old, rich_new = self.get_rich_text_diff(old_text, new_text)
                                    
                                rich_src = src_text

                                found_terms_display = []
                                if term_list:
                                    matched_src_terms = set()
                                    matched_old_terms = set()
                                    matched_new_terms = set()

                                    def check_match(term_str, target_text, partial):
                                        if partial:
                                            parts = term_str.split()
                                            if not parts: return False, []
                                            for p in parts:
                                                if ignore_case:
                                                    if p.lower() not in target_text.lower(): return False, []
                                                else:
                                                    if p not in target_text: return False, []
                                            return True, parts
                                        else:
                                            if ignore_case:
                                                match = term_str.lower() in target_text.lower()
                                            else:
                                                match = term_str in target_text
                                            return match, [term_str] if match else []

                                    for term_obj in term_list:
                                        t_src, t_tgt, is_partial = term_obj['src'], term_obj['tgt'], term_obj['is_partial']

                                        # 检查 1: 原文
                                        is_match_src, highlight_src = check_match(t_src, src_text, is_partial)
                                        if is_match_src:
                                            matched_src_terms.update(highlight_src)
                                            found_terms_display.append(f"原:{t_src}{'(部分)' if is_partial else ''}")

                                        # 检查 2: 原始译文
                                        is_match_old, highlight_old = check_match(t_tgt, old_text, is_partial)
                                        if is_match_old:
                                            matched_old_terms.update(highlight_old)
                                            found_terms_display.append(f"译:{t_tgt}{'(部分)' if is_partial else ''}")
                                            
                                        # 检查 3: AI修改后的新译文
                                        is_match_new, highlight_new = check_match(t_tgt, new_text, is_partial)
                                        if is_match_new:
                                            matched_new_terms.update(highlight_new)
                                            found_terms_display.append(f"改:{t_tgt}{'(部分)' if is_partial else ''}")

                                    # 4. 执行绿色高亮覆盖 (各自用各自的匹配词去标绿，不再错位)
                                    if matched_src_terms:
                                        rich_src = self.apply_term_rich_text(rich_src, list(matched_src_terms), ignore_case)
                                    if matched_old_terms:
                                        rich_old = self.apply_term_rich_text(rich_old, list(matched_old_terms), ignore_case)
                                    if matched_new_terms:
                                        rich_new = self.apply_term_rich_text(rich_new, list(matched_new_terms), ignore_case)
                                
                                # 5. 覆盖写入列，增加 > 0 判断防止 openpyxl 报第0列错误
                                if c_src > 0:
                                    ep_ws.cell(row=current_new_row_idx, column=c_src).value = rich_src
                                if c_tgt > 0:
                                    ep_ws.cell(row=current_new_row_idx, column=c_tgt).value = rich_old
                                    
                                ep_ws.cell(row=current_new_row_idx, column=c_res).value = rich_new
                                
                                # 6. 写入展示列信息 (去重并保持顺序)
                                if found_terms_display:
                                    unique_display = list(dict.fromkeys(found_terms_display))
                                    ep_ws.cell(row=current_new_row_idx, column=term_out_col_idx).value = " | ".join(unique_display)
                                    
                                ep_ws.cell(row=1, column=term_out_col_idx).value = "术语匹配结果"
                                
                                current_new_row_idx += 1
                            
                            safe_sheet_name = str(sheet_name).replace("/", "_").replace("\\", "_")
                            safe_ep_name = str(ep_name).replace("/", "_").replace("\\", "_")
                            out_dir = os.path.dirname(output_file_base)
                            out_name = os.path.basename(output_file_base).replace(".xlsx", "")
                            ep_output_file = os.path.join(out_dir, f"{out_name}_{safe_sheet_name}_{safe_ep_name}.xlsx")
                            ep_wb.save(ep_output_file)
                            self.log(f"  💾 【分发保存】{sheet_name} - {ep_name} 已成功保存。")
                        else:
                            for item_data in rows_in_ep:
                                row_id = item_data["id"]
                                mapping_row = int(row_id)
                                
                                # 1. 提取原文、原译文、AI修改后译文
                                try:
                                    src_text = str(item_data["original_row_values"][c_src-1])
                                except:
                                    src_text = " ".join([str(x) for x in item_data["original_row_values"] if x])
                                    
                                # --- 新增：兼容“纯翻译”与“拼写检查”模式 ---
                                if is_translation_mode:
                                    # 严谨判断 c_tgt > 0，防止出现 -1 获取到列表最后一个元素的 Bug
                                    old_text = str(item_data["original_row_values"][c_tgt-1]) if c_tgt > 0 and c_tgt-1 < len(item_data["original_row_values"]) else ""
                                    new_text = res_mapping.get(row_id, "") # 没翻译出来就为空
                                    rich_old = str(old_text)
                                    rich_new = str(new_text)
                                else:
                                    old_text = item_data["item"].get("t", "")
                                    new_text = res_mapping.get(row_id, old_text)
                                    rich_old, rich_new = self.get_rich_text_diff(old_text, new_text)
                                    
                                rich_src = src_text

                                found_terms_display = []
                                if term_list:
                                    matched_src_terms = set()
                                    matched_old_terms = set()
                                    matched_new_terms = set()

                                    def check_match(term_str, target_text, partial):
                                        if partial:
                                            parts = term_str.split()
                                            if not parts: return False, []
                                            for p in parts:
                                                if ignore_case:
                                                    if p.lower() not in target_text.lower(): return False, []
                                                else:
                                                    if p not in target_text: return False, []
                                            return True, parts 
                                        else:
                                            if ignore_case:
                                                match = term_str.lower() in target_text.lower()
                                            else:
                                                match = term_str in target_text
                                            return match, [term_str] if match else []

                                    for term_obj in term_list:
                                        t_src, t_tgt, is_partial = term_obj['src'], term_obj['tgt'], term_obj['is_partial']

                                        # 检查 1: 原文
                                        is_match_src, highlight_src = check_match(t_src, src_text, is_partial)
                                        if is_match_src:
                                            matched_src_terms.update(highlight_src)
                                            found_terms_display.append(f"原:{t_src}{'(部分)' if is_partial else ''}")

                                        # 检查 2: 原始译文
                                        is_match_old, highlight_old = check_match(t_tgt, old_text, is_partial)
                                        if is_match_old:
                                            matched_old_terms.update(highlight_old)
                                            found_terms_display.append(f"译:{t_tgt}{'(部分)' if is_partial else ''}")
                                            
                                        # 检查 3: AI修改后的新译文
                                        is_match_new, highlight_new = check_match(t_tgt, new_text, is_partial)
                                        if is_match_new:
                                            matched_new_terms.update(highlight_new)
                                            found_terms_display.append(f"改:{t_tgt}{'(部分)' if is_partial else ''}")

                                    # 3. 执行绿色高亮覆盖 (独立无错位)
                                    if matched_src_terms:
                                        rich_src = self.apply_term_rich_text(rich_src, list(matched_src_terms), ignore_case)
                                    if matched_old_terms:
                                        rich_old = self.apply_term_rich_text(rich_old, list(matched_old_terms), ignore_case)
                                    if matched_new_terms:
                                        rich_new = self.apply_term_rich_text(rich_new, list(matched_new_terms), ignore_case)

                                # 4. 覆盖写入列，增加 > 0 判断防止 openpyxl 报第0列错误
                                if c_src > 0:
                                    ws.cell(row=mapping_row, column=c_src).value = rich_src
                                if c_tgt > 0:
                                    ws.cell(row=mapping_row, column=c_tgt).value = rich_old
                                    
                                ws.cell(row=mapping_row, column=c_res).value = rich_new
                                
                                # 5. 写入展示列信息
                                if found_terms_display:
                                    unique_display = list(dict.fromkeys(found_terms_display))
                                    ws.cell(row=mapping_row, column=term_out_col_idx).value = " | ".join(unique_display)
                                    
                                # 合并模式也需要表头
                                ws.cell(row=r_start-1, column=term_out_col_idx).value = "术语匹配结果"
                                
                            wb.save(output_file_base)
                            self.log(f"  💾 【合并进度追加】{sheet_name} - {ep_name} 已安全追加至原文件。")
                    except Exception as e:
                        self.log(f"  ⚠️ 写入或保存时出错: {e}")

            # ================= 启动线程池或单线程运行 =================
            if enable_mt:
                self.log(f"\n🚀 已开启多线程并发，分配线程数: {max_threads}")
                with ThreadPoolExecutor(max_workers=max_threads) as executor:
                    futures = [executor.submit(process_episode, idx, t) for idx, t in enumerate(all_tasks, start=1)]
                    for future in as_completed(futures):
                        try:
                            future.result()
                        except Exception as e:
                            self.log(f"❌ 线程执行崩溃: {e}")
            else:
                for idx, t in enumerate(all_tasks, start=1):
                    process_episode(idx, t)

            # === 3. 生成错误报告 ===
            if error_logs:
                try:
                    err_wb = openpyxl.Workbook()
                    err_ws = err_wb.active
                    err_ws.append(["文件名", "集数", "错误信息"])
                    for log_entry in error_logs:
                        err_ws.append(log_entry)
                    
                    out_dir = os.path.dirname(output_file_base)
                    err_file = os.path.join(out_dir, f"LQA_Error_Report_{int(time.time())}.xlsx")
                    err_wb.save(err_file)
                    self.log(f"\n⚠️ 发现 {fail_count} 个请求失败的集数，错误报告已生成: {err_file}")
                except Exception as e:
                    self.log(f"\n⚠️ 错误报告生成失败: {e}")

            # ================= 循环处理结束提示 =================
            if self.stop_flag:
                self.log(f"\n====== 🛑 检查已被手动终止 ======")
                self.log(f"💰 本次共计消耗 Token: {self.total_tokens_used}")
                self.parent.after(0, lambda: messagebox.showinfo("已终止", f"检查任务已终止。\n共消耗 Token: {self.total_tokens_used}\n成功: {success_count} 集, 失败: {fail_count} 集。"))
            else:
                self.log(f"\n====== ✨ 设定任务处理完成！ ======")
                self.log(f"💰 本次共计消耗 Token: {self.total_tokens_used}")
                self.parent.after(0, lambda: messagebox.showinfo("完成", f"所选任务处理完成！\n共消耗 Token: {self.total_tokens_used}\n成功: {success_count} 集, 失败: {fail_count} 集。"))

        except Exception as e:
            self.log(f"❌ 发生致命错误: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.parent.after(0, lambda: self.btn_start.config(state="normal"))
            self.parent.after(0, lambda: self.btn_load_eps.config(state="normal"))
            self.parent.after(0, lambda: self.btn_load_sheets.config(state="normal"))
            self.parent.after(0, lambda: self.btn_stop.config(state="disabled"))
            
    def _send_batch_request(self, client, ui_model_name, actual_deployment_name, sys_prompt, batch_data):
        # 抛出异常由外层捕获拦截
        start_id = batch_data[0].get('i', batch_data[0].get('id'))
        end_id = batch_data[-1].get('i', batch_data[-1].get('id'))
        self.log(f"  -> 发送请求... [引擎: {ui_model_name}] (原表行号: {start_id} 至 {end_id})")
        
        user_prompt = f"# Work Data:\n{json.dumps(batch_data, ensure_ascii=False)}"
        
        target_temperature = 0.3
        json_resp = True
        
        if '4o' not in ui_model_name and '4.1' not in ui_model_name:
            target_temperature = 1.0
            
        if 'o1' in ui_model_name or 'o3' in ui_model_name or 'o4' in ui_model_name:
            json_resp = False

        request_kwargs = {
            "model": actual_deployment_name,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": target_temperature
        }
        
        if json_resp:
            request_kwargs["response_format"] = {"type": "json_object"}
        
        response = client.chat.completions.create(**request_kwargs)
        
        if hasattr(response, 'usage') and response.usage:
            used_tokens = response.usage.total_tokens
            self.total_tokens_used += used_tokens
            self.log(f"  [计量] 本次请求花费: {used_tokens} tokens | 累计花费: {self.total_tokens_used} tokens")
            
        resp_content = response.choices[0].message.content
        
        clean_content = resp_content.strip()
        if clean_content.startswith("```json"):
            clean_content = clean_content[7:]
        elif clean_content.startswith("```"):
            clean_content = clean_content[3:]
        if clean_content.endswith("```"):
            clean_content = clean_content[:-3]
            
        result_data = json.loads(clean_content.strip()).get("result", [])
        self.log(f"  ✅ 成功接收 {len(result_data)} 行数据。")
        
        return result_data
    

# ======= 新增：用于术语检查报告的富文本导出 =======
try:
    import openpyxl
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    RICH_TEXT_SUPPORTED = True
except ImportError:
    RICH_TEXT_SUPPORTED = False
# ==================================================

# ======= 新增：尝试引入视频帧提取与渲染库 =======
try:
    import cv2
    from PIL import Image, ImageTk
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

# ======= 新增：全局 ASS 内存暂存字典 =======
global_ass_memory_cache = {}

def safe_punct_convert(text, mode):
    """
    安全的字幕标点转换引擎（完美避开 ASS/SRT 的控制标签）
    mode: 1 (半转全去空格), 2 (全转半补空格)
    """
    if not text or mode not in (1, 2):
        return text
        
    # 利用正则切割，提取真正的文本部分，完美避开 ASS的 {...} 和 SRT的 <...> 标签
    parts = re.split(r'(\{[^}]*\}|<[^>]*>)', text)
    
    for i in range(0, len(parts), 2):
        txt = parts[i]
        if not txt: continue
        
        if mode == 1:  # ====== 半角 转 全角 ======
            txt = txt.replace('...', '…')
            trans = str.maketrans(",.?!:;()[]", "，。？！：；（）【】")
            txt = txt.translate(trans)
            txt = re.sub(r'"([^"]*)"', r'“\1”', txt)
            txt = re.sub(r"'([^']*)'", r'‘\1’', txt)
            # 删除全角标点后面多余的空格
            txt = re.sub(r'([，。？！：；）】”’…])\s+', r'\1', txt)
            
        elif mode == 2: # ====== 全角 转 半角 ======
            txt = txt.replace('…', '...')
            txt = txt.replace('—', '-').replace('——', '--')
            txt = txt.replace('“', '"').replace('”', '"').replace('‘', "'").replace('’', "'")
            trans = str.maketrans("，。？！：；（）【】", ",.?!:;()[]")
            txt = txt.translate(trans)
            
            # 智能补空格：如果半角标点后面跟着【非空格、非其他标点】的正常字符，补充一个空格
            # 自动排除了省略号、横杠、书名号、前括号和引号等不需要加空格的特例
            txt = re.sub(r'([,\?\!\:\;\)\]])(?=[^\s,\?\!\:\;\"\'\-\.《》\n])', r'\1 ', txt)
            # 单独处理句号（排除三个点...的情况）
            txt = re.sub(r'(?<!\.)\.(?!\.)(?=[^\s,\?\!\:\;\"\'\-\.《》\n])', r'. ', txt)
            
        parts[i] = txt
        
    return "".join(parts)

# ====== 新增：全局复用的滚动标签页构造器 ======
def create_scrollable_tab(notebook, text, padding=10):
    outer_frame = ttk.Frame(notebook)
    notebook.add(outer_frame, text=text)

    canvas = tk.Canvas(outer_frame, highlightthickness=0)
    scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
    inner_frame = ttk.Frame(canvas, padding=padding)

    # 动态更新滚动区域
    inner_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    # 强制内部 Frame 宽度自动拉伸，匹配 Canvas 宽度（保证 sticky="ew" 生效）
    def on_canvas_configure(event):
        canvas.itemconfig(canvas_window, width=event.width)
    canvas.bind("<Configure>", on_canvas_configure)

    canvas_window = canvas.create_window((0, 0), window=inner_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # 跨平台鼠标滚轮事件绑定 (仅当鼠标进入该区域时生效，防止干扰其他组件)
    def _on_mousewheel(event):
        if event.num == 4 or getattr(event, 'delta', 0) > 0:
            canvas.yview_scroll(-1, "units")
        elif event.num == 5 or getattr(event, 'delta', 0) < 0:
            canvas.yview_scroll(1, "units")
            
    def _bind_mouse(event):
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel) # Linux 支持
        canvas.bind_all("<Button-5>", _on_mousewheel) # Linux 支持
        
    def _unbind_mouse(event):
        canvas.unbind_all("<MouseWheel>")
        canvas.unbind_all("<Button-4>")
        canvas.unbind_all("<Button-5>")
        
    canvas.bind("<Enter>", _bind_mouse)
    canvas.bind("<Leave>", _unbind_mouse)

    return inner_frame
# ==========================================================

# ====== 新增：全局复用高级判定条件 UI 构造器与求值器 ======
def build_advanced_condition_ui(parent_widget, in_dir_var, title="判定条件"):
    f_cond = ttk.LabelFrame(parent_widget, text=title, padding=10)
    
    logic_var = tk.IntVar(value=0)
    f_logic = ttk.Frame(f_cond)
    f_logic.pack(fill=tk.X, pady=(0, 5))
    ttk.Label(f_logic, text="多条件组合逻辑:").pack(side=tk.LEFT)
    ttk.Radiobutton(f_logic, text="【与】模式 (选中条件必须同时满足才执行)", variable=logic_var, value=0).pack(side=tk.LEFT, padx=(5, 10))
    ttk.Radiobutton(f_logic, text="【或】模式 (选中条件只需满足任意一项即执行)", variable=logic_var, value=1).pack(side=tk.LEFT)

    c1_var = tk.IntVar(value=0)
    bracket_var = tk.StringVar(value="^\\[")
    f_c1 = ttk.Frame(f_cond)
    f_c1.pack(fill=tk.X, pady=2)
    ttk.Checkbutton(f_c1, text="条件1: 匹配正则表达式:", variable=c1_var).pack(side=tk.LEFT)
    ttk.Entry(f_c1, textvariable=bracket_var, width=15).pack(side=tk.LEFT, padx=5)
    ttk.Label(f_c1, text="(例如 ^\\[ 即查找文本首部存在 [ 的行)", foreground="gray").pack(side=tk.LEFT)
    
    def scan_features():
        d = in_dir_var.get().strip()
        if not d or not os.path.exists(d): return messagebox.showwarning("提示", "请先在上方输入文件夹中选择目录！")
        ass_files = [os.path.join(d, f) for f in os.listdir(d) if f.lower().endswith('.ass')]
        if not ass_files: return messagebox.showwarning("提示", "输入文件夹中未找到 .ass 文件！")
        
        effs, styles = set(), set()
        for filepath in ass_files:
            file_name = os.path.basename(filepath)
            if file_name in global_ass_memory_cache:
                lines = global_ass_memory_cache[file_name].split('\n')
            else:
                try:
                    with open(filepath, 'r', encoding='utf-8-sig') as f: lines = f.read().split('\n')
                except: continue
                
            for line in lines:
                if line.startswith('Dialogue:'):
                    p = line.split(',', 9)
                    if len(p) >= 10:
                        styles.add(p[3].strip())
                        effs.add(p[8].strip())
                        
        lb_effs.delete(0, tk.END)
        for e in sorted(list(effs)): lb_effs.insert(tk.END, e)
        
        lb_styles.delete(0, tk.END)
        for s in sorted(list(styles)): lb_styles.insert(tk.END, s)
        
        messagebox.showinfo("成功", f"扫描完毕！\n共发现 {len(effs)} 种特效说明，{len(styles)} 种样式。")

    ttk.Button(f_c1, text="🔍 扫描输入文件夹的特效与样式", command=scan_features).pack(side=tk.RIGHT, padx=5)

    c2_var = tk.IntVar(value=0)
    f_c2 = ttk.Frame(f_cond)
    f_c2.pack(fill=tk.X, pady=2)
    ttk.Checkbutton(f_c2, text="条件2: 包含在以下选中的【特效说明 Effect】内 (支持按住 Ctrl 多选):", variable=c2_var).pack(anchor="w")
    f_c2_lb = ttk.Frame(f_c2)
    f_c2_lb.pack(fill=tk.X, padx=20, pady=2)
    lb_effs = tk.Listbox(f_c2_lb, selectmode=tk.MULTIPLE, height=3, exportselection=False)
    lb_effs.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    sb_effs = ttk.Scrollbar(f_c2_lb, command=lb_effs.yview)
    sb_effs.pack(side=tk.LEFT, fill=tk.Y)
    lb_effs.config(yscrollcommand=sb_effs.set)

    c3_var = tk.IntVar(value=0)
    f_c3 = ttk.Frame(f_cond)
    f_c3.pack(fill=tk.X, pady=2)
    ttk.Checkbutton(f_c3, text="条件3: 包含在以下选中的【样式名称 Style】内 (支持按住 Ctrl 多选):", variable=c3_var).pack(anchor="w")
    f_c3_lb = ttk.Frame(f_c3)
    f_c3_lb.pack(fill=tk.X, padx=20, pady=2)
    lb_styles = tk.Listbox(f_c3_lb, selectmode=tk.MULTIPLE, height=3, exportselection=False)
    lb_styles.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    sb_styles = ttk.Scrollbar(f_c3_lb, command=lb_styles.yview)
    sb_styles.pack(side=tk.LEFT, fill=tk.Y)
    lb_styles.config(yscrollcommand=sb_styles.set)

    return f_cond, logic_var, c1_var, bracket_var, c2_var, lb_effs, c3_var, lb_styles

def evaluate_advanced_condition(fmt, p, logic_mode, use_c1, regex_pat, use_c2, sel_effs, use_c3, sel_styles):
    if not use_c1 and not use_c2 and not use_c3: return True
        
    if fmt == "ASS":
        txt = p[9]
        effect = p[8].strip()
        style = p[3].strip()
    else: # SRT
        txt = p[2]
        effect = ""
        style = ""
        use_c2 = False
        use_c3 = False
        
    c_txt = re.sub(r'\{.*?\}', '', txt).strip()
    
    match_c1 = False
    if use_c1 and regex_pat:
        try:
            # 只要有任何匹配内容即视为 True，支持多行
            match_c1 = bool(re.search(regex_pat, c_txt))
        except:
            pass # 若用户正则语法填错，则静默视为不匹配
            
    match_c2 = use_c2 and effect in sel_effs
    match_c3 = use_c3 and style in sel_styles

    if logic_mode == 0: # AND
        if use_c1 and not match_c1: return False
        if use_c2 and not match_c2: return False
        if use_c3 and not match_c3: return False
        return True
    else: # OR
        if use_c1 and match_c1: return True
        if use_c2 and match_c2: return True
        if use_c3 and match_c3: return True
        return False
# ==========================================================

# ================= 预设配置文件 (跨平台兼容版) =================
# 解决 PyInstaller 在 macOS 打包后无权限在当前目录读写 json 文件的问题
def get_config_dir():
    if platform.system() == 'Windows':
        # Windows: 存在 AppData 下
        conf_dir = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'SubtitleToolbox')
    else:
        # macOS / Linux: 存在用户主目录下的隐藏文件夹
        conf_dir = os.path.join(os.path.expanduser('~'), '.subtitle_toolbox')
    os.makedirs(conf_dir, exist_ok=True)
    return conf_dir

CONFIG_DIR = get_config_dir()
PRESET_FILE_REP = os.path.join(CONFIG_DIR, "column_presets.json")
PRESET_FILE_SPLIT = os.path.join(CONFIG_DIR, "split_presets.json")
PRESET_FILE_ASS = os.path.join(CONFIG_DIR, "ass_presets.json")

DEFAULT_PRESETS_REP = ["A, B, E", "B, C, I"]
DEFAULT_PRESETS_SPLIT = ["A, B, C, D", "A, B, C, D, E"]
DEFAULT_PRESETS_ASS = {
    "默认样式": {
        "play_res_x": "1080", "play_res_y": "1920",
        "n_font": "SimHei", "n_size": "60", "n_color": "#FFFFFF", "n_out_color": "#000000",
        "n_margin_v": "20", "n_margin_lr": "20", "n_outline": "2", "n_align": "2", "n_shadow": "0", "n_bold": 0, "n_italic": 0,
        "s_font": "SimHei", "s_size": "60", "s_color": "#26E3FF", "s_out_color": "#000000",
        "s_margin_v": "850", "s_margin_lr": "20", "s_outline": "2", "s_align": "8", "s_shadow": "0", "s_bold": 0, "s_italic": 0
    }
}

def load_presets(filepath, default_data):
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return default_data.copy() if isinstance(default_data, list) else dict(default_data)
    return default_data.copy() if isinstance(default_data, list) else dict(default_data)

def save_presets_to_file(filepath, presets):
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(presets, f, ensure_ascii=False)
    except Exception as e:
        print(f"保存预设 {filepath} 失败:", e)

def col2num(col_str):
    num = 0
    for c in col_str.upper():
        if 'A' <= c <= 'Z':
            num = num * 26 + (ord(c) - ord('A')) + 1
    return num - 1

global_ass_preset_cbs = []

def get_ass_resolution(filepath):
    """解析 ASS 文件的 PlayResX 和 PlayResY"""
    rx, ry = "1080", "1920"
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            for line in f:
                if line.startswith('PlayResX:'): rx = line.split(':')[1].strip()
                elif line.startswith('PlayResY:'): ry = line.split(':')[1].strip()
                elif line.startswith('[Events]'): break
    return rx, ry

def update_all_ass_preset_cbs():
    keys = list(current_presets_ass.keys())
    for cb in global_ass_preset_cbs:
        cb['values'] = keys

def create_ass_preset_bar(parent, n_vars, s_vars, c_btns, preset_cb_var, res_vars=None):
    f_ps = ttk.Frame(parent)
    ttk.Label(f_ps, text="样式预设:").pack(side=tk.LEFT, padx=(0,5))
    cb = ttk.Combobox(f_ps, textvariable=preset_cb_var, values=list(current_presets_ass.keys()), width=15)
    cb.pack(side=tk.LEFT, padx=5)
    global_ass_preset_cbs.append(cb)
    
    def load_p(event=None):
        name = preset_cb_var.get()
        if name in current_presets_ass:
            d = current_presets_ass[name]
            if res_vars and len(res_vars) >= 2:
                res_vars[0].set(d.get("play_res_x", "1080"))
                res_vars[1].set(d.get("play_res_y", "1920"))
            n_vars[0].set(d.get("n_font", "SimHei")); n_vars[1].set(d.get("n_size", "60"))
            n_vars[2].set(d.get("n_color", "#FFFFFF")); n_vars[3].set(d.get("n_out_color", "#000000"))
            n_vars[4].set(d.get("n_margin_v", "20")); n_vars[5].set(d.get("n_margin_lr", "20")); n_vars[6].set(d.get("n_outline", "2"))
            n_vars[7].set(d.get("n_align", "2")); n_vars[8].set(d.get("n_shadow", "0"))
            n_vars[9].set(d.get("n_bold", 0)); n_vars[10].set(d.get("n_italic", 0))
            if len(n_vars) > 11:
                n_vars[11].set(d.get("n_alpha", "00")); n_vars[12].set(d.get("n_out_alpha", "00"))
            update_color_btn(c_btns[0], d.get("n_color", "#FFFFFF")); update_color_btn(c_btns[1], d.get("n_out_color", "#000000"))
            
            if s_vars and len(s_vars) >= 11 and len(c_btns) == 4:
                s_vars[0].set(d.get("s_font", "SimHei")); s_vars[1].set(d.get("s_size", "60"))
                s_vars[2].set(d.get("s_color", "#26E3FF")); s_vars[3].set(d.get("s_out_color", "#000000"))
                s_vars[4].set(d.get("s_margin_v", "850")); s_vars[5].set(d.get("s_margin_lr", "20")); s_vars[6].set(d.get("s_outline", "2"))
                s_vars[7].set(d.get("s_align", "8")); s_vars[8].set(d.get("s_shadow", "0"))
                s_vars[9].set(d.get("s_bold", 0)); s_vars[10].set(d.get("s_italic", 0))
                if len(s_vars) > 11:
                    s_vars[11].set(d.get("s_alpha", "00")); s_vars[12].set(d.get("s_out_alpha", "00"))
                update_color_btn(c_btns[2], d.get("s_color", "#26E3FF")); update_color_btn(c_btns[3], d.get("s_out_color", "#000000"))

    def save_p():
        name = preset_cb_var.get().strip()
        if not name: return messagebox.showwarning("提示", "请输入预设名称！")
        d = current_presets_ass.get(name, DEFAULT_PRESETS_ASS["默认样式"].copy())
        if res_vars and len(res_vars) >= 2:
            d.update({"play_res_x": res_vars[0].get(), "play_res_y": res_vars[1].get()})
        d.update({"n_font": n_vars[0].get(), "n_size": n_vars[1].get(), "n_color": n_vars[2].get(), "n_out_color": n_vars[3].get(), "n_margin_v": n_vars[4].get(), "n_margin_lr": n_vars[5].get(), "n_outline": n_vars[6].get(), "n_align": n_vars[7].get(), "n_shadow": n_vars[8].get(), "n_bold": n_vars[9].get(), "n_italic": n_vars[10].get()})
        if len(n_vars) > 11:
            d.update({"n_alpha": n_vars[11].get(), "n_out_alpha": n_vars[12].get()})
        if s_vars: 
            d.update({"s_font": s_vars[0].get(), "s_size": s_vars[1].get(), "s_color": s_vars[2].get(), "s_out_color": s_vars[3].get(), "s_margin_v": s_vars[4].get(), "s_margin_lr": s_vars[5].get(), "s_outline": s_vars[6].get(), "s_align": s_vars[7].get(), "s_shadow": s_vars[8].get(), "s_bold": s_vars[9].get(), "s_italic": s_vars[10].get()})
            if len(s_vars) > 11:
                d.update({"s_alpha": s_vars[11].get(), "s_out_alpha": s_vars[12].get()})
        current_presets_ass[name] = d; save_presets_to_file(PRESET_FILE_ASS, current_presets_ass); update_all_ass_preset_cbs(); preset_cb_var.set(name); messagebox.showinfo("提示", "样式预设保存成功！")
        
    def del_p():
        name = preset_cb_var.get().strip()
        if name in current_presets_ass:
            del current_presets_ass[name]; save_presets_to_file(PRESET_FILE_ASS, current_presets_ass); update_all_ass_preset_cbs()
            keys = list(current_presets_ass.keys()); preset_cb_var.set(keys[0] if keys else ""); load_p(); messagebox.showinfo("提示", "预设已删除！")
        else: messagebox.showwarning("提示", "未找到该预设！")

    cb.bind("<<ComboboxSelected>>", load_p); ttk.Button(f_ps, text="保存预设", command=save_p).pack(side=tk.LEFT, padx=5); ttk.Button(f_ps, text="删除选中", command=del_p).pack(side=tk.LEFT, padx=5)
    if preset_cb_var.get(): load_p()
    return f_ps

# ================= 核心功能逻辑 =================

def process_ass_merge(dir1, dir2, out_dir):
    files1 = {f for f in os.listdir(dir1) if f.lower().endswith('.ass')}
    files2 = {f for f in os.listdir(dir2) if f.lower().endswith('.ass')}
    
    common_files = files1.intersection(files2)
    if not common_files: raise ValueError("两个文件夹中没有找到同名的 ASS 文件！")
    
    os.makedirs(out_dir, exist_ok=True)
    processed_count = 0
    
    for file in common_files:
        path1 = os.path.join(dir1, file)
        path2 = os.path.join(dir2, file)
        
        with open(path1, 'r', encoding='utf-8-sig') as f: lines1 = f.read().split('\n')
        with open(path2, 'r', encoding='utf-8-sig') as f: lines2 = f.read().split('\n')
        
        def parse_ass_parts(lines):
            info, styles, events = [], [], []
            curr = "info"
            for l in lines:
                ls = l.strip()
                if ls.startswith('[V4+ Styles]'): curr = "styles"
                elif ls.startswith('[Events]'): curr = "events"
                
                if curr == "info": info.append(l)
                elif curr == "styles": styles.append(l)
                elif curr == "events": events.append(l)
            return info, styles, events
            
        info1, styles1, events1 = parse_ass_parts(lines1)
        info2, styles2, events2 = parse_ass_parts(lines2)
        
        # 1. 头部信息：取文件1的
        merged_info = info1
        
        # 2. 样式去重
        merged_styles = []
        seen_style_names = set()
        
        # 优先保留 [V4+ Styles] 和 Format 行
        for l in styles1:
            if not l.strip().startswith('Style:'):
                merged_styles.append(l)
            else:
                name = l.split('Style:')[1].split(',')[0].strip()
                if name not in seen_style_names:
                    merged_styles.append(l)
                    seen_style_names.add(name)
                    
        for l in styles2:
            if l.strip().startswith('Style:'):
                name = l.split('Style:')[1].split(',')[0].strip()
                if name not in seen_style_names:
                    merged_styles.append(l)
                    seen_style_names.add(name)
                    
        # 3. 字幕行合并：文件夹2 拼在 文件夹1 顶部
        merged_events = []
        has_events_header = False
        
        # 先加入文件2的 Dialogue (跳过 [Events] 和 Format 行)
        for l in events2:
            if l.strip().startswith('[Events]') or l.strip().startswith('Format:'):
                if not has_events_header:
                    merged_events.append(l)
                    has_events_header = True
            elif l.strip().startswith('Dialogue:'):
                merged_events.append(l)
                
        # 再加入文件1的 Dialogue
        for l in events1:
            if l.strip().startswith('[Events]') or l.strip().startswith('Format:'):
                if not has_events_header:
                    merged_events.append(l)
                    has_events_header = True
            elif l.strip().startswith('Dialogue:'):
                merged_events.append(l)

        # 组合写入
        out_content = "\n".join(merged_info) + "\n" + "\n".join(merged_styles) + "\n" + "\n".join(merged_events) + "\n"
        with open(os.path.join(out_dir, file), 'w', encoding='utf-8') as f:
            f.write(out_content)
            
        processed_count += 1
        
    return processed_count

def process_time_split(in_dir, out_norm_dir, out_scr_dir, mode="SRT"):
    files = [f for f in os.listdir(in_dir) if f.lower().endswith('.srt' if mode == "SRT" else '.ass')]
    if not files: raise ValueError(f"输入文件夹中没有找到 {mode} 文件！")
    
    os.makedirs(out_norm_dir, exist_ok=True)
    os.makedirs(out_scr_dir, exist_ok=True)
    
    processed_count = 0

    def srt_time_to_ms(time_str):
        time_str = time_str.replace(',', '.')
        parts = time_str.split(':')
        h = int(parts[0])
        m = int(parts[1])
        s, ms = parts[2].split('.')
        return (h * 3600 + m * 60 + int(s)) * 1000 + int(ms)

    def ass_time_to_ms(time_str):
        parts = time_str.split(':')
        h = int(parts[0])
        m = int(parts[1])
        s, ms = parts[2].split('.')
        return (h * 3600 + m * 60 + int(s)) * 1000 + int(ms) * 10
    
    for file in files:
        filepath = os.path.join(in_dir, file)
        base_name = os.path.splitext(file)[0]
        
        if mode == "SRT":
            blocks = parse_srt_file(filepath)
            split_idx = -1
            prev_end_ms = -1
            
            # 自动扫描分界点
            for i, block in enumerate(blocks):
                start_str, end_str = block['Timeline'].split(' --> ')
                curr_start_ms = srt_time_to_ms(start_str.strip())
                curr_end_ms = srt_time_to_ms(end_str.strip())
                
                # 如果当前行开始时间早于上一行结束时间，找到分界点
                if prev_end_ms != -1 and curr_start_ms < prev_end_ms:
                    split_idx = i
                    break
                prev_end_ms = curr_end_ms
                
            # 如果没有找到任何倒挂，默认全为对白
            if split_idx == -1: split_idx = 0
            
            norm_blocks_raw = blocks[split_idx:]
            scr_blocks_raw = blocks[:split_idx]
            
            # 重新生成 ID
            norm_blocks = [f"{i+1}\n{b['Timeline']}\n{b['Text']}\n" for i, b in enumerate(norm_blocks_raw)]
            scr_blocks = [f"{i+1}\n{b['Timeline']}\n{b['Text']}\n" for i, b in enumerate(scr_blocks_raw)]
            
            with open(os.path.join(out_norm_dir, base_name + '.srt'), 'w', encoding='utf-8') as f:
                f.write("\n".join(norm_blocks))
            with open(os.path.join(out_scr_dir, base_name + '.srt'), 'w', encoding='utf-8') as f:
                f.write("\n".join(scr_blocks))
            processed_count += 1
            
        elif mode == "ASS":
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                lines = f.read().split('\n')
            
            h_lines, s_lines, ev_lines = [], [], []
            curr = "info"
            for line in lines:
                l = line.strip()
                if l.startswith('[V4+ Styles]'): curr = "styles"
                elif l.startswith('[Events]'): curr = "events"
                
                if curr == "info": h_lines.append(line)
                elif curr == "styles": s_lines.append(line)
                elif curr == "events": ev_lines.append(line)
                
            dialogue_lines = []
            other_ev_lines = []
            for ev in ev_lines:
                if ev.startswith('Dialogue:'):
                    dialogue_lines.append(ev)
                else:
                    other_ev_lines.append(ev)
                    
            split_idx = -1
            prev_end_ms = -1
            
            # 自动扫描分界点
            for i, ev in enumerate(dialogue_lines):
                parts = ev.split(',', 9)
                if len(parts) >= 10:
                    curr_start_ms = ass_time_to_ms(parts[1].strip())
                    curr_end_ms = ass_time_to_ms(parts[2].strip())
                    
                    if prev_end_ms != -1 and curr_start_ms < prev_end_ms:
                        split_idx = i
                        break
                    prev_end_ms = curr_end_ms
            
            if split_idx == -1: split_idx = 0
            
            screen_ev = other_ev_lines + dialogue_lines[:split_idx]
            normal_ev = other_ev_lines + dialogue_lines[split_idx:]
            
            with open(os.path.join(out_norm_dir, file), 'w', encoding='utf-8') as f:
                f.write("\n".join(h_lines) + "\n" + "\n".join(s_lines) + "\n" + "\n".join(normal_ev) + "\n")
            with open(os.path.join(out_scr_dir, file), 'w', encoding='utf-8') as f:
                f.write("\n".join(h_lines) + "\n" + "\n".join(s_lines) + "\n" + "\n".join(screen_ev) + "\n")
            processed_count += 1
            
    return processed_count

def process_split(input_file, out_dir_src, out_dir_tgt, cols_list):
    if input_file.lower().endswith('.csv'): df = pd.read_csv(input_file)
    else: df = pd.read_excel(input_file)
        
    c_file, c_id, c_time = col2num(cols_list[0]), col2num(cols_list[1]), col2num(cols_list[2])
    c_texts = [col2num(c) for c in cols_list[3:]]
    max_col_needed = max([c_file, c_id, c_time] + c_texts)
    if max_col_needed >= len(df.columns): raise ValueError(f"指定的列字母超出范围！\n当前表格只有 {len(df.columns)} 列。")
       
    file_data = {}
    for index, row in df.iterrows():
        filename_val = str(row.iloc[c_file]).strip()
        if pd.isna(row.iloc[c_file]) or not filename_val or filename_val == "nan": continue
        
        # 修复：防止 Excel 将 0001 吞成 1，自动强制补齐四位数字
        if filename_val.endswith('.0'): filename_val = filename_val[:-2]
        base_name = filename_val[:-4] if filename_val.lower().endswith('.srt') else filename_val
        if base_name.isdigit(): base_name = base_name.zfill(4)
        filename_val = base_name + '.srt'
            
        id_val = str(row.iloc[c_id]).strip()
        if id_val.endswith('.0'): id_val = id_val[:-2]
        time_val = str(row.iloc[c_time]).strip()
 
        if filename_val not in file_data: file_data[filename_val] = []
        file_data[filename_val].append((id_val, time_val, row))
        
    if not file_data: raise ValueError("未找到有效的数据，请检查列名是否正确！")

    for i, c_text in enumerate(c_texts):
        lang_dir = out_dir_src if i == 0 else out_dir_tgt
        if not lang_dir: continue 
        os.makedirs(lang_dir, exist_ok=True)
 
        for filename_val, rows in file_data.items():
            
            # --- 智能排序机制 ---
            def get_sort_key(item):
                try: return (0, int(item[0]))
                except: return (1, item[0])
            rows.sort(key=get_sort_key)
            
            srt_content = []
            for id_val, time_val, row in rows:
                text_val = row.iloc[c_text]
                text_val = "" if pd.isna(text_val) else str(text_val).strip()
                srt_content.append(f"{id_val}\n{time_val}\n{text_val}\n")
            with open(os.path.join(lang_dir, filename_val), 'w', encoding='utf-8') as f:
                f.write("\n".join(srt_content))
    return len(file_data)

def process_split_mode2(input_file, out_dir_src, out_dir_tgt, cols_list, sheet_name):
    """XLSX 转 SRT 模式2：无 ID 列，按行自增，支持双目录输出"""
    if input_file.lower().endswith('.csv'): df = pd.read_csv(input_file)
    else: df = pd.read_excel(input_file, sheet_name=sheet_name if sheet_name else 0)
        
    c_file, c_time = col2num(cols_list[0]), col2num(cols_list[1])
    c_texts = [col2num(c) for c in cols_list[2:]]
    max_col_needed = max([c_file, c_time] + c_texts)
    if max_col_needed >= len(df.columns): raise ValueError("指定的列字母超出表格实际范围！")
       
    file_data = {}
    for index, row in df.iterrows():
        filename_val = str(row.iloc[c_file]).strip()
        if pd.isna(row.iloc[c_file]) or not filename_val or filename_val == "nan": continue
        
        # 修复：防止 Excel 将 0001 吞成 1，自动强制补齐四位数字
        if filename_val.endswith('.0'): filename_val = filename_val[:-2]
        base_name = filename_val[:-4] if filename_val.lower().endswith('.srt') else filename_val
        if base_name.isdigit(): base_name = base_name.zfill(4)
        filename_val = base_name + '.srt'
            
        time_val = str(row.iloc[c_time]).strip()
        id_val = str(index + 1)
 
        if filename_val not in file_data: file_data[filename_val] = []
        file_data[filename_val].append((id_val, time_val, row))
        
    if not file_data: raise ValueError("未找到有效的数据，请检查列名是否正确！")

    for i, c_text in enumerate(c_texts):
        lang_dir = out_dir_src if i == 0 else out_dir_tgt
        if not lang_dir: continue
        os.makedirs(lang_dir, exist_ok=True)
 
        for filename_val, rows in file_data.items():
            srt_content = []
            for id_val, time_val, row in rows:
                text_val = row.iloc[c_text]
                text_val = "" if pd.isna(text_val) else str(text_val).strip()
                srt_content.append(f"{id_val}\n{time_val}\n{text_val}\n")
            with open(os.path.join(lang_dir, filename_val), 'w', encoding='utf-8') as f:
                f.write("\n".join(srt_content))
    return len(file_data)

def parse_srt_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f: content = f.read().strip()
    blocks = re.split(r'\n\s*\n', content)
    parsed_data = []
    for block in blocks:
        lines = block.strip().split('\n')
        if len(lines) >= 3:
            idx = lines[0].strip()
            timeline = lines[1].strip()
            text = "\n".join(lines[2:]).strip()
            parsed_data.append({'ID': idx, 'Timeline': timeline, 'Text': text})
    return parsed_data

def process_merge(src_dir, tgt_dir, src_lang_name, tgt_lang_name, output_excel):
    src_files = {f for f in os.listdir(src_dir) if f.lower().endswith('.srt')}
    tgt_files = {f for f in os.listdir(tgt_dir) if f.lower().endswith('.srt')}
    common_files = src_files.intersection(tgt_files)
    if not common_files: raise ValueError("两个目录中没有找到同名的 SRT 文件，无法合并！")
        
    # 修复：防止文件 1, 10, 2 乱序，强制使用数字自然数大小进行排序
    def get_sort_key(f):
        b = os.path.splitext(f)[0]
        return int(b) if b.isdigit() else b

    master_data = []
    for file_name in sorted(list(common_files), key=get_sort_key):
        base_name = os.path.splitext(file_name)[0]
        # 修复：强行补充四位数字（写入 Excel 的时候保证是 0001）
        if base_name.isdigit(): base_name = base_name.zfill(4)
        
        src_path, tgt_path = os.path.join(src_dir, file_name), os.path.join(tgt_dir, file_name)
        src_data, tgt_data = parse_srt_file(src_path), parse_srt_file(tgt_path)
        
        for i, src_item in enumerate(src_data):
            tgt_text = tgt_data[i]['Text'] if i < len(tgt_data) else ""
            master_data.append({
                'Episode': base_name, 'ID': src_item['ID'], 'Time': src_item['Timeline'],
                src_lang_name: src_item['Text'], tgt_lang_name: tgt_text
            })
    pd.DataFrame(master_data).to_excel(output_excel, index=False)
    return len(common_files)

def process_merge_mode2(src_dir, tgt_dir, src_lang_name, tgt_lang_name, output_excel):
    """SRT 合并为 XLSX 模式2：无ID列，保留文件和时间轴，支持单语言合并"""
    src_files = {f for f in os.listdir(src_dir) if f.lower().endswith('.srt')} if os.path.exists(src_dir) else set()
    if not src_files: raise ValueError("源语言目录中没有找到 SRT 文件！")
        
    # 修复：防止文件 1, 10, 2 乱序，强制使用数字自然数大小进行排序
    def get_sort_key(f):
        b = os.path.splitext(f)[0]
        return int(b) if b.isdigit() else b

    master_data = []
    for file_name in sorted(list(src_files), key=get_sort_key):
        base_name = os.path.splitext(file_name)[0]
        # 修复：强行补充四位数字（写入 Excel 的时候保证是 0001）
        if base_name.isdigit(): base_name = base_name.zfill(4)
        
        src_path = os.path.join(src_dir, file_name)
        src_data = parse_srt_file(src_path)
        
        tgt_data = []
        if tgt_dir and os.path.exists(tgt_dir):
            tgt_path = os.path.join(tgt_dir, file_name)
            if os.path.exists(tgt_path):
                tgt_data = parse_srt_file(tgt_path)
        
        for i, src_item in enumerate(src_data):
            tgt_text = tgt_data[i]['Text'] if i < len(tgt_data) else ""
            row_dict = {
                'Episode': base_name, 
                'Time': src_item['Timeline'],
                src_lang_name if src_lang_name else '源语言': src_item['Text']
            }
            if tgt_dir and tgt_lang_name:
                row_dict[tgt_lang_name] = tgt_text
            master_data.append(row_dict)
            
    pd.DataFrame(master_data).to_excel(output_excel, index=False, sheet_name="Subtitle Translation")
    return len(src_files)

def process_replace(report_file, srt_dir, out_summary, col_filename_str, col_id_str, col_text_str, match_mode=0):
    if report_file.lower().endswith('.csv'): df = pd.read_csv(report_file)
    else: df = pd.read_excel(report_file)
        
    c_file, c_id, c_text = col2num(col_filename_str), col2num(col_id_str), col2num(col_text_str)
    if max(c_file, c_id, c_text) >= len(df.columns): raise ValueError(f"指定的列字母超出范围！")
        
    srt_cache, summary_data = {}, []
    for index, row in df.iterrows():
        filename_val, match_val, text_val = str(row.iloc[c_file]).strip(), str(row.iloc[c_id]).strip(), row.iloc[c_text]
        # ID模式下自动去除由于Excel数值格式产生的 .0
        if match_mode == 0 and match_val.endswith('.0'): match_val = match_val[:-2]
        
        if pd.isna(text_val) or str(text_val).strip() == "" or filename_val == "nan" or match_val == "nan": continue
            
        text_val = str(text_val).strip()
        basename = os.path.basename(filename_val.replace('\\', '/'))
        if not basename.lower().endswith('.srt'): basename += '.srt'
        filepath = os.path.join(srt_dir, basename)
        if not os.path.exists(filepath): continue
            
        if basename not in srt_cache: srt_cache[basename] = parse_srt_file(filepath)
        for block in srt_cache[basename]:
            is_match = False
            if match_mode == 0:
                # 模式0：严格按 ID 匹配
                is_match = (block['ID'] == match_val)
            else:
                # 模式1：按时间轴匹配（智能无视空格，兼容英文句号与逗号的差异进行对比）
                t1 = block['Timeline'].replace(' ', '').replace(',', '.')
                t2 = match_val.replace(' ', '').replace(',', '.')
                is_match = (t1 == t2)

            if is_match:
                old_text = block['Text']
                if old_text != text_val:
                    block['Text'] = text_val
                    match_type_str = "字幕ID" if match_mode == 0 else "时间轴"
                    summary_data.append({'SRT文件名': basename, match_type_str: match_val, '原字幕内容': old_text, '替换后新内容': text_val})
                break
                
    for basename, blocks in srt_cache.items():
        srt_content = [f"{block['ID']}\n{block['Timeline']}\n{block['Text']}\n" for block in blocks]
        with open(os.path.join(srt_dir, basename), 'w', encoding='utf-8') as f:
            f.write("\n".join(srt_content))
    if summary_data: pd.DataFrame(summary_data).to_excel(out_summary, index=False)
    return len(summary_data), len(srt_cache)

def process_zip(target_dir, output_dir, max_files):
    files = sorted([f for f in os.listdir(target_dir) if os.path.isfile(os.path.join(target_dir, f))])
    if not files: raise ValueError("目标文件夹中没有任何文件！")
    total_files = len(files)
    num_zips = math.ceil(total_files / max_files)
        
    for i in range(num_zips):
        chunk = files[i*max_files : (i+1)*max_files]
        first_name, last_name = os.path.splitext(chunk[0])[0], os.path.splitext(chunk[-1])[0]
        zip_filename = f"{first_name}.zip" if first_name == last_name else f"{first_name}_{last_name}.zip"
        with zipfile.ZipFile(os.path.join(output_dir, zip_filename), 'w', compression=zipfile.ZIP_STORED) as zf:
            for f in chunk: zf.write(os.path.join(target_dir, f), f)
    return total_files, num_zips

def ass_to_hex(ass_str):
    try:
        s = ass_str.upper().replace('&H', '')
        if len(s) >= 8: s = s[-6:]
        elif len(s) < 6: s = s.zfill(6)
        b, g, r = s[0:2], s[2:4], s[4:6]
        return f"#{r}{g}{b}"
    except: return "#FFFFFF"

# ================= ASS 相关辅助模块 =================

def srt_to_ass_time(srt_time):
    srt_time = srt_time.strip()
    time_part, ms_part = srt_time.split(',') if ',' in srt_time else (srt_time.split('.') if '.' in srt_time else (srt_time, "000"))
    parts = time_part.split(':')
    h, m, s = parts if len(parts) == 3 else ("00", parts[0], parts[1])
    return f"{int(h)}:{m}:{s}.{ms_part[:2]}"

def hex2ass_with_alpha(hex_str, alpha_str="00"):
    if not hex_str or len(hex_str) != 7: hex_str = "#FFFFFF"
    r, g, b = hex_str[1:3], hex_str[3:5], hex_str[5:7]
    a = str(alpha_str).strip().upper()
    if len(a) != 2: a = "00"
    return f"&H{a}{b}{g}{r}"

def clean_ass_text(txt):
    lines = [l.strip() for l in re.split(r'\\N|\n', txt) if l.strip()]
    return "\\N".join(lines)

def rename_style_line(style_line, new_name):
    if not style_line.startswith("Style:"): return style_line
    prefix, rest = style_line.split(":", 1)
    parts = rest.split(",")
    parts[0] = f" {new_name}"
    return f"{prefix}:{','.join(parts)}"

def replace_font_in_style(style_line, new_font):
    """直接替换样式字符串中的字体名称"""
    if not style_line.startswith("Style:"): return style_line
    prefix, rest = style_line.split(":", 1)
    parts = rest.split(",")
    if len(parts) > 1:
        parts[1] = new_font
    return f"{prefix}:{','.join(parts)}"

def build_ass_style_line(name, f_name, f_size, color, out_color, mv, mlr, outline, align="2", shadow="0", bold=0, italic=0, alpha="00", out_alpha="00"):
    b_val = "-1" if str(bold) == "1" else "0"
    i_val = "-1" if str(italic) == "1" else "0"
    c_ass = color if str(color).startswith("&H") else hex2ass_with_alpha(color, alpha)
    oc_ass = out_color if str(out_color).startswith("&H") else hex2ass_with_alpha(out_color, out_alpha)
    return f"Style: {name},{f_name},{f_size},{c_ass},&H000000FF,{oc_ass},&H00000000,{b_val},{i_val},0,0,100,100,0,0,1,{outline},{shadow},{align},{mlr},{mlr},{mv},1"

def merge_ass_dialogues(dialogues_list, filename, report_list):
    if not dialogues_list: return []
    def parse_diag(line):
        prefix, content = line.split(":", 1)
        parts = content.split(",", 9)
        return {"start": parts[1].strip(), "end": parts[2].strip(), "style": parts[3].strip(), "text": parts[9].strip(), "raw": parts, "prefix": prefix}
    
    parsed = [parse_diag(d) for d in dialogues_list]
    merged, current = [], parsed[0]
    
    for nxt in parsed[1:]:
        if current["text"] == nxt["text"] and current["style"] == nxt["style"] and current["end"] == nxt["start"]:
            report_list.append({"文件名": filename, "合并文本": current["text"], "原时间轴1": f"{current['start']} --> {current['end']}", "原时间轴2": f"{nxt['start']} --> {nxt['end']}", "合并后时间轴": f"{current['start']} --> {nxt['end']}"})
            current["end"] = nxt["end"]
            current["raw"][2] = nxt["end"]
        else:
            merged.append(current["prefix"] + ":" + ",".join(current["raw"]))
            current = nxt
            
    merged.append(current["prefix"] + ":" + ",".join(current["raw"]))
    return merged

# ================= 业务流程处理 =================

def process_srt_to_ass(input_dir, out_dir, bracket_str, regex_text, custom_style_dict, do_merge, merge_report_path, style_mode, ref_cfg):
    regex_pat = bracket_str.strip()

    replacements = []
    for line in regex_text.split('\n'):
        if '>>>' in line:
            pat, repl = line.split('>>>', 1)
            repl_python = re.sub(r'\$(\d+)', r'\\\1', repl.strip())
            replacements.append((pat.strip(), repl_python))

    files = [f for f in os.listdir(input_dir) if f.lower().endswith('.srt')]
    if not files: raise ValueError("指定的输入文件夹中没有找到 .srt 文件！")

    merge_reports = []

    for file in files:
        filepath = os.path.join(input_dir, file)
        out_path = os.path.join(out_dir, file.rsplit('.', 1)[0] + '.ass')

        blocks = parse_srt_file(filepath)
        screen_events, normal_events = [], []

        for block in blocks:
            text = block['Text']
            
            is_screen = False
            if regex_pat:
                try:
                    is_screen = bool(re.search(regex_pat, text))
                except:
                    pass

            if is_screen:
                s_text, n_text = text, ""
            else:
                s_text, n_text = "", text

            n_text = clean_ass_text(n_text)
            s_text = clean_ass_text(s_text)

            for pat, repl in replacements:
                if n_text: n_text = re.sub(pat, repl, n_text)
                if s_text: s_text = re.sub(pat, repl, s_text)

            start_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[0])
            end_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[1])

            if s_text: screen_events.append(f"Dialogue: 0,{start_ass},{end_ass},画面字,,0,0,0,,{s_text}")
            if n_text: normal_events.append(f"Dialogue: 0,{start_ass},{end_ass},对白字幕,,0,0,0,,{n_text}")

        if style_mode == 0:
            resx = custom_style_dict.get('play_res_x', '1080')
            resy = custom_style_dict.get('play_res_y', '1920')
        else:
            resx, resy = get_ass_resolution(ref_cfg['ref_path'])
            
        srt_script_info = f"[Script Info]\nScriptType: v4.00+\nPlayResX: {resx}\nPlayResY: {resy}\n"
        srt_styles_block = "[V4+ Styles]\nFormat: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding\n"
        
        if style_mode == 0:
            d = custom_style_dict
            srt_styles_block += build_ass_style_line("对白字幕", d['n_font'], d['n_size'], d['n_color'], d['n_out_color'], d['n_margin_v'], d['n_margin_lr'], d['n_outline'], d.get('n_align','2'), d.get('n_shadow','0'), d.get('n_bold',0), d.get('n_italic',0), d.get('n_alpha','00'), d.get('n_out_alpha','00')) + "\n"
            srt_styles_block += build_ass_style_line("画面字", d['s_font'], d['s_size'], d['s_color'], d['s_out_color'], d['s_margin_v'], d['s_margin_lr'], d['s_outline'], d.get('s_align','8'), d.get('s_shadow','0'), d.get('s_bold',0), d.get('s_italic',0), d.get('s_alpha','00'), d.get('s_out_alpha','00'))
        else:
            ref_styles = scan_all_styles_from_ass(ref_cfg['ref_path'])
            n_line = rename_style_line(ref_styles.get(ref_cfg['n_style'], build_ass_style_line("对白字幕", "Arial", "60", "&H00FFFFFF", "&H00000000", "20", "20", "2")), "对白字幕")
            s_line = rename_style_line(ref_styles.get(ref_cfg['s_style'], build_ass_style_line("画面字", "Arial", "60", "&H00FFFFFF", "&H00000000", "850", "20", "2")), "画面字")
            
            if ref_cfg['font_mode'] == 1:
                n_line = replace_font_in_style(n_line, ref_cfg['override_font'])
                s_line = replace_font_in_style(s_line, ref_cfg['override_font'])
            
            srt_styles_block += n_line + "\n" + s_line

        ass_events_str = "[Events]\nFormat: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text\n"
        ass_events_str += "\n".join(screen_events) + ("\n" if screen_events else "")
        ass_events_str += "\n".join(normal_events) + "\n"

        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(f"{srt_script_info}\n{srt_styles_block}\n\n{ass_events_str}")

    if do_merge and merge_reports and merge_report_path:
        pd.DataFrame(merge_reports).to_excel(merge_report_path, index=False)

    return len(files)

def scan_all_styles_from_ass(filepath):
    """扫描 ASS 返回字典 {样式名: 原样式行内容}"""
    styles = {}
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            in_styles = False
            for line in f:
                line_s = line.strip()
                if line_s.startswith('[V4+ Styles]'): in_styles = True; continue
                if line_s.startswith('[Events]'): break
                if in_styles and line_s.startswith('Style:'):
                    name = line_s.split('Style:')[1].split(',')[0].strip()
                    styles[name] = line_s
    except: pass
    return styles

def process_ass_editor(input_dir, out_dir, mode_cfg):
    pass # Reserved, actual logic integrated into execute_ass_editor

def process_column_copy_batch(src_dir, tgt_dir, out_dir, err_rep, fmt, col_str, sel_blocks=None):
    is_header = col_str.startswith("Header")
    if not is_header:
        col_idx = int(col_str.split(':')[0])
        
    ext = '.srt' if fmt == 'SRT' else '.ass'
    tgt_files = [f for f in os.listdir(tgt_dir) if f.lower().endswith(ext)]
    if not tgt_files: raise ValueError(f"待接收数据的目标文件夹中没有 {ext} 文件！")
    
    all_errors = []
    processed_count = 0
    os.makedirs(out_dir, exist_ok=True)
    
    for file in tgt_files:
        src_file = os.path.join(src_dir, file)
        tgt_file = os.path.join(tgt_dir, file)
        out_file = os.path.join(out_dir, file)
    
        if not os.path.exists(src_file):
            all_errors.append({'文件名': file, '目标行号/时间轴': 'N/A', '目标文本': 'N/A', '错误说明': '提供数据的源文件夹中找不到同名文件'})
            with open(tgt_file, 'r', encoding='utf-8-sig') as f, open(out_file, 'w', encoding='utf-8') as out_f:
                out_f.write(f.read())
            continue

        if fmt == "SRT":
            src_blocks = parse_srt_file(src_file)
            tgt_blocks = parse_srt_file(tgt_file)

            if len(src_blocks) != len(tgt_blocks):
                all_errors.append({'文件名': file, '目标行号/时间轴': 'N/A', '目标文本': 'N/A', '错误说明': f'警告：行数不一致! 源:{len(src_blocks)}行, 目标:{len(tgt_blocks)}行'})

            out_blocks = []
            for i, t_b in enumerate(tgt_blocks):
                p = [t_b['ID'], t_b['Timeline'], t_b['Text']]
                if i < len(src_blocks):
                    s_b = src_blocks[i]
                    s_p = [s_b['ID'], s_b['Timeline'], s_b['Text']]

                    if t_b['Timeline'] != s_b['Timeline']:
                        all_errors.append({'文件名': file, '目标行号/时间轴': f"第{i+1}行 {t_b['Timeline']}", '目标文本': t_b['Text'], '错误说明': f"时间轴不一致! 源时间轴为 {s_b['Timeline']}"})
                    
                    p[col_idx] = s_p[col_idx] # 粗暴覆盖指定列
                else:
                    all_errors.append({'文件名': file, '目标行号/时间轴': f"第{i+1}行 {t_b['Timeline']}", '目标文本': t_b['Text'], '错误说明': '源文件在此行缺失，无法复制'})

                out_blocks.append(f"{p[0]}\n{p[1]}\n{p[2]}\n")

            with open(out_file, 'w', encoding='utf-8') as f: f.write("\n".join(out_blocks))
            processed_count += 1

        else: # ASS 格式
            with open(src_file, 'r', encoding='utf-8-sig') as f: src_lines = f.read().split('\n')
            with open(tgt_file, 'r', encoding='utf-8-sig') as f: tgt_lines = f.read().split('\n')

            if is_header:
                def parse_blocks(lines):
                    order = []
                    blocks = {}
                    curr = ""
                    for l in lines:
                        s = l.strip()
                        if s.startswith('[') and s.endswith(']'):
                            curr = s
                            if curr not in order: order.append(curr)
                        if curr not in blocks: blocks[curr] = []
                        blocks[curr].append(l)
                    return order, blocks

                src_order, src_blocks = parse_blocks(src_lines)
                tgt_order, tgt_blocks = parse_blocks(tgt_lines)
                
                out_lines = []
                
                if not sel_blocks:
                    # 全量替换文件头：用源文件的头，加上目标文件的 [Events]
                    for b in src_order:
                        if b != '[Events]':
                            out_lines.extend(src_blocks[b])
                    if '[Events]' in tgt_blocks:
                        out_lines.extend(tgt_blocks['[Events]'])
                else:
                    # 按指定块替换
                    replaced = set()
                    for b in tgt_order:
                        if b == '[Events]':
                            continue
                        if b in sel_blocks:
                            if b in src_blocks:
                                out_lines.extend(src_blocks[b])
                                replaced.add(b)
                            else:
                                out_lines.extend(tgt_blocks[b])
                        else:
                            out_lines.extend(tgt_blocks[b])
                            
                    # 如果有源文件中存在但目标文件原来没有的块，追加在 Events 之前
                    for b in sel_blocks:
                        if b not in replaced and b in src_blocks and b != '[Events]':
                            out_lines.extend(src_blocks[b])
                            
                    if '[Events]' in tgt_blocks:
                        out_lines.extend(tgt_blocks['[Events]'])
                
                with open(out_file, 'w', encoding='utf-8') as f: f.write("\n".join(out_lines))
                processed_count += 1

            else:
                # 原有的按列复制逻辑
                src_diags = [l for l in src_lines if l.strip().startswith('Dialogue:')]
                out_lines = []
                diag_idx = 0
                
                tgt_diags_count = len([l for l in tgt_lines if l.strip().startswith('Dialogue:')])
                if len(src_diags) != tgt_diags_count:
                    all_errors.append({'文件名': file, '目标行号/时间轴': 'N/A', '目标文本': 'N/A', '错误说明': f"Dialogue行数不一致! 源:{len(src_diags)}行, 目标:{tgt_diags_count}行"})

                for line in tgt_lines:
                    if line.strip().startswith('Dialogue:'):
                        t_p = line.split(',', 9)
                        if len(t_p) >= 10:
                            if diag_idx < len(src_diags):
                                s_line = src_diags[diag_idx]
                                s_p = s_line.split(',', 9)
                                if len(s_p) >= 10:
                                    t_time = f"{t_p[1]} --> {t_p[2]}"
                                    s_time = f"{s_p[1]} --> {s_p[2]}"
                                    if t_time != s_time:
                                        all_errors.append({'文件名': file, '目标行号/时间轴': f"第{diag_idx+1}条 {t_time}", '目标文本': t_p[9], '错误说明': f"时间轴不一致! 源时间轴为 {s_time}"})
                                    
                                    if col_idx < len(t_p) and col_idx < len(s_p):
                                        t_p[col_idx] = s_p[col_idx] # 粗暴覆盖指定列
                                        line = ",".join(t_p)
                            else:
                                all_errors.append({'文件名': file, '目标行号/时间轴': f"第{diag_idx+1}条", '目标文本': t_p[9] if len(t_p)>9 else 'N/A', '错误说明': '源文件在此行缺失，无法复制'})
                        out_lines.append(line)
                        diag_idx += 1
                    else:
                        out_lines.append(line)

                with open(out_file, 'w', encoding='utf-8') as f: f.write("\n".join(out_lines))
                processed_count += 1

    # 新增判定：只有非文件头替换模式下，才输出行匹配异常报告
    if all_errors and err_rep and not is_header: pd.DataFrame(all_errors).to_excel(err_rep, index=False)
    return processed_count, len(all_errors)

def process_srt_bilingual_split_batch(in_dir, out_dir1, out_dir2, suffix1, suffix2, report_path=None, split_mode=1):
    files = [f for f in os.listdir(in_dir) if f.lower().endswith('.srt')]
    if not files: raise ValueError("输入文件夹中没有找到 .srt 文件！")
    os.makedirs(out_dir1, exist_ok=True)
    os.makedirs(out_dir2, exist_ok=True)
    total_blocks = 0
    all_errors = []
    
    # 模式1：经典宽泛识别（基于宽泛 CJK 区块 vs 拉丁字母）
    def char_profile_classic(s):
        cjk = len(re.findall(r'[\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af\u0e00-\u0e7f\u0400-\u04ff]', s))
        latin = len(re.findall(r'[a-zA-Z]', s))
        return 'C' if cjk > latin else 'L'

    # 模式2：强化东亚语言特征识别（精准锁定独占字符）
    def char_profile_advanced(s):
        kana = len(re.findall(r'[\u3040-\u30ff]', s))       # 平假名/片假名
        hangul = len(re.findall(r'[\uac00-\ud7af]', s))     # 韩文谚文
        hanzi = len(re.findall(r'[\u4e00-\u9fa5]', s))      # 纯汉字
        latin = len(re.findall(r'[a-zA-Z]', s))             # 拉丁字母
        
        if hangul > 0: return 'KR' # 出现韩文
        if kana > 0: return 'JP'   # 出现假名即判定为日语
        if hanzi > 0: return 'CN'  # 有汉字且无假名，判定为中文
        if latin > 0: return 'EN'  # 纯英文或拼音
        return 'OTHER'

    for file in files:
        srt_file = os.path.join(in_dir, file)
        blocks = parse_srt_file(srt_file)
        base_name = os.path.splitext(file)[0]
        out_blocks1, out_blocks2 = [], []
        
        for block in blocks:
            lines = [l.strip() for l in block['Text'].split('\n') if l.strip()]
            text1, text2 = "", ""
            if len(lines) == 0: pass
            elif len(lines) == 1: text1, text2 = lines[0], ""
            elif len(lines) == 2: text1, text2 = lines[0], lines[1]
            elif len(lines) == 4: text1, text2 = "\n".join(lines[:2]), "\n".join(lines[2:])
            elif len(lines) == 3:
                # 根据不同模式调用不同的画像函数
                if split_mode == 2:
                    p0, p1, p2 = char_profile_advanced(lines[0]), char_profile_advanced(lines[1]), char_profile_advanced(lines[2])
                else:
                    p0, p1, p2 = char_profile_classic(lines[0]), char_profile_classic(lines[1]), char_profile_classic(lines[2])
                
                # 逻辑：如果前两行语言类型一致，且与第三行不同，则切分为 2+1
                if p1 == p0 and p1 != p2: text1, text2 = "\n".join(lines[:2]), lines[2]
                elif p1 == p2 and p1 != p0: text1, text2 = lines[0], "\n".join(lines[1:])
                else: text1, text2 = "\n".join(lines[:2]), lines[2] # 兜底策略
            else:
                half = len(lines) // 2
                text1, text2 = "\n".join(lines[:half]), "\n".join(lines[half:])
                
            if not text1.strip() or not text2.strip():
                err_msg = []
                if not text1.strip(): err_msg.append(f"上方语言({suffix1})为空")
                if not text2.strip(): err_msg.append(f"下方语言({suffix2})为空")
                all_errors.append({
                    '文件名': file, '字幕ID': block['ID'], '时间轴': block['Timeline'],
                    '原始双语文本': block['Text'], '错误说明': " & ".join(err_msg)
                })
                
            out_blocks1.append(f"{block['ID']}\n{block['Timeline']}\n{text1}\n")
            out_blocks2.append(f"{block['ID']}\n{block['Timeline']}\n{text2}\n")
            
        # 智能拼接文件名：如果有后缀就加下划线和后缀，如果没有就直接用原名
        out_name1 = f"{base_name}_{suffix1}.srt" if suffix1 else f"{base_name}.srt"
        out_name2 = f"{base_name}_{suffix2}.srt" if suffix2 else f"{base_name}.srt"
        
        with open(os.path.join(out_dir1, out_name1), 'w', encoding='utf-8') as f: f.write("\n".join(out_blocks1))
        with open(os.path.join(out_dir2, out_name2), 'w', encoding='utf-8') as f: f.write("\n".join(out_blocks2))
        total_blocks += len(blocks)
        
    if all_errors and report_path:
        pd.DataFrame(all_errors).to_excel(report_path, index=False)
        
    return len(files), total_blocks, len(all_errors)

def process_merge_srt_to_ass_batch(norm_dir, scr_dir, out_dir, custom_style_dict, style_mode, ref_cfg, regex_cfg=None):
    os.makedirs(out_dir, exist_ok=True)
    
    # --- 解析正则替换规则 ---
    replacements = []
    target_mode = ""
    if regex_cfg and regex_cfg.get('enable'):
        target_mode = regex_cfg.get('target', '画面字')
        for line in regex_cfg.get('text', '').split('\n'):
            if '>>>' in line:
                pat, repl = line.split('>>>', 1)
                # 自动将用户习惯的 $1, $2 转换为 Python 底层支持的 \1, \2
                repl_python = re.sub(r'\$(\d+)', r'\\\1', repl.strip())
                replacements.append((pat.strip(), repl_python))

    norm_files = {f for f in os.listdir(norm_dir) if f.lower().endswith('.srt')} if os.path.exists(norm_dir) else set()
    scr_files = {f for f in os.listdir(scr_dir) if f.lower().endswith('.srt')} if os.path.exists(scr_dir) else set()
    
    all_files = sorted(list(norm_files.union(scr_files)))
    if not all_files: raise ValueError("输入的文件夹中没有找到任何 .srt 文件！")
    
    if style_mode == 0:
        resx = custom_style_dict.get('play_res_x', '1080')
        resy = custom_style_dict.get('play_res_y', '1920')
    else:
        resx, resy = get_ass_resolution(ref_cfg['ref_path'])
        
    srt_script_info = f"[Script Info]\nScriptType: v4.00+\nPlayResX: {resx}\nPlayResY: {resy}\n"
    srt_styles_block = "[V4+ Styles]\nFormat: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding\n"
    
    if style_mode == 0:
        d = custom_style_dict
        srt_styles_block += build_ass_style_line("对白字幕", d['n_font'], d['n_size'], d['n_color'], d['n_out_color'], d['n_margin_v'], d['n_margin_lr'], d['n_outline'], d.get('n_align','2'), d.get('n_shadow','0'), d.get('n_bold',0), d.get('n_italic',0), d.get('n_alpha','00'), d.get('n_out_alpha','00')) + "\n"
        srt_styles_block += build_ass_style_line("画面字", d['s_font'], d['s_size'], d['s_color'], d['s_out_color'], d['s_margin_v'], d['s_margin_lr'], d['s_outline'], d.get('s_align','8'), d.get('s_shadow','0'), d.get('s_bold',0), d.get('s_italic',0), d.get('s_alpha','00'), d.get('s_out_alpha','00'))
    else:
        ref_styles = scan_all_styles_from_ass(ref_cfg['ref_path'])
        n_line = rename_style_line(ref_styles.get(ref_cfg['n_style'], build_ass_style_line("对白字幕", "Arial", "60", "&H00FFFFFF", "&H00000000", "20", "20", "2")), "对白字幕")
        s_line = rename_style_line(ref_styles.get(ref_cfg['s_style'], build_ass_style_line("画面字", "Arial", "60", "&H00FFFFFF", "&H00000000", "850", "20", "2")), "画面字")
        
        if ref_cfg['font_mode'] == 1:
            n_line = replace_font_in_style(n_line, ref_cfg['override_font'])
            s_line = replace_font_in_style(s_line, ref_cfg['override_font'])
            
        srt_styles_block += n_line + "\n" + s_line
    
    for file in all_files:
        screen_events, normal_events = [], []
        if file in scr_files:
            blocks = parse_srt_file(os.path.join(scr_dir, file))
            for block in blocks:
                s_text = clean_ass_text(block['Text'])
                if not s_text: continue
                
                # === 应用正则到画面字 ===
                if replacements and target_mode in ['画面字', '全部']:
                    for pat, repl in replacements:
                        s_text = re.sub(pat, repl, s_text)
                if not s_text.strip(): continue # 如果替换后变成了空行，则跳过
                
                start_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[0])
                end_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[1])
                screen_events.append(f"Dialogue: 0,{start_ass},{end_ass},画面字,,0,0,0,,{s_text}")
                
        if file in norm_files:
            blocks = parse_srt_file(os.path.join(norm_dir, file))
            for block in blocks:
                n_text = clean_ass_text(block['Text'])
                if not n_text: continue
                
                # === 应用正则到对白字幕 ===
                if replacements and target_mode in ['对白字幕', '全部']:
                    for pat, repl in replacements:
                        n_text = re.sub(pat, repl, n_text)
                if not n_text.strip(): continue
                
                start_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[0])
                end_ass = srt_to_ass_time(block['Timeline'].split(' --> ')[1])
                normal_events.append(f"Dialogue: 0,{start_ass},{end_ass},对白字幕,,0,0,0,,{n_text}")

        ass_events_str = "[Events]\nFormat: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text\n"
        ass_events_str += "\n".join(screen_events) + ("\n" if screen_events else "")
        ass_events_str += "\n".join(normal_events) + "\n"
        
        out_path = os.path.join(out_dir, file.rsplit('.', 1)[0] + '.ass')
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(f"{srt_script_info}\n{srt_styles_block}\n\n{ass_events_str}")
            
    return len(all_files)

def process_ass_split(in_dir, out_scr_dir, out_norm_dir, logic_mode, use_c1, bracket_str, use_c2, sel_effs, use_c3, sel_styles, to_srt=False):
    """根据组合条件将 ASS 拆分为画面字和普通字两个文件"""
    files = [f for f in os.listdir(in_dir) if f.lower().endswith('.ass')]
    if not files: raise ValueError("输入文件夹中没有找到 .ass 文件！")
    
    os.makedirs(out_scr_dir, exist_ok=True)
    os.makedirs(out_norm_dir, exist_ok=True)
    
    regex_pat = bracket_str.strip() if use_c1 else ""
        
    def ass_to_srt_time(ass_time):
        h, m, s_ms = ass_time.strip().split(':')
        s, ms = s_ms.split('.')
        return f"{int(h):02d}:{int(m):02d}:{int(s):02d},{ms.ljust(3, '0')}"

    def convert_to_srt_blocks(ev_list):
        blocks = []
        idx = 1
        for ev in ev_list:
            if ev.startswith('Dialogue:'):
                parts = ev.split(',', 9)
                if len(parts) >= 10:
                    st = ass_to_srt_time(parts[1])
                    ed = ass_to_srt_time(parts[2])
                    txt = re.sub(r'\{.*?\}', '', parts[9]).replace('\\N', '\n').replace('\\n', '\n')
                    blocks.append(f"{idx}\n{st} --> {ed}\n{txt}\n")
                    idx += 1
        return "\n".join(blocks)
        
    processed_count = 0
    for file in files:
        filepath = os.path.join(in_dir, file)
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            lines = f.read().split('\n')
            
        h_lines, s_lines, ev_lines = [], [], []
        curr = "info"
        for line in lines:
            l = line.strip()
            if l.startswith('[V4+ Styles]'): curr = "styles"
            elif l.startswith('[Events]'): curr = "events"
            
            if curr == "info": h_lines.append(line)
            elif curr == "styles": s_lines.append(line)
            elif curr == "events": ev_lines.append(line)
            
        screen_ev, normal_ev = [], []
        has_screen_dialogue = False  # 新增：标记当前文件是否真的提取到了画面字
        
        for ev in ev_lines:
            if ev.startswith('Dialogue:'):
                parts = ev.split(',', 9)
                if len(parts) >= 10:
                    style = parts[3].strip()
                    effect = parts[8].strip()
                    txt = parts[9]
                    
                    # 接入全局高级判定求值器
                    is_screen = evaluate_advanced_condition("ASS", parts, logic_mode, use_c1, regex_pat, use_c2, sel_effs, use_c3, sel_styles)
                    
                    if is_screen: 
                        screen_ev.append(ev)
                        has_screen_dialogue = True  # 记录：确实找到了符合条件的画面字
                    else: 
                        normal_ev.append(ev)
                else:
                    screen_ev.append(ev)
                    normal_ev.append(ev)
            else:
                screen_ev.append(ev)
                normal_ev.append(ev)
                
        base_name = os.path.splitext(file)[0]
        if to_srt:
            # 只有在找到画面字时，才真正生成并输出画面字的 SRT 文件
            if has_screen_dialogue:
                with open(os.path.join(out_scr_dir, base_name + '.srt'), 'w', encoding='utf-8') as f:
                    f.write(convert_to_srt_blocks(screen_ev))
            # 普通字文件始终输出
            with open(os.path.join(out_norm_dir, base_name + '.srt'), 'w', encoding='utf-8') as f:
                f.write(convert_to_srt_blocks(normal_ev))
        else:
            # 只有在找到画面字时，才真正生成并输出画面字的 ASS 文件
            if has_screen_dialogue:
                with open(os.path.join(out_scr_dir, file), 'w', encoding='utf-8') as f:
                    f.write("\n".join(h_lines) + "\n" + "\n".join(s_lines) + "\n" + "\n".join(screen_ev) + "\n")
            # 普通字文件始终输出
            with open(os.path.join(out_norm_dir, file), 'w', encoding='utf-8') as f:
                f.write("\n".join(h_lines) + "\n" + "\n".join(s_lines) + "\n" + "\n".join(normal_ev) + "\n")
            
        processed_count += 1
    return processed_count
# ================= UI 交互回调 =================
def process_timeline_op(in_dir, out_dir, logic_mode, use_c1, bracket_str, use_c2, sel_effs, use_c3, sel_styles, opt1, opt2, tie_mode, threshold, do_concat, concat_pos, do_del, do_regex, do_final_regex, regex_rules):
    files = [f for f in os.listdir(in_dir) if f.lower().endswith('.ass')]
    if not files: raise ValueError("输入文件夹中没有找到 .ass 文件！")
    os.makedirs(out_dir, exist_ok=True)
    
    regex_pat = bracket_str.strip() if use_c1 else ""
    
    def ass_time_to_ms(time_str):
        parts = time_str.strip().split(':')
        h = int(parts[0])
        m = int(parts[1])
        s, ms = parts[2].split('.')
        return (h * 3600 + m * 60 + int(s)) * 1000 + int(ms) * 10

    processed_count = 0
    for file in files:
        filepath = os.path.join(in_dir, file)
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            lines = f.read().split('\n')
            
        h_lines, s_lines, ev_lines = [], [], []
        curr = "info"
        for line in lines:
            l = line.strip()
            if l.startswith('[V4+ Styles]'): curr = "styles"
            elif l.startswith('[Events]'): curr = "events"
            
            if curr == "info": h_lines.append(line)
            elif curr == "styles": s_lines.append(line)
            elif curr == "events": ev_lines.append(line)
            
        st_events, dt_events, parsed_order = [], [], []

        # 接入全局高级判定求值器
        for file_idx, ev in enumerate(ev_lines):
            if ev.startswith('Dialogue:'):
                parts = ev.split(',', 9)
                if len(parts) >= 10:
                    is_screen = evaluate_advanced_condition("ASS", parts, logic_mode, use_c1, regex_pat, use_c2, sel_effs, use_c3, sel_styles)
                    
                    st_ms = ass_time_to_ms(parts[1])
                    ed_ms = ass_time_to_ms(parts[2])
                    txt = parts[9]
                    
                    if is_screen:
                        # 画面字：新增记录原始出现时间 st_ms 和 原文件行号 file_idx，用于后续的绝对时间排序
                        obj = {"parts": list(parts), "st": st_ms, "ed": ed_ms, "orig_txt": txt, "deleted": False, "orig_idx": file_idx}
                        st_events.append(obj)
                    else:
                        # 对白字：新增 matched_texts 列表，用于安全挂载所有跟它重叠的画面字
                        obj = {"parts": list(parts), "st": st_ms, "ed": ed_ms, "orig_txt": txt, "deleted": False, "matched_texts": []}
                        dt_events.append(obj)
                        
                    parsed_order.append([obj])
                else:
                    parsed_order.append([ev])
            else:
                parsed_order.append([ev])

        # 选项2：重叠时间调整与合并
        if opt2:
            for st in st_events:
                best_dt, max_overlap, best_dt_st = None, -1, -1

                # 【满足要求2】完全解耦的重叠扫描：画面字之间如何重叠无所谓，只找重叠最长的那个对白
                for dt in dt_events:
                    overlap = min(st['ed'], dt['ed']) - max(st['st'], dt['st'])
                    if overlap > threshold:
                        if overlap > max_overlap:
                            max_overlap, best_dt, best_dt_st = overlap, dt, dt['st']
                        elif overlap == max_overlap:
                            # 发生平局，采用打破平局策略
                            if tie_mode == 0 and dt['st'] < best_dt_st:
                                best_dt, best_dt_st = dt, dt['st']
                            elif tie_mode == 1 and dt['st'] > best_dt_st:
                                best_dt, best_dt_st = dt, dt['st']
                
                # 如果找到了最佳对齐对白字幕
                if best_dt:
                    st['parts'][1] = best_dt['parts'][1]
                    st['parts'][2] = best_dt['parts'][2]
                    
                    if do_concat:
                        txt_to_add = st['orig_txt']
                        
                        # 【满足要求3】合并前，在此处对每一个画面字进行“单独”的正则替换
                        if do_regex:
                            for pat, repl in regex_rules:
                                txt_to_add = re.sub(pat, repl, txt_to_add)
                        
                        # 不直接傻瓜拼接，而是打包暂存到目标的列表中，保留它的原始时间信息
                        best_dt['matched_texts'].append({
                            "txt": txt_to_add,
                            "time": st['st'],      # 画面字的初始出现时间
                            "idx": st['orig_idx']  # 画面字在文件中的行号
                        })
                        
                        if do_del: st['deleted'] = True
                    else:
                        st['st'] = best_dt['st']
                        st['ed'] = best_dt['ed']

        # 【满足要求1】将缓冲好的拼接文本正式应用，并绝对保证按首次出现顺序
        if opt2 and do_concat:
            for dt in dt_events:
                if dt['matched_texts']:
                    # 核心排序魔法：优先根据画面字的初始时间戳排序；时间戳一样时，按文件原有的上下先后排序
                    dt['matched_texts'].sort(key=lambda x: (x['time'], x['idx']))
                    
                    combined_add_txt = "".join([x['txt'] for x in dt['matched_texts']])
                    
                    if concat_pos == 0:
                        dt['parts'][9] = combined_add_txt + dt['parts'][9]
                    else:
                        dt['parts'][9] = dt['parts'][9] + combined_add_txt

        final_events = []
        if opt1:
            # 选项1开启：先排头部结构/注释，再排画面字，最后排对白
            for group in parsed_order:
                for item in group:
                    if isinstance(item, str): final_events.append(item)
            for st in st_events:
                if not st['deleted']: final_events.append(",".join(st['parts']))
            for dt in dt_events:
                if not dt['deleted']: final_events.append(",".join(dt['parts']))
        else:
            # 不开启重排序：严格遵循原位组装
            for group in parsed_order:
                for item in group:
                    if isinstance(item, str): final_events.append(item)
                    elif not item['deleted']: final_events.append(",".join(item['parts']))

        # ====== 新增：在所有合并、排序操作完成后，进行最终的全局正则替换 ======
        if do_final_regex and regex_rules:
            for i in range(len(final_events)):
                if final_events[i].startswith('Dialogue:'):
                    parts = final_events[i].split(',', 9)
                    if len(parts) >= 10:
                        for pat, repl in regex_rules:
                            parts[9] = re.sub(pat, repl, parts[9])
                        final_events[i] = ",".join(parts)
        # ======================================================================
                
        with open(os.path.join(out_dir, file), 'w', encoding='utf-8') as f:
            f.write("\n".join(h_lines) + "\n" + "\n".join(s_lines) + "\n" + "\n".join(final_events) + "\n")
            
        processed_count += 1
    return processed_count

def run_ass_merge():
    dir1 = am_dir1_var.get().strip()
    dir2 = am_dir2_var.get().strip()
    out_dir = am_out_var.get().strip()
    
    if not dir1 or not dir2 or not out_dir:
        return messagebox.showwarning("警告", "请完整选择两个输入文件夹和输出文件夹！")
        
    try:
        count = process_ass_merge(dir1, dir2, out_dir)
        messagebox.showinfo("完成", f"ASS 合并成功！\n共合并了 {count} 个文件。")
    except Exception as e:
        messagebox.showerror("错误", f"合并失败:\n{str(e)}")

def run_time_split():
    in_dir = ts_in_var.get().strip()
    out_norm = ts_norm_var.get().strip()
    out_scr = ts_scr_var.get().strip()
    mode = ts_mode_var.get()
    
    if not in_dir or not out_norm or not out_scr:
        return messagebox.showwarning("警告", "请完整填写输入和输出文件夹！")
        
    try:
        count = process_time_split(in_dir, out_norm, out_scr, mode)
        messagebox.showinfo("完成", f"按时间轴拆分成功！\n共自动处理 {count} 个文件。")
    except Exception as e:
        messagebox.showerror("错误", f"拆分失败:\n{str(e)}")

def run_split():
    infile = split_in_var.get().strip()
    out_src = split_out_src_var.get().strip()
    out_tgt = split_out_tgt_var.get().strip()
    
    parts = [p.strip() for p in split_cols_var.get().strip().replace('，', ',').split(',') if p.strip()]
    if not infile or not out_src: return messagebox.showwarning("警告", "请至少填写输入文件和源语言输出目录！")
    
    if split_mode_var.get() == 1:
        if len(parts) < 4: return messagebox.showwarning("警告", "模式1 需要至少 4 个列名字母 (文件, ID, 时间, 内容)！")
        try:
            count = process_split(infile, out_src, out_tgt, parts)
            messagebox.showinfo("完成", f"成功拆分 {count} 个视频文件！")
        except Exception as e: messagebox.showerror("错误", f"拆分失败:\n{str(e)}")
    else:
        if len(parts) < 3: return messagebox.showwarning("警告", "模式2 需要至少 3 个列名字母 (文件, 时间, 内容)！")
        try:
            count = process_split_mode2(infile, out_src, out_tgt, parts, split_sheet_var.get().strip())
            messagebox.showinfo("完成", f"成功拆分 {count} 个视频文件！")
        except Exception as e: messagebox.showerror("错误", f"拆分失败:\n{str(e)}")

def run_merge():
    src_dir, tgt_dir, out_xls = merge_src_var.get().strip(), merge_tgt_var.get().strip(), merge_out_var.get().strip()
    src_name, tgt_name = merge_src_name_var.get().strip(), merge_tgt_name_var.get().strip()
    
    if not src_name: return messagebox.showerror("错误", "请输入源语言列名！")
    if not src_dir or not out_xls: return messagebox.showwarning("警告", "请完整填写源目录和输出文件路径！")
    
    if merge_mode_var.get() == 1:
        if not tgt_dir or not tgt_name: return messagebox.showwarning("警告", "模式1(双语对照)下，目标语言目录和列名必填！")
        try:
            count = process_merge(src_dir, tgt_dir, src_name, tgt_name, out_xls)
            messagebox.showinfo("完成", f"成功合并 {count} 个 SRT 文件到 Excel！")
        except Exception as e: messagebox.showerror("错误", f"合并失败:\n{str(e)}")
    else:
        try:
            count = process_merge_mode2(src_dir, tgt_dir, src_name, tgt_name, out_xls)
            messagebox.showinfo("完成", f"成功按模式2合并了 {count} 个 SRT 文件到 Excel！")
        except Exception as e: messagebox.showerror("错误", f"合并失败:\n{str(e)}")

def run_replace():
    report_file, srt_dir, out_summary = rep_report_var.get().strip(), rep_srt_var.get().strip(), rep_out_var.get().strip()
    parts = [p.strip() for p in rep_cols_var.get().strip().replace('，', ',').split(',')]
    if not report_file or not srt_dir or not out_summary: return messagebox.showwarning("警告", "请完整选择文件路径！")
    if len(parts) != 3 or not all(parts): return messagebox.showwarning("警告", "列名格式错误！")
    try:
        # === 核心修改：将界面选择的匹配模式传递给底层函数 ===
        mode = rep_match_mode_var.get()
        rep_count, file_count = process_replace(report_file, srt_dir, out_summary, parts[0], parts[1], parts[2], mode)
        messagebox.showinfo("完成", f"替换完毕！\n共影响了 {file_count} 个 SRT 文件。\n合计成功替换 {rep_count} 条字幕，已生成展示表格。")
    except Exception as e: messagebox.showerror("错误", f"替换失败:\n{str(e)}")

def run_zip():
    target_dir, out_dir = zip_target_var.get().strip(), zip_out_var.get().strip()
    if not target_dir or not out_dir: return messagebox.showwarning("警告", "请选择需要打包的文件夹和输出文件夹！")
    try: max_f = int(zip_max_var.get().strip())
    except: return messagebox.showwarning("警告", "最大文件数必须是整数！")
    try:
        total_files, num_zips = process_zip(target_dir, out_dir, max_f)
        messagebox.showinfo("完成", f"打包完成！\n共处理了 {total_files} 个文件，\n生成了 {num_zips} 个分卷压缩包。")
    except Exception as e: messagebox.showerror("错误", f"打包失败:\n{str(e)}")

def run_ass_convert():
    srt_dir, out_dir = ass_srt_var.get().strip(), ass_out_var.get().strip()
    if not srt_dir or not out_dir: return messagebox.showwarning("警告", "请完整选择仅限 SRT 输入文件夹和 ASS 输出文件夹！")
    
    bracket = ass_bracket_var.get().strip()
    # 修改点：如果开关没打开，就强制将 regex_text 置为空字符串，底层就不会执行替换了
    regex_text = ass_regex_text.get("1.0", tk.END) if ass_enable_regex_var.get() == 1 else ""
    do_merge = ass_merge_var.get() == 1
    merge_report_path = ass_merge_report_var.get().strip()
    if do_merge and not merge_report_path: return messagebox.showwarning("警告", "勾选了合并功能，请指定合并报告的保存路径！")
    
    custom_style = {
        "play_res_x": ass_resx_var.get(), "play_res_y": ass_resy_var.get(),
        "n_font": ass_n_font_var.get(), "n_size": ass_n_size_var.get(), "n_color": ass_n_color_var.get(), "n_alpha": ass_n_alpha_var.get(), "n_out_color": ass_n_outcolor_var.get(), "n_out_alpha": ass_n_outalpha_var.get(),
        "n_margin_v": ass_n_marginv_var.get(), "n_margin_lr": ass_n_marginlr_var.get(), "n_outline": ass_n_outline_var.get(),
        "n_align": ass_n_align_var.get(), "n_shadow": ass_n_shadow_var.get(), "n_bold": ass_n_bold_var.get(), "n_italic": ass_n_italic_var.get(),
        "s_font": ass_s_font_var.get(), "s_size": ass_s_size_var.get(), "s_color": ass_s_color_var.get(), "s_alpha": ass_s_alpha_var.get(), "s_out_color": ass_s_outcolor_var.get(), "s_out_alpha": ass_s_outalpha_var.get(),
        "s_margin_v": ass_s_marginv_var.get(), "s_margin_lr": ass_s_marginlr_var.get(), "s_outline": ass_s_outline_var.get(),
        "s_align": ass_s_align_var.get(), "s_shadow": ass_s_shadow_var.get(), "s_bold": ass_s_bold_var.get(), "s_italic": ass_s_italic_var.get()
    }
    
    ref_cfg = None
    if ass_style_mode_5.get() == 1:
        ref_cfg = {
            'ref_path': ass_ref_path_5.get().strip(),
            'n_style': ass_ref_n_var_5.get().strip(),
            's_style': ass_ref_s_var_5.get().strip(),
            'font_mode': ass_ref_font_mode_5.get(),
            'override_font': ass_ref_override_font_5.get().strip()
        }
        if not ref_cfg['ref_path'] or not os.path.exists(ref_cfg['ref_path']):
            return messagebox.showwarning("警告", "请选择有效的外部参考 ASS 文件！")
        if not ref_cfg['n_style'] or not ref_cfg['s_style']:
            return messagebox.showwarning("警告", "请在应用前先扫描参考文件并选择样式！")
            
    try:
        count = process_srt_to_ass(srt_dir, out_dir, bracket, regex_text, custom_style, do_merge, merge_report_path, ass_style_mode_5.get(), ref_cfg)
        messagebox.showinfo("完成", f"转换完成！\n成功处理了 {count} 个 SRT 文件，已全部转为 ASS 并应用样式。")
    except Exception as e: messagebox.showerror("错误", f"转换失败:\n{str(e)}")

def run_column_copy():
    src_dir, tgt_dir = eff_src_var.get().strip(), eff_tgt_var.get().strip()
    out_dir, err_rep = eff_out_var.get().strip(), eff_err_var.get().strip()
    fmt, col_str = eff_fmt_var.get(), eff_col_var.get()
    
    if not src_dir or not tgt_dir or not out_dir: return messagebox.showwarning("警告", "请完整选择目录！")
    
    sel_blocks = []
    if fmt == "ASS" and col_str.startswith("Header"):
        sel_blocks = [lb_eff_headers.get(i) for i in lb_eff_headers.curselection()]
        
    try:
        processed_count, err_count = process_column_copy_batch(src_dir, tgt_dir, out_dir, err_rep, fmt, col_str, sel_blocks)
        
        if col_str.startswith("Header"):
            messagebox.showinfo("完成", f"完美处理 {processed_count} 个文件！所选文件头区块已成功同步覆盖。")
        else:
            if err_count > 0: messagebox.showwarning("部分完成", f"成功提取 {processed_count} 个文件！但检测到 {err_count} 处异常(行数缺失或时间轴错位)，已忽略时间轴强行复用并导出报告。")
            else: messagebox.showinfo("完成", f"完美处理 {processed_count} 个文件！所选列已严格按行数映射全部复制成功。")
    except Exception as e: messagebox.showerror("错误", f"处理失败:\n{str(e)}")

def run_srt_bilingual_split():
    in_d = bi_srt_var.get().strip()
    out_d1, out_d2 = bi_out_dir1_var.get().strip(), bi_out_dir2_var.get().strip()
    s1, s2 = bi_suf1_var.get().strip(), bi_suf2_var.get().strip()
    err_rep = bi_err_rep_var.get().strip()
    s_mode = bi_split_mode_var.get()
    
    if not in_d or not out_d1 or not out_d2: return messagebox.showwarning("警告", "请完整选择输入和两个输出目录！")
    # 彻底移除了对 s1 和 s2 的强制输入校验，允许它们为空字符串
    try:
        # 将 s_mode 传递给底层函数
        file_count, block_count, err_count = process_srt_bilingual_split_batch(in_d, out_d1, out_d2, s1, s2, err_rep, s_mode)
        if err_count > 0:
            msg = f"批量拆分完成！\n共处理 {file_count} 个文件（合计 {block_count} 条字幕）。\n\n⚠️ 发现 {err_count} 处拆分后存在空白行的情况！"
            if err_rep: msg += f"\n详细空白行报错报告已导出至：\n{err_rep}"
            messagebox.showwarning("部分完成", msg)
        else:
            messagebox.showinfo("完成", f"批量拆分成功！\n共处理 {file_count} 个文件（合计 {block_count} 条字幕），两种语言已分别完美导出。")
    except Exception as e: messagebox.showerror("错误", f"拆分失败:\n{str(e)}")

def run_merge_srt_to_ass():
    norm_d, scr_d, out_d = ms_norm_var.get().strip(), ms_scr_var.get().strip(), ms_out_var.get().strip()
    if not out_d or (not norm_d and not scr_d): return messagebox.showwarning("警告", "请至少选择一个输入和输出文件夹！")
    
    custom_style = {
        "play_res_x": m5_resx_var.get(), "play_res_y": m5_resy_var.get(),
        "n_font": m5_n_font_var.get(), "n_size": m5_n_size_var.get(), "n_color": m5_n_color_var.get(), "n_alpha": m5_n_alpha_var.get(), "n_out_color": m5_n_outcolor_var.get(), "n_out_alpha": m5_n_outalpha_var.get(),
        "n_margin_v": m5_n_marginv_var.get(), "n_margin_lr": m5_n_marginlr_var.get(), "n_outline": m5_n_outline_var.get(),
        "n_align": m5_n_align_var.get(), "n_shadow": m5_n_shadow_var.get(), "n_bold": m5_n_bold_var.get(), "n_italic": m5_n_italic_var.get(),
        "s_font": m5_s_font_var.get(), "s_size": m5_s_size_var.get(), "s_color": m5_s_color_var.get(), "s_alpha": m5_s_alpha_var.get(), "s_out_color": m5_s_outcolor_var.get(), "s_out_alpha": m5_s_outalpha_var.get(),
        "s_margin_v": m5_s_marginv_var.get(), "s_margin_lr": m5_s_marginlr_var.get(), "s_outline": m5_s_outline_var.get(),
        "s_align": m5_s_align_var.get(), "s_shadow": m5_s_shadow_var.get(), "s_bold": m5_s_bold_var.get(), "s_italic": m5_s_italic_var.get()
    }
    
    ref_cfg = None
    if ms_style_mode_9.get() == 1:
        ref_cfg = {
            'ref_path': ms_ref_path_9.get().strip(),
            'n_style': ms_ref_n_var_9.get().strip(),
            's_style': ms_ref_s_var_9.get().strip(),
            'font_mode': ms_ref_font_mode_9.get(),
            'override_font': ms_ref_override_font_9.get().strip()
        }
        if not ref_cfg['ref_path'] or not os.path.exists(ref_cfg['ref_path']):
            return messagebox.showwarning("警告", "请选择有效的外部参考 ASS 文件！")
        if not ref_cfg['n_style'] or not ref_cfg['s_style']:
            return messagebox.showwarning("警告", "请在应用前先扫描参考文件并选择样式！")
            
    regex_cfg = {
        'enable': ms_enable_regex_var.get() == 1,
        'target': ms_regex_target_var.get().strip(),
        'text': ms_regex_text.get("1.0", tk.END)
    }
            
    try:
        count = process_merge_srt_to_ass_batch(norm_d, scr_d, out_d, custom_style, ms_style_mode_9.get(), ref_cfg, regex_cfg)
        messagebox.showinfo("完成", f"合并转换完成！\n成功将 {count} 个双源 SRT 合并为了 ASS。")
    except Exception as e: messagebox.showerror("错误", f"处理失败:\n{str(e)}")

def run_ass_split():
    i_d = split_ass_in_var.get().strip()
    o_s = split_ass_out_scr_var.get().strip()
    o_n = split_ass_out_norm_var.get().strip()
    
    if not i_d or not o_s or not o_n: return messagebox.showwarning("警告", "请完整选择输入和两个输出文件夹！")
        
    logic_mode = split_ass_logic_var.get()
    u1, u2, u3 = split_ass_c1_var.get() == 1, split_ass_c2_var.get() == 1, split_ass_c3_var.get() == 1
    b_str = split_ass_bracket_var.get().strip()
    sel_effs = [lb_split_effs.get(i) for i in lb_split_effs.curselection()]
    sel_styles = [lb_split_styles.get(i) for i in lb_split_styles.curselection()]
    to_srt = split_ass_to_srt_var.get() == 1
    
    if not (u1 or u2 or u3): return messagebox.showwarning("警告", "请至少勾选一个拆分条件！")
    if u2 and not sel_effs: return messagebox.showwarning("警告", "勾选了特效条件，但未在列表中选中任何特效！")
    if u3 and not sel_styles: return messagebox.showwarning("警告", "勾选了样式条件，但未在列表中选中任何样式！")
    
    try:
        count = process_ass_split(i_d, o_s, o_n, logic_mode, u1, b_str, u2, sel_effs, u3, sel_styles, to_srt)
        messagebox.showinfo("完成", f"拆分成功！\n共处理了 {count} 个 ASS 文件，已分别输出。")
    except Exception as e: messagebox.showerror("错误", f"拆分失败:\n{str(e)}")

# ----- 预设通用操作 -----
def action_save_preset(var, current_list, combobox, filepath, min_len):
    val = var.get().strip().replace('，', ',')
    parts = [p.strip().upper() for p in val.split(',') if p.strip()]
    if len(parts) < min_len: return messagebox.showwarning("格式不规范", f"请输入标准的 {min_len} 个字母格式后再保存")
    val_clean = ", ".join(parts)
    if val_clean not in current_list:
        current_list.append(val_clean)
        combobox['values'] = current_list
        var.set(val_clean)
        save_presets_to_file(filepath, current_list)
        messagebox.showinfo("提示", "预设保存成功！")
    else: messagebox.showinfo("提示", "该预设已存在！")

def action_del_preset(var, current_list, combobox, filepath):
    val = var.get().strip()
    if val in current_list:
        current_list.remove(val)
        combobox['values'] = current_list
        var.set(current_list[0] if current_list else "")
        save_presets_to_file(filepath, current_list)
        messagebox.showinfo("提示", "预设已删除！")
    else: messagebox.showwarning("提示", "列表中没有此预设，无法删除。")

def choose_color(var, btn):
    # 动态获取当前触发按钮所在的顶层窗口
    parent_win = btn.winfo_toplevel()
    
    # 强制将颜色面板绑定到该子窗口上
    c = colorchooser.askcolor(title="选择颜色", initialcolor=var.get() or "#FFFFFF", parent=parent_win)
    if c[1]:
        var.set(c[1].upper())
        btn.config(bg=c[1].upper())
        
    # 选完颜色后，强制把这个子窗口拉回最顶层并恢复输入焦点
    parent_win.lift()
    parent_win.focus_set()

def update_color_btn(btn, color_hex):
    try: btn.config(bg=color_hex)
    except: pass

def ask_file(var, title, filetypes): var.set(filedialog.askopenfilename(title=title, filetypes=filetypes))
def ask_dir(var, title): var.set(filedialog.askdirectory(title=title))
def ask_save_file(var, title, filetypes, defaultextension): var.set(filedialog.asksaveasfilename(title=title, filetypes=filetypes, defaultextension=defaultextension))


def switch_category():
    cat = category_var.get()
    nb_srt.pack_forget()
    nb_ass.pack_forget()
    nb_other.pack_forget()
    
    if cat == "SRT": nb_srt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    elif cat == "ASS": nb_ass.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    else: nb_other.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

def scan_sheets():
    f = split_in_var.get().strip()
    if not f or not os.path.exists(f): return messagebox.showwarning("警告", "请先选择有效的Excel文件！")
    try:
        xl = pd.ExcelFile(f)
        cb_sheet['values'] = xl.sheet_names
        if xl.sheet_names: split_sheet_var.set(xl.sheet_names[0])
        messagebox.showinfo("成功", f"扫描到 {len(xl.sheet_names)} 个 Sheet")
    except Exception as e: messagebox.showerror("错误", str(e))

# ================= GUI 界面构建 =================

root = tk.Tk()
root.title("字幕拆分与合并工具箱")
root.geometry("770x670")
root.minsize(500, 500)

# 跨平台主题与字体自适应
os_name = platform.system()
if os_name == 'Darwin':  # macOS 环境
    GLOBAL_FONT = 'PingFang SC'
    GLOBAL_FONT_TUPLE = (GLOBAL_FONT, 12) # Mac 渲染策略不同，字体需稍大
else:  # Windows / Linux 环境
    GLOBAL_FONT = 'Microsoft YaHei'
    GLOBAL_FONT_TUPLE = (GLOBAL_FONT, 10)

style = ttk.Style(root)
themes = style.theme_names()
if os_name == 'Darwin' and 'aqua' in themes:
    style.theme_use('aqua') # 激活原生 macOS 拟物化水晶主题
elif 'vista' in themes: 
    style.theme_use('vista')
elif 'winnative' in themes: 
    style.theme_use('winnative')

style.configure('TButton', font=GLOBAL_FONT_TUPLE)

# --- 新增：顶部导航分类栏 ---
nav_frame = tk.Frame(root)
nav_frame.pack(fill=tk.X, padx=10, pady=(10, 0))

category_var = tk.StringVar(value="SRT")

ttk.Radiobutton(nav_frame, text=" 📝 基本功能 ", variable=category_var, value="SRT", command=switch_category, style='Toolbutton').pack(side=tk.LEFT, padx=(0, 5), ipadx=10, ipady=3)
ttk.Radiobutton(nav_frame, text=" 🎬 高级功能 ", variable=category_var, value="ASS", command=switch_category, style='Toolbutton').pack(side=tk.LEFT, padx=5, ipadx=10, ipady=3)
ttk.Radiobutton(nav_frame, text=" 🛠️ 其他功能 ", variable=category_var, value="OTHER", command=switch_category, style='Toolbutton').pack(side=tk.LEFT, padx=5, ipadx=10, ipady=3)

nb_srt = ttk.Notebook(root)
nb_ass = ttk.Notebook(root)
nb_other = ttk.Notebook(root)

nb_srt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
# ================= TAB 1: 拆分 =================
tab_split = ttk.Frame(nb_srt, padding=20)
nb_srt.add(tab_split, text=" XLSX 拆分为 SRT ")
tab_split.columnconfigure(1, weight=1)
current_presets_split = load_presets(PRESET_FILE_SPLIT, DEFAULT_PRESETS_SPLIT)
split_in_var = tk.StringVar()
split_out_src_var, split_out_tgt_var = tk.StringVar(), tk.StringVar()
split_cols_var = tk.StringVar(value=current_presets_split[0] if current_presets_split else "A, B, C, D")

split_mode_var = tk.IntVar(value=1)
split_sheet_var = tk.StringVar()

f_mode = ttk.Frame(tab_split)
f_mode.grid(row=0, column=0, columnspan=3, sticky="w", pady=5)
ttk.Radiobutton(f_mode, text="模式1: 番茄导出格式", variable=split_mode_var, value=1, command=lambda: update_split_mode()).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_mode, text="模式2: CPP格式", variable=split_mode_var, value=2, command=lambda: update_split_mode()).pack(side=tk.LEFT, padx=5)


btn_scan_sheet = ttk.Button(f_mode, text="扫描Sheet", command=scan_sheets, state="disabled")
btn_scan_sheet.pack(side=tk.LEFT, padx=5)
cb_sheet = ttk.Combobox(f_mode, textvariable=split_sheet_var, state="disabled", width=15)
cb_sheet.pack(side=tk.LEFT, padx=5)

ttk.Label(tab_split, text="双语 XLSX/CSV 文件:").grid(row=1, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_split, textvariable=split_in_var).grid(row=1, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_split, text="浏览...", command=lambda: ask_file(split_in_var, "选择文件", [("Excel", "*.xlsx"), ("CSV", "*.csv")])).grid(row=1, column=2, padx=(5,0), pady=10)

f_s = ttk.Frame(tab_split)
f_s.grid(row=2, column=0, columnspan=3, sticky="w", pady=5)
lbl_split_hint = ttk.Label(f_s, text="列名 (模式1: 文件, ID, 时间, 内容):")
lbl_split_hint.pack(side=tk.LEFT, padx=(0,5))
cb_s = ttk.Combobox(f_s, textvariable=split_cols_var, values=current_presets_split, width=15)
cb_s.pack(side=tk.LEFT, padx=(0, 10))
ttk.Button(f_s, text="保存预设", command=lambda: action_save_preset(split_cols_var, current_presets_split, cb_s, PRESET_FILE_SPLIT, 3)).pack(side=tk.LEFT, padx=5)
ttk.Button(f_s, text="删除预设", command=lambda: action_del_preset(split_cols_var, current_presets_split, cb_s, PRESET_FILE_SPLIT)).pack(side=tk.LEFT, padx=5)

def update_split_mode():
    if split_mode_var.get() == 1:
        cb_sheet.config(state="disabled")
        btn_scan_sheet.config(state="disabled")
        lbl_split_hint.config(text="列名 (模式1: 文件名, ID, 时间, 内容1, (内容2)):")
        split_cols_var.set("A, B, C, D")
    else:
        cb_sheet.config(state="readonly")
        btn_scan_sheet.config(state="normal")
        lbl_split_hint.config(text="列名 (模式2: 文件名, 时间, 内容1, (内容2)):")
        split_cols_var.set("A, B, C, D")

ttk.Label(tab_split, text="源语言 SRT 输出目录:").grid(row=3, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_split, textvariable=split_out_src_var).grid(row=3, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_split, text="浏览...", command=lambda: ask_dir(split_out_src_var, "选择目录")).grid(row=3, column=2, padx=(5,0), pady=10)

ttk.Label(tab_split, text="目标语言 SRT 输出目录(可选):").grid(row=4, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_split, textvariable=split_out_tgt_var).grid(row=4, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_split, text="浏览...", command=lambda: ask_dir(split_out_tgt_var, "选择目录")).grid(row=4, column=2, padx=(5,0), pady=10)

ttk.Button(tab_split, text="开始拆分", command=run_split, style='TButton').grid(row=5, column=0, columnspan=3, pady=20, ipadx=20, ipady=5)

# ================= TAB 2: 合并 =================
tab_merge = ttk.Frame(nb_srt, padding=20)
nb_srt.add(tab_merge, text=" SRT 合并为 XLSX ")
tab_merge.columnconfigure(1, weight=1)

merge_mode_var = tk.IntVar(value=1)
merge_src_var, merge_tgt_var, merge_out_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
merge_src_name_var, merge_tgt_name_var = tk.StringVar(), tk.StringVar()

f_merge_mode = ttk.Frame(tab_merge)
f_merge_mode.grid(row=0, column=0, columnspan=3, sticky="w", pady=5)
ttk.Radiobutton(f_merge_mode, text="模式1: 番茄导出格式", variable=merge_mode_var, value=1).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_merge_mode, text="模式2: CPP格式", variable=merge_mode_var, value=2).pack(side=tk.LEFT, padx=5)

ttk.Label(tab_merge, text="源语言 SRT 目录:").grid(row=1, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_merge, textvariable=merge_src_var).grid(row=1, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_merge, text="浏览...", command=lambda: ask_dir(merge_src_var, "选择目录")).grid(row=1, column=2, padx=(5,0), pady=10)

ttk.Label(tab_merge, text="目标语言 SRT 目录(可选):").grid(row=2, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_merge, textvariable=merge_tgt_var).grid(row=2, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_merge, text="浏览...", command=lambda: ask_dir(merge_tgt_var, "选择目录")).grid(row=2, column=2, padx=(5,0), pady=10)

f_m = ttk.Frame(tab_merge)
f_m.grid(row=3, column=1, sticky="w", pady=5)
ttk.Label(f_m, text="源列名:").pack(side=tk.LEFT)
ttk.Entry(f_m, textvariable=merge_src_name_var, width=10).pack(side=tk.LEFT, padx=(5,15))
ttk.Label(f_m, text="目标列名:").pack(side=tk.LEFT)
ttk.Entry(f_m, textvariable=merge_tgt_name_var, width=10).pack(side=tk.LEFT, padx=(5,0))

ttk.Label(tab_merge, text="保存为 XLSX 文件:").grid(row=4, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_merge, textvariable=merge_out_var).grid(row=4, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_merge, text="浏览...", command=lambda: ask_save_file(merge_out_var, "保存", [("Excel", "*.xlsx")], ".xlsx")).grid(row=4, column=2, padx=(5,0), pady=10)
ttk.Button(tab_merge, text="开始合并", command=run_merge, style='TButton').grid(row=5, column=0, columnspan=3, pady=20, ipadx=20, ipady=5)

# ================= TAB 3: 替换 =================
tab_rep = ttk.Frame(nb_srt, padding=20)
nb_srt.add(tab_rep, text=" 根据报告修改 SRT ")
tab_rep.columnconfigure(1, weight=1)
current_presets_rep = load_presets(PRESET_FILE_REP, DEFAULT_PRESETS_REP)
rep_report_var, rep_srt_var, rep_out_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
rep_cols_var = tk.StringVar(value=current_presets_rep[0] if current_presets_rep else "A, B, E")

# ====== 新增：匹配模式变量 ======
rep_match_mode_var = tk.IntVar(value=0)

ttk.Label(tab_rep, text="QA 报告 (Excel/CSV):").grid(row=0, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_rep, textvariable=rep_report_var).grid(row=0, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_rep, text="浏览...", command=lambda: ask_file(rep_report_var, "选择文件", [("Excel", "*.xlsx"), ("CSV", "*.csv")])).grid(row=0, column=2, padx=(5,0), pady=10)

# ====== 新增：匹配模式选择 UI ======
f_rep_mode = ttk.Frame(tab_rep)
f_rep_mode.grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 5), padx=20)
ttk.Radiobutton(f_rep_mode, text="根据【字幕ID】匹配替换", variable=rep_match_mode_var, value=0).pack(side=tk.LEFT, padx=(0, 15))
ttk.Radiobutton(f_rep_mode, text="根据【时间轴】匹配替换 (智能忽略空格与逗号差异)", variable=rep_match_mode_var, value=1).pack(side=tk.LEFT)

f_r = ttk.Frame(tab_rep)
f_r.grid(row=2, column=0, columnspan=3, sticky="w", pady=5, padx=20)
ttk.Label(f_r, text="列名 (文件, ID/时间轴, 内容):").pack(side=tk.LEFT, padx=(0,5))
cb_r = ttk.Combobox(f_r, textvariable=rep_cols_var, values=current_presets_rep, width=15)
cb_r.pack(side=tk.LEFT, padx=(0, 10))
ttk.Button(f_r, text="保存预设", command=lambda: action_save_preset(rep_cols_var, current_presets_rep, cb_r, PRESET_FILE_REP, 3)).pack(side=tk.LEFT, padx=5)
ttk.Button(f_r, text="删除预设", command=lambda: action_del_preset(rep_cols_var, current_presets_rep, cb_r, PRESET_FILE_REP)).pack(side=tk.LEFT, padx=5)

ttk.Label(tab_rep, text="需修改的 SRT 文件夹:").grid(row=3, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_rep, textvariable=rep_srt_var).grid(row=3, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_rep, text="浏览...", command=lambda: ask_dir(rep_srt_var, "选择目录")).grid(row=3, column=2, padx=(5,0), pady=10)

ttk.Label(tab_rep, text="保存替换展示表格:").grid(row=4, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_rep, textvariable=rep_out_var).grid(row=4, column=1, sticky="ew", padx=5, pady=10)
ttk.Button(tab_rep, text="浏览...", command=lambda: ask_save_file(rep_out_var, "保存", [("Excel", "*.xlsx")], ".xlsx")).grid(row=4, column=2, padx=(5,0), pady=10)

ttk.Button(tab_rep, text="开始替换", command=run_replace, style='TButton').grid(row=5, column=0, columnspan=3, pady=15, ipadx=20, ipady=5)

# ================= TAB 8: 双语 SRT 批量拆分 =================
tab_bi = ttk.Frame(nb_srt, padding=20)
nb_srt.add(tab_bi, text=" 双语 SRT 批量拆分 ")
tab_bi.columnconfigure(1, weight=1)

bi_srt_var = tk.StringVar()
bi_out_dir1_var, bi_out_dir2_var = tk.StringVar(), tk.StringVar()
bi_suf1_var, bi_suf2_var = tk.StringVar(value="语言1"), tk.StringVar(value="语言2")
bi_err_rep_var = tk.StringVar() # 新增：报告路径变量
bi_split_mode_var = tk.IntVar(value=1) # 新增：默认选中模式1

ttk.Label(tab_bi, text="双语 SRT 输入文件夹:").grid(row=0, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_srt_var).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Button(tab_bi, text="浏览...", command=lambda: ask_dir(bi_srt_var, "选择输入文件夹")).grid(row=0, column=2, padx=5)

ttk.Label(tab_bi, text="【上方语言】保存目录:").grid(row=1, column=0, sticky="e", pady=5, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_out_dir1_var).grid(row=1, column=1, sticky="ew", padx=5)
ttk.Button(tab_bi, text="浏览...", command=lambda: ask_dir(bi_out_dir1_var, "选择上方语言输出目录")).grid(row=1, column=2, padx=5)

ttk.Label(tab_bi, text="上方语言文件后缀:").grid(row=2, column=0, sticky="e", pady=5, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_suf1_var, width=15).grid(row=2, column=1, sticky="w", padx=5)

ttk.Label(tab_bi, text="【下方语言】保存目录:").grid(row=3, column=0, sticky="e", pady=5, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_out_dir2_var).grid(row=3, column=1, sticky="ew", padx=5)
ttk.Button(tab_bi, text="浏览...", command=lambda: ask_dir(bi_out_dir2_var, "选择下方语言输出目录")).grid(row=3, column=2, padx=5)

ttk.Label(tab_bi, text="下方语言文件后缀:").grid(row=4, column=0, sticky="e", pady=5, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_suf2_var, width=15).grid(row=4, column=1, sticky="w", padx=5)

# 新增：空白行报告导出选项
ttk.Label(tab_bi, text="空白异常报告保存至(可选):").grid(row=5, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_bi, textvariable=bi_err_rep_var).grid(row=5, column=1, sticky="ew", padx=5)
ttk.Button(tab_bi, text="浏览...", command=lambda: ask_save_file(bi_err_rep_var, "保存异常报告", [("Excel", "*.xlsx")], ".xlsx")).grid(row=5, column=2, padx=5)

# ====== 新增：拆分模式选择区域 ======
f_bi_mode = ttk.LabelFrame(tab_bi, text="3行字幕拆分策略 (语言识别引擎)", padding=10)
f_bi_mode.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(15, 5))
ttk.Radiobutton(f_bi_mode, text="经典模式 (基于中日韩区块/纯英文字母)", variable=bi_split_mode_var, value=1).pack(side=tk.LEFT, padx=10)
ttk.Radiobutton(f_bi_mode, text="强化东亚语言模式 (识别假名/谚文)", variable=bi_split_mode_var, value=2).pack(side=tk.LEFT, padx=10)
# ===================================

ttk.Button(tab_bi, text="开始批量拆分双语", command=run_srt_bilingual_split, style='TButton').grid(row=7, column=0, columnspan=3, pady=10, ipadx=20, ipady=5)
# ================= TAB 10: 自动时间轴拆分 =================
tab_ts = ttk.Frame(nb_srt, padding=20)
nb_srt.add(tab_ts, text=" 自动时间轴拆分 ")
tab_ts.columnconfigure(1, weight=1)

ts_in_var = tk.StringVar()
ts_norm_var = tk.StringVar()
ts_scr_var = tk.StringVar()
ts_mode_var = tk.StringVar(value="SRT")

ttk.Label(tab_ts, text="输入文件夹:").grid(row=0, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_ts, textvariable=ts_in_var).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Button(tab_ts, text="浏览...", command=lambda: ask_dir(ts_in_var, "选择输入文件夹")).grid(row=0, column=2, padx=5)

f_ts_mode = ttk.Frame(tab_ts)
f_ts_mode.grid(row=1, column=0, columnspan=3, sticky="w", pady=5)
ttk.Radiobutton(f_ts_mode, text="处理 SRT", variable=ts_mode_var, value="SRT").pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_ts_mode, text="处理 ASS", variable=ts_mode_var, value="ASS").pack(side=tk.LEFT, padx=5)

ttk.Label(tab_ts, text="对白输出目录:").grid(row=2, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_ts, textvariable=ts_norm_var).grid(row=2, column=1, sticky="ew", padx=5)
ttk.Button(tab_ts, text="浏览...", command=lambda: ask_dir(ts_norm_var, "选择输出目录")).grid(row=2, column=2, padx=5)

ttk.Label(tab_ts, text="画面字输出目录:").grid(row=3, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_ts, textvariable=ts_scr_var).grid(row=3, column=1, sticky="ew", padx=5)
ttk.Button(tab_ts, text="浏览...", command=lambda: ask_dir(ts_scr_var, "选择输出目录")).grid(row=3, column=2, padx=5)

ttk.Label(tab_ts, text="* 智能侦测：工具会逐行扫描文件，一旦发现某行字幕的【开始时间】早于上一行的【结束时间】，\n即自动判定该行为分界点。该行及下方划为对白字幕，上方划为画面字。", foreground="gray").grid(row=4, column=0, columnspan=3, pady=(5,10), sticky="w")
ttk.Button(tab_ts, text="开始自动拆分", command=run_time_split, style='TButton').grid(row=5, column=0, columnspan=3, pady=10, ipadx=20, ipady=5)

# ================= TAB 12: 时间轴操作 =================
tab_time_op = create_scrollable_tab(nb_srt, " 时间轴操作 ", padding=20)
tab_time_op.columnconfigure(1, weight=1)

time_op_in_var, time_op_out_var = tk.StringVar(), tk.StringVar()
time_op_bracket_var = tk.StringVar(value="^\\[")

ttk.Label(tab_time_op, text="ASS 输入文件夹:").grid(row=0, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_time_op, textvariable=time_op_in_var).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Button(tab_time_op, text="浏览...", command=lambda: ask_dir(time_op_in_var, "选择目录")).grid(row=0, column=2, padx=5)

ttk.Label(tab_time_op, text="ASS 输出文件夹:").grid(row=1, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_time_op, textvariable=time_op_out_var).grid(row=1, column=1, sticky="ew", padx=5)
ttk.Button(tab_time_op, text="浏览...", command=lambda: ask_dir(time_op_out_var, "选择目录")).grid(row=1, column=2, padx=5)

f_time_cond, time_logic_var, time_c1_var, time_bracket_var, time_c2_var, lb_time_effs, time_c3_var, lb_time_styles = build_advanced_condition_ui(tab_time_op, time_op_in_var, "第一步：画面字判定条件 (仅支持 ASS 格式)")
f_time_cond.grid(row=2, column=0, columnspan=3, sticky="ew", pady=10, padx=5)

f_time_opts = ttk.LabelFrame(tab_time_op, text="第二步：执行操作 (基于上述判定分离出的结果)", padding=10)
f_time_opts.grid(row=3, column=0, columnspan=3, sticky="ew", pady=5, padx=5)

time_opt1_var = tk.IntVar(value=1)
ttk.Checkbutton(f_time_opts, text="选项1：重排序重组 (将所有分离出的画面字统一挪动至上方，对白字幕至下方)", variable=time_opt1_var).pack(anchor="w", pady=2)

time_opt2_var = tk.IntVar(value=1)
ttk.Checkbutton(f_time_opts, text="选项2：调整画面字的重叠时间轴 (对齐至与之重叠时间最长的对白字幕)", variable=time_opt2_var).pack(anchor="w", pady=(10, 2))

f_opt2_sub1 = ttk.Frame(f_time_opts)
f_opt2_sub1.pack(fill=tk.X, padx=20, pady=2)
time_tie_var = tk.IntVar(value=0)
ttk.Label(f_opt2_sub1, text="重叠相同时优先对齐至:").pack(side=tk.LEFT)
ttk.Radiobutton(f_opt2_sub1, text="最早出现的字幕", variable=time_tie_var, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_opt2_sub1, text="最晚出现的字幕", variable=time_tie_var, value=1).pack(side=tk.LEFT, padx=5)

time_thresh_var = tk.StringVar(value="0")
ttk.Label(f_opt2_sub1, text="   |   重叠安全阈值(ms):").pack(side=tk.LEFT, padx=(15, 5))
ttk.Entry(f_opt2_sub1, textvariable=time_thresh_var, width=6).pack(side=tk.LEFT)
ttk.Label(f_opt2_sub1, text="(<=该阈值不调整)", foreground="gray").pack(side=tk.LEFT, padx=5)

time_concat_var = tk.IntVar(value=0)
ttk.Checkbutton(f_time_opts, text="选项2子选项：合并文本 (将调整好时间的画面字，拼接到对齐的对白字幕轨道中)", variable=time_concat_var).pack(anchor="w", pady=(10, 2))

f_opt2_sub2 = ttk.Frame(f_time_opts)
f_opt2_sub2.pack(fill=tk.X, padx=20, pady=2)
time_concat_pos = tk.IntVar(value=0)
ttk.Radiobutton(f_opt2_sub2, text="拼接到首部", variable=time_concat_pos, value=0).pack(side=tk.LEFT)
ttk.Radiobutton(f_opt2_sub2, text="拼接到尾部", variable=time_concat_pos, value=1).pack(side=tk.LEFT, padx=10)
time_concat_del = tk.IntVar(value=1)
ttk.Checkbutton(f_opt2_sub2, text="合并后删除被合并的原独立画面字", variable=time_concat_del).pack(side=tk.LEFT, padx=20)

time_regex_var = tk.IntVar(value=0)

def toggle_final_regex():
    if time_regex_var.get() == 1:
        cb_time_final_regex.config(state="normal")
    else:
        cb_time_final_regex.config(state="disabled")
        time_final_regex_var.set(0) # 父选项取消时，子选项自动取消

ttk.Checkbutton(f_time_opts, text="合并前先对有重叠画面字正则替换:", variable=time_regex_var, command=toggle_final_regex).pack(anchor="w", padx=20, pady=(10, 2))

# 新增：全局后续正则替换选项（初始禁用）
time_final_regex_var = tk.IntVar(value=0)
cb_time_final_regex = ttk.Checkbutton(f_time_opts, text="所有操作后再进行正则替换 (使用下方相同规则)", variable=time_final_regex_var, state="disabled")
cb_time_final_regex.pack(anchor="w", padx=40, pady=(0, 2))

time_regex_text = tk.Text(f_time_opts, height=3, width=70, font=('Arial', 9))
time_regex_text.pack(anchor="w", padx=40, pady=2)
time_regex_text.insert(tk.END, r"""1. 查找指定数量的特定字符，此处为{2}，查找\\N或者\n: ^(?:.*(?:\\N|\n)){2}[\s\S]*$
2. []替换为()并加上换行: \[([^\]]*)\] >>> ($1)\\N
3. 只替换首尾的[]为()：^\[([\s\S]*)\]$ >>> ($1)""")

def run_time_op():
    in_d = time_op_in_var.get().strip()
    out_d = time_op_out_var.get().strip()
    if not in_d or not out_d: return messagebox.showwarning("警告", "请完整选择输入和输出文件夹！")
    
    logic_mode = time_logic_var.get()
    u1 = time_c1_var.get() == 1
    b_str = time_bracket_var.get().strip()
    u2 = time_c2_var.get() == 1
    sel_effs = [lb_time_effs.get(i) for i in lb_time_effs.curselection()]
    u3 = time_c3_var.get() == 1
    sel_styles = [lb_time_styles.get(i) for i in lb_time_styles.curselection()]

    if not (u1 or u2 or u3): return messagebox.showwarning("警告", "请至少勾选一个画面字判定条件！")
    if u2 and not sel_effs: return messagebox.showwarning("警告", "第一步：勾选了特效条件，但未选中任何特效！")
    if u3 and not sel_styles: return messagebox.showwarning("警告", "第一步：勾选了样式条件，但未选中任何样式！")

    opt1 = time_opt1_var.get() == 1
    opt2 = time_opt2_var.get() == 1
    tie_mode = time_tie_var.get()
    try: threshold = float(time_thresh_var.get().strip())
    except: return messagebox.showwarning("警告", "重叠安全阈值必须是有效数字！")
    
    do_concat = time_concat_var.get() == 1
    concat_pos = time_concat_pos.get()
    do_del = time_concat_del.get() == 1
    do_regex = time_regex_var.get() == 1
    do_final_regex = time_final_regex_var.get() == 1  # 新增：获取子选项状态
    
    regex_rules = []
    if do_regex:
        raw_text = time_regex_text.get("1.0", tk.END).split('\n')
        for line in raw_text:
            line_clean = line.strip('\r\n')
            if not line_clean: continue
            
            if '>>>' in line_clean:
                pat, repl = line_clean.split('>>>', 1)
                if pat.endswith(' '): pat = pat[:-1]
                if repl.startswith(' '): repl = repl[1:]
            else:
                pat = line_clean
                repl = ""
                
            repl_python = re.sub(r'\$(\d+)', r'\\\1', repl)
            regex_rules.append((pat, repl_python))
    
    try:
        # 新增：将 do_final_regex 传给底层引擎
        count = process_timeline_op(in_d, out_d, logic_mode, u1, b_str, u2, sel_effs, u3, sel_styles, opt1, opt2, tie_mode, threshold, do_concat, concat_pos, do_del, do_regex, do_final_regex, regex_rules)
        messagebox.showinfo("完成", f"时间轴操作成功！\n共完美处理了 {count} 个 ASS 文件。")
    except Exception as e:
        messagebox.showerror("错误", f"处理失败:\n{str(e)}")

ttk.Button(tab_time_op, text="▶ 执行时间轴操作", command=run_time_op, style='TButton').grid(row=4, column=0, columnspan=3, pady=10, ipadx=20, ipady=5)

# ================= TAB 4: 打包 =================
tab_zip = ttk.Frame(nb_other, padding=20)
nb_other.add(tab_zip, text=" 分部分打包 ")
tab_zip.columnconfigure(1, weight=1)
zip_target_var, zip_out_var, zip_max_var = tk.StringVar(), tk.StringVar(), tk.StringVar(value="50")

ttk.Label(tab_zip, text="需打包的文件夹:").grid(row=0, column=0, sticky="e", padx=(0,10), pady=20)
ttk.Entry(tab_zip, textvariable=zip_target_var).grid(row=0, column=1, sticky="ew", padx=5, pady=20)
ttk.Button(tab_zip, text="浏览...", command=lambda: ask_dir(zip_target_var, "选择")).grid(row=0, column=2, padx=(5,0), pady=20)
ttk.Label(tab_zip, text="单包最大文件数:").grid(row=1, column=0, sticky="e", padx=(0,10), pady=10)
ttk.Entry(tab_zip, textvariable=zip_max_var, width=15).grid(row=1, column=1, sticky="w", padx=5, pady=10)
ttk.Label(tab_zip, text="ZIP 输出文件夹:").grid(row=2, column=0, sticky="e", padx=(0,10), pady=20)
ttk.Entry(tab_zip, textvariable=zip_out_var).grid(row=2, column=1, sticky="ew", padx=5, pady=20)
ttk.Button(tab_zip, text="浏览...", command=lambda: ask_dir(zip_out_var, "选择")).grid(row=2, column=2, padx=(5,0), pady=20)
ttk.Button(tab_zip, text="开始打包", command=run_zip, style='TButton').grid(row=4, column=0, columnspan=3, pady=25, ipadx=20, ipady=5)

# ================= TAB: LQA 智能拼写检查 =================
# 直接调用 subtitle_tool 的通用滚动标签页生成器
tab_lqa = create_scrollable_tab(nb_other, " LQA 拼写检查 ", padding=10)

# 此时传进去的 tab_lqa 已经是一个自带滚动的内层 Frame 了
lqa_tool_instance = LQA_App(tab_lqa)


def build_style_tab(parent, font_v, size_v, col_v, ocol_v, mv_v, mlr_v, out_v, align_v, shad_v, bold_v, ita_v, alpha_v=None, oalpha_v=None):
    container = ttk.Frame(parent)
    # 将 rowspan 改为 2，完美适配全新的极限两行布局
    container.grid(row=0, column=0, rowspan=2, columnspan=8, sticky="ew")
    
    # ---- 第一行：字体排版与颜色 (极致紧凑整合) ----
    r1 = ttk.Frame(container)
    r1.pack(fill=tk.X, pady=(2, 2), anchor="w")
    
    ttk.Label(r1, text="字体:").pack(side=tk.LEFT)
    cb = ttk.Combobox(r1, textvariable=font_v, width=12)
    cb.pack(side=tk.LEFT, padx=(2, 5))
    
    ttk.Label(r1, text="字号:").pack(side=tk.LEFT)
    ttk.Entry(r1, textvariable=size_v, width=3).pack(side=tk.LEFT, padx=(2, 5))
    
    ttk.Checkbutton(r1, text="加粗", variable=bold_v).pack(side=tk.LEFT, padx=(0, 2))
    ttk.Checkbutton(r1, text="斜体", variable=ita_v).pack(side=tk.LEFT, padx=(0, 5))
    
    ttk.Label(r1, text="主色:").pack(side=tk.LEFT)
    c_btn = tk.Button(r1, width=2, bg="#FFFFFF", relief="ridge", command=lambda: choose_color(col_v, c_btn))
    c_btn.pack(side=tk.LEFT, padx=(2, 5))
    if alpha_v is not None:
        ttk.Label(r1, text="透:").pack(side=tk.LEFT)
        ttk.Entry(r1, textvariable=alpha_v, width=2).pack(side=tk.LEFT, padx=(2, 5))
        
    ttk.Label(r1, text="描边色:").pack(side=tk.LEFT)
    oc_btn = tk.Button(r1, width=2, bg="#000000", relief="ridge", command=lambda: choose_color(ocol_v, oc_btn))
    oc_btn.pack(side=tk.LEFT, padx=(2, 5))
    if oalpha_v is not None:
        ttk.Label(r1, text="透:").pack(side=tk.LEFT)
        ttk.Entry(r1, textvariable=oalpha_v, width=2).pack(side=tk.LEFT, padx=(2, 0))

    # ---- 第二行：特效参数与边距调参 (极致紧凑整合) ----
    r2 = ttk.Frame(container)
    r2.pack(fill=tk.X, pady=(2, 5), anchor="w")
    
    ttk.Label(r2, text="粗细:").pack(side=tk.LEFT)
    ttk.Entry(r2, textvariable=out_v, width=2).pack(side=tk.LEFT, padx=(2, 5))
    
    ttk.Label(r2, text="阴影:").pack(side=tk.LEFT)
    ttk.Entry(r2, textvariable=shad_v, width=2).pack(side=tk.LEFT, padx=(2, 5))

    ttk.Label(r2, text="对齐:").pack(side=tk.LEFT)
    ttk.Combobox(r2, textvariable=align_v, values=[str(i) for i in range(1,10)], width=2, state="readonly").pack(side=tk.LEFT, padx=(2, 5))
    
    ttk.Label(r2, text="上下边距:").pack(side=tk.LEFT)
    ttk.Entry(r2, textvariable=mv_v, width=3).pack(side=tk.LEFT, padx=(2, 5))
    
    ttk.Label(r2, text="左右边距:").pack(side=tk.LEFT)
    ttk.Entry(r2, textvariable=mlr_v, width=3).pack(side=tk.LEFT, padx=(2, 5))
    
    # 嵌套 open_visual_adjuster (内部逻辑和绑定完全一致)
    def open_visual_adjuster(parent_win, mv_var, mlr_var, align_var, font_var, size_var, col_var, ocol_var, out_var, bold_var, ita_var, alpha_var, oalpha_var):
        res_win = tk.Toplevel(parent_win)
        res_win.title("输入画板参数")
        res_win.resizable(False, False)
        
        f_res = ttk.Frame(res_win, padding=20)
        f_res.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(f_res, text="宽 (Width):").grid(row=0, column=0, pady=5, padx=5, sticky="e")
        ent_w = ttk.Entry(f_res, width=15); ent_w.insert(0, "1080"); ent_w.grid(row=0, column=1, pady=5, sticky="w")
        ttk.Label(f_res, text="高 (Height):").grid(row=1, column=0, pady=5, padx=5, sticky="e")
        ent_h = ttk.Entry(f_res, width=15); ent_h.insert(0, "1920"); ent_h.grid(row=1, column=1, pady=5, sticky="w")
        
        bg_color_var = tk.StringVar(value="#000000")
        ttk.Label(f_res, text="画板背景色:").grid(row=2, column=0, pady=5, padx=5, sticky="e")
        f_bg = ttk.Frame(f_res)
        f_bg.grid(row=2, column=1, sticky="w", pady=5)
        bg_btn = tk.Button(f_bg, width=4, bg="#000000", command=lambda: choose_color(bg_color_var, bg_btn))
        bg_btn.pack(side=tk.LEFT)
        
        def open_canvas():
            try: w, h = int(ent_w.get()), int(ent_h.get())
            except: return messagebox.showwarning("错误", "请输入有效的数字！")
            
            vid_bg_color = bg_color_var.get() or "#000000"
            res_win.destroy()
            
            cv_win = tk.Toplevel(parent_win)
            cv_win.title("拖拽调整边距 (完美支持透明度与色彩混合预览)")
            
            screen_w = parent_win.winfo_screenwidth()
            screen_h = parent_win.winfo_screenheight()
            max_h = screen_h * 0.75
            max_w = screen_w * 0.75
            scale = min(max_w / w, max_h / h)
            
            cw, ch = int(w * scale), int(h * scale)
            cv_win.resizable(False, False) 
            
            top_frame = tk.Frame(cv_win)
            top_frame.pack(fill=tk.X, pady=10)
            
            canvas = tk.Canvas(cv_win, width=cw, height=ch, bg="#333333", highlightthickness=2, highlightbackground="#555555")
            canvas.pack(padx=20, pady=(0, 20))
            
            rect_id = canvas.create_rectangle(0, 0, cw, ch, fill=vid_bg_color, outline="gray")
            
            def get_val(var, default, is_int=True):
                if not var: return default
                try: 
                    val = var.get()
                    return int(val) if is_int else (val if val else default)
                except: return default
                 
            def blend_hex(fg_hex, bg_hex, alpha_hex):
                try:
                    a = int(alpha_hex, 16)
                    if a < 0: a = 0
                    if a > 255: a = 255
                    opacity = (255 - a) / 255.0

                    fg = fg_hex.lstrip('#')
                    bg = bg_hex.lstrip('#')
                    if len(fg) != 6: fg = "FFFFFF"
                    if len(bg) != 6: bg = "000000"

                    r_fg, g_fg, b_fg = int(fg[0:2], 16), int(fg[2:4], 16), int(fg[4:6], 16)
                    r_bg, g_bg, b_bg = int(bg[0:2], 16), int(bg[2:4], 16), int(bg[4:6], 16)

                    r = int(r_fg * opacity + r_bg * (1 - opacity))
                    g = int(g_fg * opacity + g_bg * (1 - opacity))
                    b = int(b_fg * opacity + b_bg * (1 - opacity))

                    return f"#{r:02X}{g:02X}{b:02X}"
                except:
                    return fg_hex if fg_hex.startswith("#") else "#FFFFFF"

            drag_data = {"mv": get_val(mv_var, 20), "mlr": get_val(mlr_var, 20)}
            text_ids = []
            crosshair_ids = []
            mouse_pos = {"x": cw // 2, "y": ch // 2, "show": False}
            
            def draw_preview():
                for tid in text_ids + crosshair_ids: canvas.delete(tid)
                text_ids.clear()
                crosshair_ids.clear()
                
                align = get_val(align_var, 2)
                mv, mlr = drag_data["mv"], drag_data["mlr"]
                
                try:
                    f_name = get_val(font_var, GLOBAL_FONT, False)
                except NameError:
                    f_name = get_val(font_var, "Microsoft YaHei", False)
                f_size = max(8, int(get_val(size_var, 60) * scale * 0.8))
                b_str = "bold" if get_val(bold_var, 0) == 1 else "normal"
                i_str = "italic" if get_val(ita_var, 0) == 1 else "roman"
                tk_font = (f_name, f_size, b_str, i_str)
                
                fg_col_raw = get_val(col_var, "#FFFFFF", False)
                bg_col_raw = get_val(ocol_var, "#000000", False)
                fg_alpha = get_val(alpha_var, "00", False)
                bg_alpha = get_val(oalpha_var, "00", False)
                
                fg_col = blend_hex(fg_col_raw, vid_bg_color, fg_alpha)
                bg_col = blend_hex(bg_col_raw, vid_bg_color, bg_alpha)
                
                out_w = max(0, int(get_val(out_var, 2) * scale))
                
                if align in [1, 2, 3]: cy = ch - (mv * scale)
                elif align in [7, 8, 9]: cy = mv * scale
                else: cy = ch / 2
                
                if align in [1, 4, 7]: cx = mlr * scale
                elif align in [3, 6, 9]: cx = cw - (mlr * scale)
                else: cx = cw / 2
                
                anc_map = {1: "sw", 2: "s", 3: "se", 4: "w", 5: "center", 6: "e", 7: "nw", 8: "n", 9: "ne"}
                anc = anc_map.get(align, "s")
                
                if out_w > 0:
                    offsets = [(out_w,0), (-out_w,0), (0,out_w), (0,-out_w), (out_w,out_w), (-out_w,-out_w), (out_w,-out_w), (-out_w,out_w)]
                    for dx, dy in offsets:
                        tid = canvas.create_text(cx+dx, cy+dy, text="【示例字幕预览】", fill=bg_col, font=tk_font, anchor=anc)
                        text_ids.append(tid)
                        
                tid = canvas.create_text(cx, cy, text="【示例字幕预览】", fill=fg_col, font=tk_font, anchor=anc)
                text_ids.append(tid)

                if mouse_pos["show"]:
                    mx, my = mouse_pos["x"], mouse_pos["y"]
                    l1 = canvas.create_line(0, my, cw, my, fill="#00FF00", dash=(4, 4))
                    l2 = canvas.create_line(mx, 0, mx, ch, fill="#00FF00", dash=(4, 4))
                    crosshair_ids.extend([l1, l2])
                    
                    info_txt = f"左右边距(Margin L/R): {drag_data['mlr']} \n上下边距(Margin V): {drag_data['mv']}"
                    tx = mx + 10 if mx < cw - 180 else mx - 10
                    ty = my + 10 if my < ch - 40 else my - 30
                    anc_txt = "nw" if mx < cw - 180 else "ne"
                    
                    tid_shadow = canvas.create_text(tx+1, ty+1, text=info_txt, fill="#000000", font=("Microsoft YaHei", 9, "bold"), anchor=anc_txt)
                    tid_main = canvas.create_text(tx, ty, text=info_txt, fill="#00FF00", font=("Microsoft YaHei", 9, "bold"), anchor=anc_txt)
                    crosshair_ids.extend([tid_shadow, tid_main])
            
            draw_preview()
            
            def update_mouse(event, is_drag=False):
                nx, ny = max(0, min(event.x, cw)), max(0, min(event.y, ch))
                mouse_pos["x"] = nx
                mouse_pos["y"] = ny
                mouse_pos["show"] = True
                
                if is_drag:
                    align = get_val(align_var, 2)
                    if align in [1, 2, 3]: new_mv = (ch - ny) / scale
                    elif align in [7, 8, 9]: new_mv = ny / scale
                    else: new_mv = drag_data["mv"] 
                    
                    if align in [1, 4, 7]: new_mlr = nx / scale
                    elif align in [3, 6, 9]: new_mlr = (cw - nx) / scale
                    else: new_mlr = drag_data["mlr"]
                    
                    drag_data["mv"] = int(new_mv)
                    drag_data["mlr"] = int(new_mlr)
                    
                draw_preview()

            canvas.bind("<Motion>", lambda e: update_mouse(e, False))
            canvas.bind("<B1-Motion>", lambda e: update_mouse(e, True))
            canvas.bind("<Leave>", lambda e: [mouse_pos.update({"show": False}), draw_preview()])
            
            def confirm_and_sync():
                mv_var.set(str(drag_data["mv"]))
                mlr_var.set(str(drag_data["mlr"]))
                draw_preview() 
                messagebox.showinfo("同步成功", "边距参数已写入主界面！\n画板已根据主界面最新的颜色/对齐/透明度等参数刷新预览。", parent=cv_win)
                
            ttk.Button(top_frame, text="✅ 确认修改并同步预览", command=confirm_and_sync, style='TButton').pack(ipadx=10, ipady=2)
            ttk.Label(top_frame, text="拖拽文字调整边距。如果在主界面修改了透明度/颜色/字体等，点击上方按钮即可刷新预览！", foreground="gray").pack(pady=(5,0))
            
        ttk.Button(f_res, text="确定并打开画板", command=open_canvas, style='TButton').grid(row=3, column=0, columnspan=2, pady=(15, 0), ipadx=10, ipady=2)

    ttk.Button(r2, text="📐 图形调参", command=lambda: open_visual_adjuster(parent, mv_v, mlr_v, align_v, font_v, size_v, col_v, ocol_v, out_v, bold_v, ita_v, alpha_v, oalpha_v)).pack(side=tk.LEFT, padx=(5, 0))

    return cb, c_btn, oc_btn

def scan_ass_for_styles(path):
    s = {}
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8-sig') as f:
            in_s = False
            for line in f:
                l = line.strip()
                if l.startswith('[V4+ Styles]'): in_s = True; continue
                if l.startswith('[Events]'): break
                if in_s and l.startswith('Style:'): s[l.split('Style:')[1].split(',')[0].strip()] = l
    return s

# ================= TAB 5: 纯净 SRT 转 ASS =================
tab_ass = ttk.Frame(nb_ass, padding=10)
nb_ass.add(tab_ass, text=" SRT转ASS ")
tab_ass.columnconfigure(1, weight=1)

ass_srt_var, ass_out_var = tk.StringVar(), tk.StringVar()
ass_bracket_var = tk.StringVar(value="^\\[")

ttk.Label(tab_ass, text="仅限 SRT 输入文件夹:").grid(row=0, column=0, sticky="e", padx=(0,5), pady=5)
ttk.Entry(tab_ass, textvariable=ass_srt_var).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
ttk.Button(tab_ass, text="浏览...", command=lambda: ask_dir(ass_srt_var, "选择目录")).grid(row=0, column=2, padx=(5,0), pady=5)

ttk.Label(tab_ass, text="ASS 输出文件夹:").grid(row=1, column=0, sticky="e", padx=(0,5), pady=5)
ttk.Entry(tab_ass, textvariable=ass_out_var).grid(row=1, column=1, sticky="ew", padx=5, pady=5)
ttk.Button(tab_ass, text="浏览...", command=lambda: ask_dir(ass_out_var, "选择目录")).grid(row=1, column=2, padx=(5,0), pady=5)

f_ass_txt = ttk.LabelFrame(tab_ass, text="文本处理 (画面字提取、正则替换、合并)", padding=10)
f_ass_txt.grid(row=2, column=0, columnspan=3, sticky="ew", pady=10, padx=5)
f_ass_txt.columnconfigure(1, weight=1)

ttk.Label(f_ass_txt, text="画面字正则条件:").grid(row=0, column=0, sticky="e", padx=(0,5))
ttk.Entry(f_ass_txt, textvariable=ass_bracket_var).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Label(f_ass_txt, text="例如: ^\\[ (只要匹配成功，该条即全划为画面字)", font=("Arial", 8)).grid(row=0, column=2, sticky="w", padx=5)

# 新增：正则替换功能的开关变量（默认设为 0，即关闭状态）
ass_enable_regex_var = tk.IntVar(value=0)
# 把原来的 Label 替换成带有绑定变量的 Checkbutton
ttk.Checkbutton(f_ass_txt, text="开启正则批量替换:", variable=ass_enable_regex_var).grid(row=1, column=0, sticky="ne", padx=(0,5), pady=5)

ass_regex_text = tk.Text(f_ass_txt, height=3, width=50, font=('Arial', 9))
ass_regex_text.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
ass_regex_text.insert(tk.END, r"""1. 查找指定数量的特定字符，此处为{2}，查找\\N或者\n: ^(?:.*(?:\\N|\n)){2}[\s\S]*$
2. []替换为()并加上换行: \[([^\]]*)\] >>> ($1)\\N
3. 只替换首尾的[]为()：^\[([\s\S]*)\]$ >>> ($1)""")

ass_merge_var = tk.IntVar(value=0)
ass_merge_report_var = tk.StringVar()
ttk.Checkbutton(f_ass_txt, text="合并相邻且相同的字幕 (时间轴无缝衔接)", variable=ass_merge_var).grid(row=3, column=0, columnspan=2, sticky="w", padx=5, pady=5)
ttk.Label(f_ass_txt, text="合并报告导出至:").grid(row=4, column=0, sticky="e", padx=(0,5))
ttk.Entry(f_ass_txt, textvariable=ass_merge_report_var).grid(row=4, column=1, sticky="ew", padx=5)
ttk.Button(f_ass_txt, text="浏览...", command=lambda: ask_save_file(ass_merge_report_var, "保存合并报告", [("Excel", "*.xlsx")], ".xlsx")).grid(row=4, column=2, padx=5)

f_ass_style = ttk.LabelFrame(tab_ass, text="样式设置", padding=10)
f_ass_style.grid(row=3, column=0, columnspan=3, sticky="ew", pady=5, padx=5)

ass_style_mode_5 = tk.IntVar(value=0)

ass_frame_custom = ttk.Frame(f_ass_style)
f_ass_ref_5 = ttk.Frame(f_ass_style)

# --- 插入：分辨率 UI ---
ass_resx_var, ass_resy_var = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_ass_res_5 = ttk.Frame(ass_frame_custom)
f_ass_res_5.pack(fill=tk.X, pady=(0, 5))
ttk.Label(f_ass_res_5, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_ass_res_5, textvariable=ass_resx_var, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_ass_res_5, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_ass_res_5, textvariable=ass_resy_var, width=8).pack(side=tk.LEFT, padx=5)
# ---------------------

def update_ass_style_mode_5():
    if ass_style_mode_5.get() == 0:
        ass_frame_custom.pack(fill=tk.BOTH, expand=True, pady=5)
        f_ass_ref_5.pack_forget()
    else:
        ass_frame_custom.pack_forget()
        f_ass_ref_5.pack(fill=tk.BOTH, expand=True, pady=5)

f_mode_btns_5 = ttk.Frame(f_ass_style)
f_mode_btns_5.pack(fill=tk.X, pady=2)
ttk.Radiobutton(f_mode_btns_5, text="使用下方自定义样式 (分 对白字幕/画面字)", variable=ass_style_mode_5, value=0, command=update_ass_style_mode_5).pack(side=tk.LEFT, padx=(0, 20))
ttk.Radiobutton(f_mode_btns_5, text="直接复用外部 ASS 样式 (完美保留排版/特效等)", variable=ass_style_mode_5, value=1, command=update_ass_style_mode_5).pack(side=tk.LEFT)

current_presets_ass = load_presets(PRESET_FILE_ASS, DEFAULT_PRESETS_ASS)
ass_preset_cb_var = tk.StringVar()

ass_n_font_var, ass_n_size_var, ass_n_color_var, ass_n_outcolor_var = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
ass_n_marginv_var, ass_n_marginlr_var, ass_n_outline_var = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
ass_n_align_var, ass_n_shadow_var, ass_n_bold_var, ass_n_italic_var = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)

ass_s_font_var, ass_s_size_var, ass_s_color_var, ass_s_outcolor_var = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#26E3FF"), tk.StringVar(value="#000000")
ass_s_marginv_var, ass_s_marginlr_var, ass_s_outline_var = tk.StringVar(value="850"), tk.StringVar(value="20"), tk.StringVar(value="2")
ass_s_align_var, ass_s_shadow_var, ass_s_bold_var, ass_s_italic_var = tk.StringVar(value="8"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)

style_notebook = ttk.Notebook(ass_frame_custom)
style_notebook.pack(fill=tk.X, pady=5)
tab_normal = ttk.Frame(style_notebook, padding=10); style_notebook.add(tab_normal, text=" 对白字幕 ")
tab_screen = ttk.Frame(style_notebook, padding=10); style_notebook.add(tab_screen, text=" 画面字 ")

ass_n_alpha_var = tk.StringVar(value="00"); ass_n_outalpha_var = tk.StringVar(value="00")
ass_s_alpha_var = tk.StringVar(value="00"); ass_s_outalpha_var = tk.StringVar(value="00")
n_cb, n_c_btn, n_oc_btn = build_style_tab(tab_normal, ass_n_font_var, ass_n_size_var, ass_n_color_var, ass_n_outcolor_var, ass_n_marginv_var, ass_n_marginlr_var, ass_n_outline_var, ass_n_align_var, ass_n_shadow_var, ass_n_bold_var, ass_n_italic_var, ass_n_alpha_var, ass_n_outalpha_var)
s_cb, s_c_btn, s_oc_btn = build_style_tab(tab_screen, ass_s_font_var, ass_s_size_var, ass_s_color_var, ass_s_outcolor_var, ass_s_marginv_var, ass_s_marginlr_var, ass_s_outline_var, ass_s_align_var, ass_s_shadow_var, ass_s_bold_var, ass_s_italic_var, ass_s_alpha_var, ass_s_outalpha_var)

if current_presets_ass: ass_preset_cb_var.set(list(current_presets_ass.keys())[0])
f_ps = create_ass_preset_bar(
    ass_frame_custom, 
    [ass_n_font_var, ass_n_size_var, ass_n_color_var, ass_n_outcolor_var, ass_n_marginv_var, ass_n_marginlr_var, ass_n_outline_var, ass_n_align_var, ass_n_shadow_var, ass_n_bold_var, ass_n_italic_var, ass_n_alpha_var, ass_n_outalpha_var],
    [ass_s_font_var, ass_s_size_var, ass_s_color_var, ass_s_outcolor_var, ass_s_marginv_var, ass_s_marginlr_var, ass_s_outline_var, ass_s_align_var, ass_s_shadow_var, ass_s_bold_var, ass_s_italic_var, ass_s_alpha_var, ass_s_outalpha_var],
    [n_c_btn, n_oc_btn, s_c_btn, s_oc_btn], ass_preset_cb_var, [ass_resx_var, ass_resy_var]
)

f_ps.pack(fill=tk.X, pady=(5, 5))
ttk.Button(f_ps, text="▶ 开始 SRT 转 ASS", command=run_ass_convert, style='TButton').pack(side=tk.LEFT, padx=(40, 0), ipadx=15, ipady=2)

# REF MODE 5
ass_ref_path_5 = tk.StringVar()
f_ref_top_5 = ttk.Frame(f_ass_ref_5)
f_ref_top_5.pack(anchor="w", pady=5)
ttk.Label(f_ref_top_5, text="外部 ASS 文件:").pack(side=tk.LEFT)
ttk.Entry(f_ref_top_5, textvariable=ass_ref_path_5, width=30).pack(side=tk.LEFT, padx=5)
ttk.Button(f_ref_top_5, text="浏览...", command=lambda: ask_file(ass_ref_path_5, "选择ASS", [("ASS","*.ass")])).pack(side=tk.LEFT)

f_ref_mid_5 = ttk.Frame(f_ass_ref_5)
f_ref_mid_5.pack(anchor="w", pady=5)

ass_ref_n_var_5, ass_ref_s_var_5 = tk.StringVar(), tk.StringVar()
# 修复：将下拉框的父容器改为 f_ref_mid_5
cb_ref_n_5 = ttk.Combobox(f_ref_mid_5, textvariable=ass_ref_n_var_5, width=15)
cb_ref_s_5 = ttk.Combobox(f_ref_mid_5, textvariable=ass_ref_s_var_5, width=15)

def scan_ref_5():
    s = scan_ass_for_styles(ass_ref_path_5.get().strip())
    k = list(s.keys())
    cb_ref_n_5['values'] = k; cb_ref_s_5['values'] = k
    if k: ass_ref_n_var_5.set(k[0]); ass_ref_s_var_5.set(k[0])

ttk.Button(f_ref_top_5, text="扫描样式 ->", command=scan_ref_5).pack(side=tk.LEFT, padx=5)

ttk.Label(f_ref_mid_5, text="赋予给普通字:").pack(side=tk.LEFT)
cb_ref_n_5.pack(side=tk.LEFT, padx=5)
ttk.Label(f_ref_mid_5, text="赋予给画面字:").pack(side=tk.LEFT, padx=10)
cb_ref_s_5.pack(side=tk.LEFT, padx=5)

f_ref_bot_5 = ttk.Frame(f_ass_ref_5)
f_ref_bot_5.pack(anchor="w", pady=5)
ass_ref_font_mode_5 = tk.IntVar(value=0)
ass_ref_override_font_5 = tk.StringVar(value="SimHei")
ttk.Radiobutton(f_ref_bot_5, text="保留参考样式的原始字体", variable=ass_ref_font_mode_5, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_ref_bot_5, text="覆盖字体为:", variable=ass_ref_font_mode_5, value=1).pack(side=tk.LEFT, padx=5)
cb_ref_font_5 = ttk.Combobox(f_ref_bot_5, textvariable=ass_ref_override_font_5, width=25)
cb_ref_font_5.pack(side=tk.LEFT, padx=5)
ttk.Button(f_ref_bot_5, text="▶ 开始 SRT 转 ASS", command=run_ass_convert, style='TButton').pack(side=tk.LEFT, padx=(40, 0), ipadx=15, ipady=2)

update_ass_style_mode_5()

# ================= TAB 9: SRT 分类合并转 ASS =================
tab_ms = ttk.Frame(nb_ass, padding=10)
nb_ass.add(tab_ms, text=" SRT合并/转ASS ")
tab_ms.columnconfigure(1, weight=1)

ms_norm_var, ms_scr_var, ms_out_var = tk.StringVar(), tk.StringVar(), tk.StringVar()

ttk.Label(tab_ms, text="对白字幕 SRT 文件夹:").grid(row=0, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ms, textvariable=ms_norm_var).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Button(tab_ms, text="浏览...", command=lambda: ask_dir(ms_norm_var, "选择对白字幕目录")).grid(row=0, column=2, padx=5)

ttk.Label(tab_ms, text="画面字 SRT 文件夹:").grid(row=1, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ms, textvariable=ms_scr_var).grid(row=1, column=1, sticky="ew", padx=5)
ttk.Button(tab_ms, text="浏览...", command=lambda: ask_dir(ms_scr_var, "选择画面字目录")).grid(row=1, column=2, padx=5)

ttk.Label(tab_ms, text="合成 ASS 输出文件夹:").grid(row=2, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ms, textvariable=ms_out_var).grid(row=2, column=1, sticky="ew", padx=5)
ttk.Button(tab_ms, text="浏览...", command=lambda: ask_dir(ms_out_var, "选择输出目录")).grid(row=2, column=2, padx=5)

# --- 新增：文本正则替换面板 ---
f_ms_txt = ttk.LabelFrame(tab_ms, text="文本处理 (合并前对源字幕文件进行正则替换，不影响源文件)", padding=10)
f_ms_txt.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(10, 5), padx=5)

ms_enable_regex_var = tk.IntVar(value=0)
ttk.Checkbutton(f_ms_txt, text="开启正则替换", variable=ms_enable_regex_var).grid(row=0, column=0, sticky="w", padx=5)

f_ms_txt_sub = ttk.Frame(f_ms_txt)
f_ms_txt_sub.grid(row=0, column=1, sticky="w", padx=10)
ttk.Label(f_ms_txt_sub, text="应用到:").pack(side=tk.LEFT)
ms_regex_target_var = tk.StringVar(value="画面字")
ttk.Combobox(f_ms_txt_sub, textvariable=ms_regex_target_var, values=["画面字", "对白字幕", "全部"], width=10).pack(side=tk.LEFT, padx=5)

ms_regex_text = tk.Text(f_ms_txt, height=3, width=50, font=('Arial', 9))
ms_regex_text.grid(row=1, column=0, columnspan=3, sticky="ew", padx=5, pady=5)


f_ms_style = ttk.LabelFrame(tab_ms, text="样式设置 (画面字将自动排在普通字上层)", padding=10)
f_ms_style.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(5, 10), padx=5)

ms_style_mode_9 = tk.IntVar(value=0)

f_ms_frame_custom = ttk.Frame(f_ms_style)
f_ms_ref_9 = ttk.Frame(f_ms_style)

# --- 插入：分辨率 UI ---
m5_resx_var, m5_resy_var = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_ms_res = ttk.Frame(f_ms_frame_custom)
f_ms_res.pack(fill=tk.X, pady=(0, 5))
ttk.Label(f_ms_res, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_ms_res, textvariable=m5_resx_var, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_ms_res, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_ms_res, textvariable=m5_resy_var, width=8).pack(side=tk.LEFT, padx=5)
# ---------------------

def update_ms_style_mode_9():
    if ms_style_mode_9.get() == 0:
        f_ms_frame_custom.pack(fill=tk.BOTH, expand=True, pady=5)
        f_ms_ref_9.pack_forget()
    else:
        f_ms_frame_custom.pack_forget()
        f_ms_ref_9.pack(fill=tk.BOTH, expand=True, pady=5)

ttk.Radiobutton(f_ms_style, text="使用下方自定义样式 (转换时分别分配给 对白字幕 和 画面字)", variable=ms_style_mode_9, value=0, command=update_ms_style_mode_9).pack(anchor="w", pady=2)
ttk.Radiobutton(f_ms_style, text="直接复用外部 ASS 样式 (完美保留所有排版与特效等私有参数)", variable=ms_style_mode_9, value=1, command=update_ms_style_mode_9).pack(anchor="w", pady=2)

m5_n_font_var, m5_n_size_var, m5_n_color_var, m5_n_outcolor_var = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
m5_n_marginv_var, m5_n_marginlr_var, m5_n_outline_var = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
m5_n_align_var, m5_n_shadow_var, m5_n_bold_var, m5_n_italic_var = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)

m5_s_font_var, m5_s_size_var, m5_s_color_var, m5_s_outcolor_var = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#26E3FF"), tk.StringVar(value="#000000")
m5_s_marginv_var, m5_s_marginlr_var, m5_s_outline_var = tk.StringVar(value="850"), tk.StringVar(value="20"), tk.StringVar(value="2")
m5_s_align_var, m5_s_shadow_var, m5_s_bold_var, m5_s_italic_var = tk.StringVar(value="8"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)

style_notebook_ms = ttk.Notebook(f_ms_frame_custom)
style_notebook_ms.pack(fill=tk.X, pady=5)
tab_norm_ms = ttk.Frame(style_notebook_ms, padding=10); style_notebook_ms.add(tab_norm_ms, text=" 对白字幕 ")
tab_scr_ms = ttk.Frame(style_notebook_ms, padding=10); style_notebook_ms.add(tab_scr_ms, text=" 画面字 ")

m5_n_alpha_var = tk.StringVar(value="00"); m5_n_outalpha_var = tk.StringVar(value="00")
m5_s_alpha_var = tk.StringVar(value="00"); m5_s_outalpha_var = tk.StringVar(value="00")
cb_msn, c_btn_msn, oc_btn_msn = build_style_tab(tab_norm_ms, m5_n_font_var, m5_n_size_var, m5_n_color_var, m5_n_outcolor_var, m5_n_marginv_var, m5_n_marginlr_var, m5_n_outline_var, m5_n_align_var, m5_n_shadow_var, m5_n_bold_var, m5_n_italic_var, m5_n_alpha_var, m5_n_outalpha_var)
cb_mss, c_btn_mss, oc_btn_mss = build_style_tab(tab_scr_ms, m5_s_font_var, m5_s_size_var, m5_s_color_var, m5_s_outcolor_var, m5_s_marginv_var, m5_s_marginlr_var, m5_s_outline_var, m5_s_align_var, m5_s_shadow_var, m5_s_bold_var, m5_s_italic_var, m5_s_alpha_var, m5_s_outalpha_var)

ms_preset_var = tk.StringVar()
if current_presets_ass: ms_preset_var.set(list(current_presets_ass.keys())[0])
f_ms_ps = create_ass_preset_bar(
    f_ms_frame_custom,
    [m5_n_font_var, m5_n_size_var, m5_n_color_var, m5_n_outcolor_var, m5_n_marginv_var, m5_n_marginlr_var, m5_n_outline_var, m5_n_align_var, m5_n_shadow_var, m5_n_bold_var, m5_n_italic_var, m5_n_alpha_var, m5_n_outalpha_var],
    [m5_s_font_var, m5_s_size_var, m5_s_color_var, m5_s_outcolor_var, m5_s_marginv_var, m5_s_marginlr_var, m5_s_outline_var, m5_s_align_var, m5_s_shadow_var, m5_s_bold_var, m5_s_italic_var, m5_s_alpha_var, m5_s_outalpha_var],
    [c_btn_msn, oc_btn_msn, c_btn_mss, oc_btn_mss], ms_preset_var, [m5_resx_var, m5_resy_var]
)

f_ms_ps.pack(fill=tk.X, pady=5)
ttk.Button(f_ms_ps, text="▶ 开始双源合并转 ASS", command=run_merge_srt_to_ass, style='TButton').pack(side=tk.LEFT, padx=(40, 0), ipadx=15, ipady=2)

# REF MODE 9
ms_ref_path_9 = tk.StringVar()
f_ref_top_9 = ttk.Frame(f_ms_ref_9)
f_ref_top_9.pack(anchor="w", pady=5)
ttk.Label(f_ref_top_9, text="外部 ASS 文件:").pack(side=tk.LEFT)
ttk.Entry(f_ref_top_9, textvariable=ms_ref_path_9, width=30).pack(side=tk.LEFT, padx=5)
ttk.Button(f_ref_top_9, text="浏览...", command=lambda: ask_file(ms_ref_path_9, "选择ASS", [("ASS","*.ass")])).pack(side=tk.LEFT)

f_ref_mid_9 = ttk.Frame(f_ms_ref_9)
f_ref_mid_9.pack(anchor="w", pady=5)

ms_ref_n_var_9, ms_ref_s_var_9 = tk.StringVar(), tk.StringVar()
# 修复：将下拉框的父容器改为 f_ref_mid_9
cb_ref_n_9 = ttk.Combobox(f_ref_mid_9, textvariable=ms_ref_n_var_9, width=15)
cb_ref_s_9 = ttk.Combobox(f_ref_mid_9, textvariable=ms_ref_s_var_9, width=15)

def scan_ref_9():
    s = scan_ass_for_styles(ms_ref_path_9.get().strip())
    k = list(s.keys())
    cb_ref_n_9['values'] = k; cb_ref_s_9['values'] = k
    if k: ms_ref_n_var_9.set(k[0]); ms_ref_s_var_9.set(k[0])

ttk.Button(f_ref_top_9, text="扫描样式 ->", command=scan_ref_9).pack(side=tk.LEFT, padx=5)

ttk.Label(f_ref_mid_9, text="赋予给普通字:").pack(side=tk.LEFT)
cb_ref_n_9.pack(side=tk.LEFT, padx=5)
ttk.Label(f_ref_mid_9, text="赋予给画面字:").pack(side=tk.LEFT, padx=10)
cb_ref_s_9.pack(side=tk.LEFT, padx=5)

f_ref_bot_9 = ttk.Frame(f_ms_ref_9)
f_ref_bot_9.pack(anchor="w", pady=5)
ms_ref_font_mode_9 = tk.IntVar(value=0)
ms_ref_override_font_9 = tk.StringVar(value="SimHei")
ttk.Radiobutton(f_ref_bot_9, text="保留参考样式的原始字体", variable=ms_ref_font_mode_9, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_ref_bot_9, text="覆盖字体为:", variable=ms_ref_font_mode_9, value=1).pack(side=tk.LEFT, padx=5)
cb_ref_font_9 = ttk.Combobox(f_ref_bot_9, textvariable=ms_ref_override_font_9, width=25)
cb_ref_font_9.pack(side=tk.LEFT, padx=5)
ttk.Button(f_ref_bot_9, text="▶ 开始双源合并转 ASS", command=run_merge_srt_to_ass, style='TButton').pack(side=tk.LEFT, padx=(40, 0), ipadx=15, ipady=2)

update_ms_style_mode_9()

# ================= TAB 6: ASS 样式修改与复用 =================
tab_edit = ttk.Frame(nb_ass, padding=10)
nb_ass.add(tab_edit, text=" SRT/ASS编辑器 ")
tab_edit.columnconfigure(1, weight=1)

edit_in_var, edit_out_var = tk.StringVar(), tk.StringVar()

# ================= TAB 11: ASS 合并 =================
tab_am = ttk.Frame(nb_ass, padding=20)
nb_ass.add(tab_am, text=" ASS 合并 ")
tab_am.columnconfigure(1, weight=1)

am_dir1_var = tk.StringVar()
am_dir2_var = tk.StringVar()
am_out_var = tk.StringVar()

ttk.Label(tab_am, text="说明: 此功能会将 文件夹2 的字幕行 置于 文件夹1 的字幕行顶部，并自动合并去重样式头。", foreground="gray").grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="w")

ttk.Label(tab_am, text="文件夹1 (主文件/在下):").grid(row=1, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_am, textvariable=am_dir1_var).grid(row=1, column=1, sticky="ew", padx=5)
ttk.Button(tab_am, text="浏览...", command=lambda: ask_dir(am_dir1_var, "选择文件夹1")).grid(row=1, column=2, padx=5)

ttk.Label(tab_am, text="文件夹2 (要拼接顶部的):").grid(row=2, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_am, textvariable=am_dir2_var).grid(row=2, column=1, sticky="ew", padx=5)
ttk.Button(tab_am, text="浏览...", command=lambda: ask_dir(am_dir2_var, "选择文件夹2")).grid(row=2, column=2, padx=5)

ttk.Label(tab_am, text="合并后输出文件夹:").grid(row=3, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_am, textvariable=am_out_var).grid(row=3, column=1, sticky="ew", padx=5)
ttk.Button(tab_am, text="浏览...", command=lambda: ask_dir(am_out_var, "选择输出文件夹")).grid(row=3, column=2, padx=5)

ttk.Button(tab_am, text="开始合并 ASS", command=run_ass_merge, style='TButton').grid(row=4, column=0, columnspan=3, pady=20, ipadx=20, ipady=5)

# ====== 核心修复：当输入文件夹发生改变时，自动清空旧的内存暂存 ======
def clear_memory_on_dir_change(*args):
    global_ass_memory_cache.clear()
edit_in_var.trace_add("write", clear_memory_on_dir_change)
# ==============================================================

ttk.Label(tab_edit, text="字幕输入文件夹:").grid(row=0, column=0, sticky="e", padx=(0,5), pady=5)
ttk.Entry(tab_edit, textvariable=edit_in_var).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
# === 第 0 行：输入文件夹的浏览按钮 + 暂存按钮 ===
f_edit_in_btns = ttk.Frame(tab_edit)
f_edit_in_btns.grid(row=0, column=2, sticky="w", padx=(5,0), pady=5)
ttk.Button(f_edit_in_btns, text="浏览...", command=lambda: ask_dir(edit_in_var, "选择目录")).pack(side=tk.LEFT)
ttk.Button(f_edit_in_btns, text="💾 暂存到内存", command=lambda: execute_ass_editor(stage_only=True)).pack(side=tk.LEFT, padx=(10, 0))

ttk.Label(tab_edit, text="修改后字幕输出:").grid(row=1, column=0, sticky="e", padx=(0,5), pady=5)
ttk.Entry(tab_edit, textvariable=edit_out_var).grid(row=1, column=1, sticky="ew", padx=5, pady=5)

# === 第 1 行：输出文件夹的浏览按钮 + 执行按钮 ===
f_edit_out_btns = ttk.Frame(tab_edit)
f_edit_out_btns.grid(row=1, column=2, sticky="w", padx=(5,0), pady=5)
ttk.Button(f_edit_out_btns, text="浏览...", command=lambda: ask_dir(edit_out_var, "选择目录")).pack(side=tk.LEFT)
ttk.Button(f_edit_out_btns, text="▶ 批处理输出", command=lambda: execute_ass_editor(stage_only=False)).pack(side=tk.LEFT, padx=(10, 0))
# ================= 修复：赋予子面板垂直拉伸能力 =================
# 新增：告诉底层 grid 系统，把多余或缩小的垂直空间全都分配给 row=2（即子面板所在行）
tab_edit.rowconfigure(2, weight=1) 

edit_nb = ttk.Notebook(tab_edit)
# 修改：将 sticky="ew" 改为 sticky="nsew" (North-South-East-West)，允许上下左右全方位拉伸
edit_nb.grid(row=2, column=0, columnspan=3, sticky="nsew", pady=10, padx=5)
# ==============================================================

ASS_COLS = ["0: Layer", "1: Start", "2: End", "3: Style", "4: Name", "5: MarginL", "6: MarginR", "7: MarginV", "8: Effect", "9: Text"]
SRT_COLS = ["0: ID", "1: Timeline", "2: Text"]

# ------ 功能0: 依指定列多选改样式 (替代原功能1和功能4) ------
etab_m0 = ttk.Frame(edit_nb, padding=10)
edit_nb.add(etab_m0, text="按指定列改样式")
etab_m0.columnconfigure(1, weight=1)

f_m0_top = ttk.Frame(etab_m0)
f_m0_top.grid(row=0, column=0, columnspan=3, sticky="ew", pady=5)
ttk.Label(f_m0_top, text="选择要扫描检索的列:").pack(side=tk.LEFT)
m0_col_var = tk.StringVar(value=ASS_COLS[3]) # 默认选择 Style 列
ttk.Combobox(f_m0_top, textvariable=m0_col_var, values=ASS_COLS, state="readonly", width=12).pack(side=tk.LEFT, padx=5)

def scan_m2_features():
    d = edit_in_var.get().strip()
    if not d or not os.path.exists(d): return messagebox.showwarning("提示", "请先在上方【字幕输入文件夹】中选择目录！")
    ass_files = [os.path.join(d, f) for f in os.listdir(d) if f.lower().endswith('.ass')]
    if not ass_files: return messagebox.showwarning("提示", "输入文件夹中未找到 .ass 文件！")
    
    effs, styles = set(), set()
    for filepath in ass_files:
        file_name = os.path.basename(filepath)
        # 优先读取内存暂存，保证流式处理最新数据
        if file_name in global_ass_memory_cache:
            lines = global_ass_memory_cache[file_name].split('\n')
        else:
            try:
                with open(filepath, 'r', encoding='utf-8-sig') as f: lines = f.read().split('\n')
            except: continue
            
        for line in lines:
            if line.startswith('Dialogue:'):
                p = line.split(',', 9)
                if len(p) >= 10:
                    styles.add(p[3].strip())
                    effs.add(p[8].strip())
                    
    lb_m2_effs.delete(0, tk.END)
    for e in sorted(list(effs)): lb_m2_effs.insert(tk.END, e)
    
    lb_m2_styles.delete(0, tk.END)
    for s in sorted(list(styles)): lb_m2_styles.insert(tk.END, s)
    
    messagebox.showinfo("成功", f"扫描完毕！\n共发现 {len(effs)} 种特效说明，{len(styles)} 种样式。")

def scan_m0_cols():
    d = edit_in_var.get().strip()
    if not d or not os.path.exists(d): return messagebox.showwarning("提示", "请先在上方选择 ASS 输入文件夹！")
    ass_files = [os.path.join(d, f) for f in os.listdir(d) if f.lower().endswith('.ass')]
    if not ass_files: return messagebox.showwarning("提示", "输入文件夹中未找到 .ass 文件！")
    
    col_idx = int(m0_col_var.get().split(':')[0])
    vals = set()
    for filepath in ass_files:
        file_name = os.path.basename(filepath)
        # --- 核心：优先从内存暂存中读取 ---
        if file_name in global_ass_memory_cache:
            lines = global_ass_memory_cache[file_name].split('\n')
        else:
            try:
                with open(filepath, 'r', encoding='utf-8-sig') as f: lines = f.read().split('\n')
                global_ass_memory_cache[file_name] = "\n".join(lines)
            except: continue
            
        for line in lines:
            if line.startswith('Dialogue:'):
                p = line.strip().split(',', 9)
                if len(p) > col_idx and p[col_idx].strip():
                    vals.add(p[col_idx].strip())
                        
    lb_m0_vals.delete(0, tk.END)
    names = sorted(list(vals))
    for n in names: lb_m0_vals.insert(tk.END, n)
    if names: messagebox.showinfo("成功", f"扫描完毕，在当前暂存文件列中发现了 {len(names)} 个唯一数据！")
    else: messagebox.showwarning("提示", "未扫描到任何内容，该列可能全为空。")

ttk.Button(f_m0_top, text="🔍 一键扫描提取数据", command=scan_m0_cols).pack(side=tk.LEFT, padx=10)

f_m0_lb = ttk.Frame(etab_m0)
f_m0_lb.grid(row=1, column=0, columnspan=3, sticky="ew", pady=5)
ttk.Label(f_m0_lb, text="在下方选中要修改的数据项 (支持鼠标拖拽 / 按住Ctrl或Shift多选):").pack(anchor="w")
lb_m0_vals = tk.Listbox(f_m0_lb, selectmode=tk.EXTENDED, height=5, exportselection=False)
lb_m0_vals.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
sb_m0 = ttk.Scrollbar(f_m0_lb, command=lb_m0_vals.yview)
sb_m0.pack(side=tk.LEFT, fill=tk.Y)
lb_m0_vals.config(yscrollcommand=sb_m0.set)

f_m0_mid = ttk.Frame(etab_m0)
f_m0_mid.grid(row=2, column=0, columnspan=3, sticky="ew", pady=5)
ttk.Label(f_m0_mid, text="将选中的字幕行统一替换为新样式，命名为:").pack(side=tk.LEFT)
m0_target_style_var = tk.StringVar(value="Mod_Style")
ttk.Entry(f_m0_mid, textvariable=m0_target_style_var, width=15).pack(side=tk.LEFT, padx=5)

edit_m0_mode = tk.IntVar(value=0)
def update_m0_ui():
    if edit_m0_mode.get() == 0:
        f_m0_custom.grid(row=4, column=0, columnspan=3, sticky="ew")
        f_m0_ref.grid_remove()
    else:
        f_m0_custom.grid_remove()
        f_m0_ref.grid(row=4, column=0, columnspan=3, sticky="w")

ttk.Radiobutton(f_m0_mid, text="赋予下方自定义样式", variable=edit_m0_mode, value=0, command=update_m0_ui).pack(side=tk.LEFT, padx=(20, 5))
ttk.Radiobutton(f_m0_mid, text="从外部 ASS 偷取样式", variable=edit_m0_mode, value=1, command=update_m0_ui).pack(side=tk.LEFT)

# 自定义样式面板
f_m0_custom = ttk.Frame(etab_m0)
e_m0_font, e_m0_size, e_m0_col, e_m0_ocol = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
e_m0_mv, e_m0_mlr, e_m0_outl = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
e_m0_align, e_m0_shad, e_m0_bold, e_m0_ita = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)
e_m0_alpha, e_m0_outalpha = tk.StringVar(value="00"), tk.StringVar(value="00")
cb_m0, c_btn_m0, oc_btn_m0 = build_style_tab(f_m0_custom, e_m0_font, e_m0_size, e_m0_col, e_m0_ocol, e_m0_mv, e_m0_mlr, e_m0_outl, e_m0_align, e_m0_shad, e_m0_bold, e_m0_ita, e_m0_alpha, e_m0_outalpha)

e_m0_resx, e_m0_resy = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_m0_res = ttk.Frame(f_m0_custom)
f_m0_res.grid(row=3, column=0, columnspan=8, sticky="w", pady=(10, 0))
ttk.Label(f_m0_res, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_m0_res, textvariable=e_m0_resx, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_m0_res, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_m0_res, textvariable=e_m0_resy, width=8).pack(side=tk.LEFT, padx=5)

m0_preset_var = tk.StringVar()
if current_presets_ass: m0_preset_var.set(list(current_presets_ass.keys())[0])
f_m0_ps = create_ass_preset_bar(f_m0_custom, [e_m0_font, e_m0_size, e_m0_col, e_m0_ocol, e_m0_mv, e_m0_mlr, e_m0_outl, e_m0_align, e_m0_shad, e_m0_bold, e_m0_ita, e_m0_alpha, e_m0_outalpha], None, [c_btn_m0, oc_btn_m0], m0_preset_var, [e_m0_resx, e_m0_resy])
f_m0_ps.grid(row=4, column=0, columnspan=8, sticky="w", pady=5)

# 外部参考面板
f_m0_ref = ttk.Frame(etab_m0)
m0_ref_path, m0_ref_style = tk.StringVar(), tk.StringVar()
f_m0_rtop = ttk.Frame(f_m0_ref)
f_m0_rtop.pack(anchor="w", pady=5)
ttk.Label(f_m0_rtop, text="外部 ASS 文件:").pack(side=tk.LEFT)
ttk.Entry(f_m0_rtop, textvariable=m0_ref_path, width=30).pack(side=tk.LEFT, padx=5)
ttk.Button(f_m0_rtop, text="浏览...", command=lambda: ask_file(m0_ref_path, "选择", [("ASS","*.ass")])).pack(side=tk.LEFT, padx=5)

m0_ref_cb = ttk.Combobox(f_m0_rtop, textvariable=m0_ref_style, width=15)
ttk.Button(f_m0_rtop, text="扫描样式 ->", command=lambda: scan_ref_for_cb(m0_ref_path.get(), m0_ref_cb, m0_ref_style)).pack(side=tk.LEFT, padx=5)
m0_ref_cb.pack(side=tk.LEFT, padx=5)

f_m0_rbot = ttk.Frame(f_m0_ref)
f_m0_rbot.pack(anchor="w", pady=5)
e_m0_font_mode, e_m0_override_font = tk.IntVar(value=0), tk.StringVar(value="SimHei")
ttk.Radiobutton(f_m0_rbot, text="保留参考样式的原始字体", variable=e_m0_font_mode, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_m0_rbot, text="覆盖字体为:", variable=e_m0_font_mode, value=1).pack(side=tk.LEFT, padx=5)
cb_m0_ref_font = ttk.Combobox(f_m0_rbot, textvariable=e_m0_override_font, width=25)
cb_m0_ref_font.pack(side=tk.LEFT, padx=5)

update_m0_ui()

# === 注意：确保完全删除了 原功能4（etab_eff） 区块的所有代码 ===

# ------ 功能2: 根据字符重新分配 ------
etab_tag = create_scrollable_tab(edit_nb, "根据[]重划定对白/画面字")
etab_tag.columnconfigure(1, weight=1)

f_m2_cond, m2_logic_var, m2_c1_var, edit_m2_bracket, m2_c2_var, lb_m2_effs, m2_c3_var, lb_m2_styles = build_advanced_condition_ui(etab_tag, edit_in_var, "重划定判定条件 (都不勾选则默认全部划为普通对白)")
f_m2_cond.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 10))
m2_c1_var.set(1) # 默认勾选条件1

edit_m2_mode = tk.IntVar(value=0)
f_m2_container = ttk.Frame(etab_tag)  

def update_m2_ui():
    if edit_m2_mode.get() == 0:
        f_m2_container.grid(row=2, column=0, columnspan=3, sticky="ew")
        f_m2_ref.grid_remove()
    else:
        f_m2_container.grid_remove()
        f_m2_ref.grid(row=2, column=0, columnspan=3, sticky="w")

# 【已撤销上个版本的居右 Frame 方案，改回原始 grid 左右并列方案】
ttk.Radiobutton(etab_tag, text="重新划分后，使用下方自定义样式赋予", variable=edit_m2_mode, value=0, command=update_m2_ui).grid(row=1, column=0, sticky="w", pady=(10, 5))
ttk.Radiobutton(etab_tag, text="重新划分后，从外部 ASS 偷取样式赋予", variable=edit_m2_mode, value=1, command=update_m2_ui).grid(row=1, column=1, sticky="w", pady=(10, 5))

f_m2_custom = ttk.Notebook(f_m2_container)
f_m2_custom.pack(fill=tk.BOTH, expand=True)
t2_n, t2_s = ttk.Frame(f_m2_custom, padding=10), ttk.Frame(f_m2_custom, padding=10)
f_m2_custom.add(t2_n, text=" 新的对白字幕样式 "); f_m2_custom.add(t2_s, text=" 新的画面字样式 ")

e_m2n_font, e_m2n_size, e_m2n_col, e_m2n_ocol = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
e_m2n_mv, e_m2n_mlr, e_m2n_outl = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
e_m2n_align, e_m2n_shad, e_m2n_bold, e_m2n_ita = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)
e_m2n_alpha, e_m2n_outalpha = tk.StringVar(value="00"), tk.StringVar(value="00")
cb_m2n, c_btn_m2n, oc_btn_m2n = build_style_tab(t2_n, e_m2n_font, e_m2n_size, e_m2n_col, e_m2n_ocol, e_m2n_mv, e_m2n_mlr, e_m2n_outl, e_m2n_align, e_m2n_shad, e_m2n_bold, e_m2n_ita, e_m2n_alpha, e_m2n_outalpha)

# --- 这里补回了被吞掉的画面字变量声明 ---
e_m2s_font, e_m2s_size, e_m2s_col, e_m2s_ocol = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#26E3FF"), tk.StringVar(value="#000000")
e_m2s_mv, e_m2s_mlr, e_m2s_outl = tk.StringVar(value="850"), tk.StringVar(value="20"), tk.StringVar(value="2")
e_m2s_align, e_m2s_shad, e_m2s_bold, e_m2s_ita = tk.StringVar(value="8"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)
# ----------------------------------------

e_m2s_alpha, e_m2s_outalpha = tk.StringVar(value="00"), tk.StringVar(value="00")
cb_m2s, c_btn_m2s, oc_btn_m2s = build_style_tab(t2_s, e_m2s_font, e_m2s_size, e_m2s_col, e_m2s_ocol, e_m2s_mv, e_m2s_mlr, e_m2s_outl, e_m2s_align, e_m2s_shad, e_m2s_bold, e_m2s_ita, e_m2s_alpha, e_m2s_outalpha)

# --- 新增：分辨率 UI ---
e_m2_resx, e_m2_resy = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_m2_res = ttk.Frame(f_m2_container)
f_m2_res.pack(fill=tk.X, pady=(5, 0))
ttk.Label(f_m2_res, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_m2_res, textvariable=e_m2_resx, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_m2_res, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_m2_res, textvariable=e_m2_resy, width=8).pack(side=tk.LEFT, padx=5)

m2_preset_var = tk.StringVar()
if current_presets_ass: m2_preset_var.set(list(current_presets_ass.keys())[0])
f_m2_ps = create_ass_preset_bar(
    f_m2_container,
    [e_m2n_font, e_m2n_size, e_m2n_col, e_m2n_ocol, e_m2n_mv, e_m2n_mlr, e_m2n_outl, e_m2n_align, e_m2n_shad, e_m2n_bold, e_m2n_ita, e_m2n_alpha, e_m2n_outalpha],
    [e_m2s_font, e_m2s_size, e_m2s_col, e_m2s_ocol, e_m2s_mv, e_m2s_mlr, e_m2s_outl, e_m2s_align, e_m2s_shad, e_m2s_bold, e_m2s_ita, e_m2s_alpha, e_m2s_outalpha],
    [c_btn_m2n, oc_btn_m2n, c_btn_m2s, oc_btn_m2s], m2_preset_var, [e_m2_resx, e_m2_resy]
)
f_m2_ps.pack(fill=tk.X, pady=5)

f_m2_ref = ttk.Frame(etab_tag)
m2_ref_path, m2_ref_n, m2_ref_s = tk.StringVar(), tk.StringVar(), tk.StringVar()
f_m2_top = ttk.Frame(f_m2_ref)
f_m2_top.pack(anchor="w", pady=5)
ttk.Label(f_m2_top, text="提供样式的外部 ASS 文件:").pack(side=tk.LEFT)
ttk.Entry(f_m2_top, textvariable=m2_ref_path, width=30).pack(side=tk.LEFT, padx=5)
ttk.Button(f_m2_top, text="浏览...", command=lambda: ask_file(m2_ref_path, "选择", [("ASS","*.ass")])).pack(side=tk.LEFT, padx=5)

f_m2_mid = ttk.Frame(f_m2_ref)
f_m2_mid.pack(anchor="w", pady=5)

# 修复：将下拉框的父容器改为 f_m2_mid
m2_n_cb = ttk.Combobox(f_m2_mid, textvariable=m2_ref_n, width=15)
m2_s_cb = ttk.Combobox(f_m2_mid, textvariable=m2_ref_s, width=15)

def scan_ref_2():
    s = scan_ass_for_styles(m2_ref_path.get().strip())
    k = list(s.keys())
    m2_n_cb['values'] = k; m2_s_cb['values'] = k
    if k: m2_ref_n.set(k[0]); m2_ref_s.set(k[0])

ttk.Button(f_m2_top, text="扫描该参考文件中的样式 ->", command=scan_ref_2).pack(side=tk.LEFT, padx=5)

ttk.Label(f_m2_mid, text="将其赋予给对白字幕:").pack(side=tk.LEFT)
m2_n_cb.pack(side=tk.LEFT, padx=5)
ttk.Label(f_m2_mid, text="将其赋予给画面字:").pack(side=tk.LEFT, padx=10)
m2_s_cb.pack(side=tk.LEFT, padx=5)

f_m2_bot = ttk.Frame(f_m2_ref)
f_m2_bot.pack(anchor="w", pady=5)
e_m2_font_mode = tk.IntVar(value=0)
e_m2_override_font = tk.StringVar(value="SimHei")
ttk.Radiobutton(f_m2_bot, text="保留参考样式的原始字体", variable=e_m2_font_mode, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_m2_bot, text="覆盖字体为:", variable=e_m2_font_mode, value=1).pack(side=tk.LEFT, padx=5)
cb_m2_ref_font = ttk.Combobox(f_m2_bot, textvariable=e_m2_override_font, width=25)
cb_m2_ref_font.pack(side=tk.LEFT, padx=5)


# ------ 功能3: 基于时间轴同步复制样式/特效 ------
etab_sync = ttk.Frame(edit_nb, padding=10)
edit_nb.add(etab_sync, text="依据时间轴复用样式")
etab_sync.columnconfigure(1, weight=1)

m3_ref_dir = tk.StringVar()
ttk.Label(etab_sync, text="提供样式/特效的参考文件夹:").grid(row=0, column=0, sticky="e", pady=5)
ttk.Entry(etab_sync, textvariable=m3_ref_dir).grid(row=0, column=1, sticky="ew", padx=5)
ttk.Button(etab_sync, text="浏览...", command=lambda: ask_dir(m3_ref_dir, "选择参考目录")).grid(row=0, column=2)

m3_sync_type = tk.IntVar(value=0)
ttk.Radiobutton(etab_sync, text="复用样式 (Style) 及排版", variable=m3_sync_type, value=0, command=lambda: m3_keep_font_cb.config(state="normal")).grid(row=1, column=0, sticky="w", pady=5)
ttk.Radiobutton(etab_sync, text="仅复用特效说明列 (Effect)", variable=m3_sync_type, value=1, command=lambda: m3_keep_font_cb.config(state="disabled")).grid(row=1, column=1, sticky="w", pady=5)

m3_keep_font = tk.IntVar(value=0)
m3_keep_font_cb = ttk.Checkbutton(etab_sync, text="替换样式时，保留原本的字体名称 (仅替换颜色、大小等)", variable=m3_keep_font)
m3_keep_font_cb.grid(row=2, column=0, columnspan=3, sticky="w", padx=20)

m3_err_rep = tk.StringVar()
ttk.Label(etab_sync, text="时间轴不匹配报错报告保存至:").grid(row=3, column=0, sticky="e", pady=10)
ttk.Entry(etab_sync, textvariable=m3_err_rep).grid(row=3, column=1, sticky="ew", padx=5)
ttk.Button(etab_sync, text="浏览...", command=lambda: ask_save_file(m3_err_rep, "保存", [("Excel", "*.xlsx")], ".xlsx")).grid(row=3, column=2)

# ------ 功能3: 批量/条件定位正则替换 (合并原功能5和6) ------
etab_f4 = ttk.Frame(edit_nb, padding=10)
edit_nb.add(etab_f4, text="批量/条件正则")

f4_format_var = tk.StringVar(value="ASS")
def update_f4_cols():
    if f4_format_var.get() == "ASS":
        cb_f4_tgt['values'] = ASS_COLS
        f4_target_col.set(ASS_COLS[9])
    else:
        cb_f4_tgt['values'] = SRT_COLS
        f4_target_col.set(SRT_COLS[2])

f4_top = ttk.Frame(etab_f4)
f4_top.grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 10))
ttk.Radiobutton(f4_top, text="处理 ASS 格式", variable=f4_format_var, value="ASS", command=update_f4_cols).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f4_top, text="处理 SRT 格式", variable=f4_format_var, value="SRT", command=update_f4_cols).pack(side=tk.LEFT, padx=5)

# 新增：只查找不替换 & 输出报告
f4_find_only_var = tk.IntVar(value=0)
ttk.Checkbutton(f4_top, text="只查找不替换", variable=f4_find_only_var).pack(side=tk.LEFT, padx=(20, 5))
f4_report_var = tk.IntVar(value=0)
ttk.Checkbutton(f4_top, text="输出查找/替换报告 (CSV格式)", variable=f4_report_var).pack(side=tk.LEFT, padx=5)

f4_target_col = tk.StringVar(value=ASS_COLS[9])
ttk.Label(etab_f4, text="需要应用正则替换的列:").grid(row=1, column=0, sticky="e", pady=5)
cb_f4_tgt = ttk.Combobox(etab_f4, textvariable=f4_target_col, values=ASS_COLS, width=15, state="readonly")
cb_f4_tgt.grid(row=1, column=1, sticky="w", padx=5)

# === 新增：标点符号智能转换区域 ===
f4_punct_mode = tk.IntVar(value=0)
f_punct = ttk.Frame(etab_f4)
# 把它放置在需要替换的列的下拉框右边
f_punct.grid(row=1, column=2, columnspan=2, sticky="w") 
ttk.Label(f_punct, text="字幕文本标点转换:").pack(side=tk.LEFT, padx=(15, 5))
ttk.Radiobutton(f_punct, text="不转换", variable=f4_punct_mode, value=0).pack(side=tk.LEFT)
ttk.Radiobutton(f_punct, text="半转全(删空格)", variable=f4_punct_mode, value=1).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_punct, text="全转半(补空格)", variable=f4_punct_mode, value=2).pack(side=tk.LEFT)
# ==================================

ttk.Label(etab_f4, text="正则替换规则:").grid(row=2, column=0, sticky="ne", pady=5)
f4_regex_text = tk.Text(etab_f4, height=3, width=45, font=('Arial', 9))
f4_regex_text.grid(row=2, column=1, columnspan=3, sticky="w", padx=5, pady=5)
f4_regex_text.insert(tk.END, r"""1. 查找指定数量的特定字符，此处为{2}，查找\\N或者\n: ^(?:.*(?:\\N|\n)){2}[\s\S]*$
2. []替换为()并加上换行: \[([^\]]*)\] >>> ($1)\\N
3. 只替换首尾的[]为()：^\[([\s\S]*)\]$ >>> ($1)""")

f_f4_cond, f4_logic_var, f4_c1_var, f4_bracket_var, f4_c2_var, lb_f4_effs, f4_c3_var, lb_f4_styles = build_advanced_condition_ui(etab_f4, edit_in_var, "定位条件 (都不勾选则为针对目标列的全量批量替换)")
f_f4_cond.grid(row=3, column=0, columnspan=4, sticky="ew", pady=10, padx=5)

# ------ 功能7: 手动选中修改样式 ------
etab_f7 = ttk.Frame(edit_nb, padding=5)
edit_nb.add(etab_f7, text="手动选中行改样式")

f7_top = ttk.Frame(etab_f7)
f7_top.pack(fill=tk.X, pady=5)
m7_file_var = tk.StringVar()

def load_m7_files():
    d = edit_in_var.get().strip()
    if not d or not os.path.exists(d): return messagebox.showwarning("提示", "请先在上方选择输入文件夹！")
    files = [f for f in os.listdir(d) if f.lower().endswith('.ass')]
    m7_file_cb['values'] = files
    if files: m7_file_var.set(files[0])
    
def load_m7_content():
    f = m7_file_var.get().strip()
    if not f: return
    p = os.path.join(edit_in_var.get().strip(), f)
    
    # --- 核心：优先从内存暂存中读取以保证多步操作连贯 ---
    if f in global_ass_memory_cache:
        lines = global_ass_memory_cache[f].split('\n')
    else:
        if not os.path.exists(p): return
        with open(p, 'r', encoding='utf-8-sig') as f_in: lines = f_in.read().split('\n')
        global_ass_memory_cache[f] = "\n".join(lines)
    
    m7_tree.delete(*m7_tree.get_children())
    curr, ev_lines_temp = "info", []
    for l in lines:
        if l.strip().startswith('[V4+ Styles]'): curr = "styles"
        elif l.strip().startswith('[Events]'): curr = "events"
        if curr == "events": ev_lines_temp.append(l)
        
    for i, ev in enumerate(ev_lines_temp):
        if ev.startswith('Dialogue:'):
            pts = ev.split(',', 9)
            if len(pts) >= 10: m7_tree.insert('', 'end', values=(i, f"{pts[1]}->{pts[2]}", pts[3], pts[8], pts[9]))

def open_video_preview():
    import threading
    import subprocess
    import tempfile
    
    if not HAS_CV2:
        return messagebox.showerror("缺少依赖", "请先在终端安装库：\npip install opencv-python pillow")
    
    f_name = m7_file_var.get().strip()
    v_dir = m7_video_dir.get().strip()
    if not f_name or not v_dir or not os.path.exists(v_dir): return messagebox.showwarning("警告", "请确保已选择有效的视频文件夹并加载了对应字幕！")
    
    base = os.path.splitext(f_name)[0]
    v_path = next((os.path.join(v_dir, base + ext) for ext in ['.mp4', '.mkv', '.avi', '.mov'] if os.path.exists(os.path.join(v_dir, base + ext))), None)
    if not v_path: return messagebox.showwarning("警告", f"未在视频文件夹找到与 {base} 同名的视频文件！")
    
    sel = m7_tree.selection()
    if not sel: return messagebox.showwarning("警告", "请先在列表中点击一句字幕，以作为预览起点！")

    def t2ms(t_str):
        h, m, s_ms = t_str.split(':')
        s, ms = s_ms.split('.')
        return (int(h)*3600 + int(m)*60 + int(s))*1000 + int(ms)*10

    start_ms = t2ms(m7_tree.item(sel[0], 'values')[1].split('->')[0])
    
    pv_win = tk.Toplevel(root)
    pv_win.title(f"实时视频调参 - {f_name}")
    
    try:
        cap = cv2.VideoCapture(v_path, cv2.CAP_ANY, [cv2.CAP_PROP_HW_ACCELERATION, cv2.VIDEO_ACCELERATION_ANY])
        if not cap.isOpened(): cap = cv2.VideoCapture(v_path)
    except:
        cap = cv2.VideoCapture(v_path)

    if not cap.isOpened():
        pv_win.destroy()
        return messagebox.showerror("错误", "无法读取视频流！")

    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    vid_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    vid_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    if vid_w == 0 or vid_h == 0:
        pv_win.destroy()
        return messagebox.showerror("错误", "视频分辨率读取异常！")

    screen_w, screen_h = root.winfo_screenwidth() * 0.7, root.winfo_screenheight() * 0.7
    scale = min(screen_w / vid_w, screen_h / vid_h)
    cw, ch = int(vid_w * scale), int(vid_h * scale)
    
    pv_win.geometry(f"{cw + 40}x{ch + 80}")
    
    top_f = tk.Frame(pv_win); top_f.pack(fill=tk.X, pady=5)
    canvas = tk.Canvas(pv_win, width=cw, height=ch, bg="black", highlightthickness=0)
    canvas.pack(pady=5, expand=True)
    
    state = {
        "play": False, "ms": start_ms, "seek_req": start_ms, "auto_sync": False,
        "drag_mv": 20, "drag_mlr": 20, "drag_align": 2, "drag_target_id": sel[0] if sel else None,
        "mx": cw//2, "my": ch//2, "cross": False, "bg_img": None, "loop_id": None,
        "last_cache": "", "play_res_x": 1080.0, "play_res_y": 1920.0, 
        "styles": {}, "events": [], "hide_overlay": False, "is_ffmpeging": False
    }

    # ================= 核心1：实时同步解析内存大闸 =================
    def sync_ass_memory():
        content = global_ass_memory_cache.get(f_name, "")
        if content == state["last_cache"]: return False
        
        lines = content.split('\n')
        state["play_res_x"], state["play_res_y"] = 1080.0, 1920.0
        state["styles"].clear()
        state["events"].clear()
        
        tree_ids = m7_tree.get_children()
        idx = 0
        curr = ""
        
        for line in lines:
            l = line.strip()
            if l.startswith('[V4+ Styles]'): curr = "styles"
            elif l.startswith('[Events]'): curr = "events"
            elif l.startswith('PlayResX:'): 
                try: state["play_res_x"] = float(l.split(':')[1].strip())
                except: pass
            elif l.startswith('PlayResY:'): 
                try: state["play_res_y"] = float(l.split(':')[1].strip())
                except: pass
            
            if curr == "styles" and l.startswith('Style:'):
                pts = l.split('Style:')[1].split(',')
                if len(pts) >= 22: state["styles"][pts[0].strip()] = pts
            elif curr == "events" and l.startswith('Dialogue:'):
                pts = l.split(',', 9)
                if len(pts) >= 10:
                    tid = tree_ids[idx] if idx < len(tree_ids) else None
                    state["events"].append({"id": tid, "st": t2ms(pts[1].strip()), "ed": t2ms(pts[2].strip()), "style": pts[3].strip(), "text": pts[9].replace('\\N', '\n')})
                    idx += 1
                    
        state["last_cache"] = content
        return True

    # 初始化拉取一次数据
    sync_ass_memory()

    def parse_color(ass_c):
        try:
            s = ass_c.strip().upper().replace('&H', '')
            if len(s) >= 8: a, b, g, r = s[-8:-6], s[-6:-4], s[-4:-2], s[-2:]
            elif len(s) == 6: a, b, g, r = "00", s[-6:-4], s[-4:-2], s[-2:]
            else: s = s.zfill(6); a, b, g, r = "00", s[-6:-4], s[-4:-2], s[-2:]
            return f"#{r}{g}{b}"
        except: return "#FFFFFF"

    # ================= 核心2：FFmpeg 真实压制效果一键桥接 =================
    def render_ffmpeg_frame():
        if state["is_ffmpeging"]: return
        state["is_ffmpeging"] = True
        btn_ff.config(text="⏳ 正在调用 FFmpeg 渲染...", state=tk.DISABLED)
        
        def task():
            try:
                with tempfile.TemporaryDirectory() as td:
                    ass_path = os.path.join(td, "temp.ass")
                    with open(ass_path, "w", encoding="utf-8-sig") as f:
                        f.write(global_ass_memory_cache[f_name])
                    
                    cmd = ['ffmpeg', '-y', '-ss', str(state["ms"]/1000.0), '-i', v_path, '-vf', 'ass=temp.ass', '-vframes', '1', '-f', 'image2pipe', '-vcodec', 'png', '-']
                    
                    # 隐藏 Windows 弹出的黑色控制台窗口
                    startupinfo = None
                    if platform.system() == "Windows":
                        startupinfo = subprocess.STARTUPINFO()
                        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                        
                    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, cwd=td, startupinfo=startupinfo)
                    img_data, _ = proc.communicate()
                    
                    if img_data:
                        from io import BytesIO
                        img = Image.open(BytesIO(img_data)).convert("RGB")
                        img = img.resize((cw, ch), Image.Resampling.LANCZOS)
                        
                        def update_ui():
                            state["bg_img"] = ImageTk.PhotoImage(image=img)
                            canvas.delete("bg")
                            canvas.create_image(0, 0, anchor=tk.NW, image=state["bg_img"], tags="bg")
                            state["hide_overlay"] = True  # 隐藏 Tkinter 假字幕，展示纯纯的 FFmpeg 真字幕！
                            update_overlay()
                            btn_ff.config(text="📸 获取 FFmpeg 100% 真实压制帧", state=tk.NORMAL)
                            state["is_ffmpeging"] = False
                            
                        pv_win.after(0, update_ui)
                    else: raise Exception("FFmpeg 截取失败")
            except Exception as e:
                pv_win.after(0, lambda: [btn_ff.config(text="📸 获取 FFmpeg 100% 真实压制帧", state=tk.NORMAL), state.update({"is_ffmpeging": False}), messagebox.showerror("调用失败", f"请确保系统已安装 FFmpeg 并配置了环境变量！\n报错详情: {e}", parent=pv_win)])
                
        threading.Thread(target=task, daemon=True).start()

    # ================= 核心3：精准坐标轴 UI 渲染 (Subtitle Edit 高仿版) =================
    def update_overlay():
        canvas.delete("overlay")
        if state["hide_overlay"]: return
        
        active_subs = [sub for sub in state["events"] if sub["st"] <= state["ms"] <= sub["ed"]]
                
        if active_subs:
            state["auto_sync"] = True
            try:
                current_sel = m7_tree.selection()
                active_ids = [sub["id"] for sub in active_subs if sub["id"]]
                if set(current_sel) != set(active_ids) and active_ids:
                    m7_tree.selection_set(active_ids)
                    m7_tree.see(active_ids[0])
            except: pass
            state["auto_sync"] = False

            sys_fonts = tkfont.families()
            font_alias = {"SimHei": "黑体", "Microsoft YaHei": "微软雅黑", "SimSun": "宋体", "黑体": "SimHei", "微软雅黑": "Microsoft YaHei", "宋体": "SimSun"}
            
            ass_scale_x = cw / state["play_res_x"]
            ass_scale_y = ch / state["play_res_y"]
            
            # 建立真实包围盒缓存矩阵，用于实现和 Subtitle Edit 一致的真实物理碰撞检测
            occupied_bboxes = []
            
            for sub in active_subs:
                s_name = sub["style"]
                if s_name not in state["styles"]: continue

                pts = state["styles"][s_name]
                f_name_ass = pts[1].strip()
                if f_name_ass not in sys_fonts and f_name_ass in font_alias and font_alias[f_name_ass] in sys_fonts:
                    f_name_ass = font_alias[f_name_ass]

                f_size_ass = float(pts[2].strip())
                fg, bg = parse_color(pts[3].strip()), parse_color(pts[5].strip())
                bold = 1 if pts[7].strip() == '-1' else 0
                italic = 1 if pts[8].strip() == '-1' else 0
                out_w = max(0, int(float(pts[16].strip()) * ass_scale_y))
                align, mlr, mv = int(pts[18].strip()), int(pts[19].strip()), int(pts[21].strip())
                
                # 若当前字幕是被鼠标选中拖拽的焦点，覆盖为 UI 的临时坐标
                if sub["id"] == state.get("drag_target_id"):
                    mv, mlr, align = state["drag_mv"], state["drag_mlr"], state["drag_align"]

                # [核心修正 1]: Tkinter 的字体视觉上比 ASS 原生渲染略大，添加 0.85 黄金缩放系数
                px_size = max(8, int(f_size_ass * ass_scale_y * 0.85))
                tk_f = (f_name_ass, -px_size, "bold" if bold else "normal", "italic" if italic else "roman")
                
                cy = ch - (mv * ass_scale_y) if align in [1,2,3] else (mv * ass_scale_y if align in [7,8,9] else ch/2)
                cx = mlr * ass_scale_x if align in [1,4,7] else (cw - (mlr * ass_scale_x) if align in [3,6,9] else cw/2)
                anc = {1: "sw", 2: "s", 3: "se", 4: "w", 5: "center", 6: "e", 7: "nw", 8: "n", 9: "ne"}.get(align, "s")
                
                # [核心修正 2]: 多行字幕的断行对齐逻辑，修复排版乱跳
                just = {1: tk.LEFT, 4: tk.LEFT, 7: tk.LEFT, 3: tk.RIGHT, 6: tk.RIGHT, 9: tk.RIGHT}.get(align, tk.CENTER)
                
                # [核心修正 3]: 剥离 {xxx} 这种 Tkinter 无法识别渲染的 ASS 特效标签代码
                txt = re.sub(r'\{.*?\}', '', sub["text"])

                # [核心修正 4]: 获取该段文字真实的宽高包围盒，用于精准计算碰撞弹开
                dummy_id = canvas.create_text(cx, cy, text=txt, font=tk_f, anchor=anc, justify=just)
                bbox = canvas.bbox(dummy_id)
                canvas.delete(dummy_id)

                if bbox:
                    b_l, b_t, b_r, b_b = bbox
                    shift_y = 0
                    
                    # 模拟 Subtitle Edit 的堆叠碰撞逻辑 (Collision: Normal)
                    while True:
                        collision = False
                        for (u_l, u_t, u_r, u_b) in occupied_bboxes:
                            # 判断矩形是否相交 (留 2px 余量)
                            if not (b_r < u_l or b_l > u_r or b_b + shift_y + 2 < u_t or b_t + shift_y - 2 > u_b):
                                collision = True
                                break
                        if collision:
                            # 底部字幕向上堆叠，顶部字幕向下堆叠
                            if align in [1, 2, 3]: shift_y -= (b_b - b_t + 2)
                            elif align in [7, 8, 9]: shift_y += (b_b - b_t + 2)
                            else: shift_y -= (b_b - b_t + 2)
                        else:
                            break
                            
                    cy += shift_y
                    occupied_bboxes.append((b_l, b_t + shift_y, b_r, b_b + shift_y))

                # [核心修正 5]: 高精度自适应描边渲染 (消除描边十字间隙)
                if out_w > 0:
                    steps = max(8, int(out_w * 3)) # 自适应增加描边密度
                    for i in range(steps):
                        angle = i * (2 * math.pi / steps)
                        dx = out_w * math.cos(angle)
                        dy = out_w * math.sin(angle)
                        canvas.create_text(cx+dx, cy+dy, text=txt, fill=bg, font=tk_f, anchor=anc, justify=just, tags="overlay")
                
                # 绘制主字幕
                canvas.create_text(cx, cy, text=txt, fill=fg, font=tk_f, anchor=anc, justify=just, tags="overlay")
        
        # 拖拽辅助十字线 UI 绘制
        if not state["play"] and state["cross"]:
            mx, my = state["mx"], state["my"]
            canvas.create_line(0, my, cw, my, fill="#00FF00", dash=(4, 4), tags="overlay")
            canvas.create_line(mx, 0, mx, ch, fill="#00FF00", dash=(4, 4), tags="overlay")
            info = f"Margin L/R: {state['drag_mlr']}\nMargin V: {state['drag_mv']}"
            tx, ty = mx + 10 if mx < cw - 150 else mx - 10, my + 10 if my < ch - 40 else my - 30
            anc_txt = "nw" if mx < cw - 150 else "ne"
            canvas.create_text(tx+1, ty+1, text=info, fill="#000000", font=("Microsoft YaHei", 9, "bold"), anchor=anc_txt, tags="overlay")
            canvas.create_text(tx, ty, text=info, fill="#00FF00", font=("Microsoft YaHei", 9, "bold"), anchor=anc_txt, tags="overlay")

    def video_loop():
        if not pv_win.winfo_exists():
            cap.release(); return
        
        # 时刻监视内存变化，只要在主界面点了“暂存”，下一帧立刻更新排版！
        if sync_ass_memory(): update_overlay()
        
        frame_updated = False
        if state["seek_req"] is not None:
            cap.set(cv2.CAP_PROP_POS_MSEC, state["seek_req"])
            state["seek_req"] = None
            ret, frame = cap.read()
            if ret:
                state["ms"] = cap.get(cv2.CAP_PROP_POS_MSEC)
                frame = cv2.cvtColor(cv2.resize(frame, (cw, ch)), cv2.COLOR_BGR2RGB)
                state["bg_img"] = ImageTk.PhotoImage(image=Image.fromarray(frame))
                canvas.delete("bg")
                canvas.create_image(0, 0, anchor=tk.NW, image=state["bg_img"], tags="bg")
                frame_updated = True
        
        elif state["play"]:
            ret, frame = cap.read()
            if ret:
                state["ms"] = cap.get(cv2.CAP_PROP_POS_MSEC)
                frame = cv2.cvtColor(cv2.resize(frame, (cw, ch)), cv2.COLOR_BGR2RGB)
                state["bg_img"] = ImageTk.PhotoImage(image=Image.fromarray(frame))
                canvas.delete("bg")
                canvas.create_image(0, 0, anchor=tk.NW, image=state["bg_img"], tags="bg")
                frame_updated = True
            else: state["play"] = False 
                
        if frame_updated or state["play"]: update_overlay()
            
        state["loop_id"] = pv_win.after(int(1000 / fps) if state["play"] else 50, video_loop)

    def toggle_play(e=None):
        state["play"] = not state["play"]
        state["hide_overlay"] = False # 恢复实时 UI
        if not state["play"]: update_overlay()

    pv_win.bind("<space>", toggle_play)

    def on_mouse_move(e, is_drag=False):
        state["cross"], state["mx"], state["my"] = True, max(0, min(e.x, cw)), max(0, min(e.y, ch))
        if is_drag and not state["play"]:
            state["hide_overlay"] = False
            ass_scale_x, ass_scale_y = cw / state["play_res_x"], ch / state["play_res_y"]
            align = state["drag_align"]
            if align in [1, 2, 3]: state["drag_mv"] = int((ch - state["my"]) / ass_scale_y)
            elif align in [7, 8, 9]: state["drag_mv"] = int(state["my"] / ass_scale_y)
            if align in [1, 4, 7]: state["drag_mlr"] = int(state["mx"] / ass_scale_x)
            elif align in [3, 6, 9]: state["drag_mlr"] = int((cw - state["mx"]) / ass_scale_x)
        if not state["play"]: update_overlay()

    canvas.bind("<Motion>", lambda e: on_mouse_move(e, False))
    canvas.bind("<B1-Motion>", lambda e: on_mouse_move(e, True))
    canvas.bind("<Leave>", lambda e: state.update({"cross": False}) or (update_overlay() if not state["play"] else None))
    
    def on_tree_sync(e):
        if state["auto_sync"]: return
        sel = m7_tree.selection()
        if not sel: return
        
        focus_item = m7_tree.focus()
        target_item = focus_item if focus_item in sel else sel[0]
        
        # 【核心修复2】：将用户显式点击的这行字幕记录为“唯一拖拽目标”
        state["drag_target_id"] = target_item
        
        s_name = m7_tree.item(target_item, 'values')[2]
        if s_name in state["styles"]:
            state["drag_mv"] = int(state["styles"][s_name][21].strip())
            state["drag_mlr"] = int(state["styles"][s_name][19].strip())
            state["drag_align"] = int(state["styles"][s_name][18].strip())
        
        target_st = t2ms(m7_tree.item(target_item, 'values')[1].split('->')[0])
        target_ed = t2ms(m7_tree.item(target_item, 'values')[1].split('->')[1])
        if not (target_st <= state["ms"] <= target_ed): 
            state["seek_req"] = target_st
            state["hide_overlay"] = False # 恢复实时 UI
        else:
            # 如果仅仅是切换了同画面的拖拽焦点，也要刷新一下
            update_overlay()
            
    m7_tree.bind("<<TreeviewSelect>>", on_tree_sync)
    
    ttk.Button(top_f, text="▶/⏸ 播放/暂停视频 (Space)", command=toggle_play, style='TButton').pack(side=tk.LEFT, padx=10)
    ttk.Button(top_f, text="✅ 点击应用拖拽后的边距", command=lambda: e_m7_mv.set(str(state["drag_mv"])) or e_m7_mlr.set(str(state["drag_mlr"])), style='TButton').pack(side=tk.LEFT, padx=10)
    
    btn_ff = ttk.Button(top_f, text="📸 获取 FFmpeg 100% 真实压制帧", command=render_ffmpeg_frame, style='TButton')
    btn_ff.pack(side=tk.LEFT, padx=10)
    
    video_loop()

# === 将原本拥挤在一行的组件，分装到上下两个子容器中 ===
f7_top_1 = ttk.Frame(f7_top)
f7_top_1.pack(fill=tk.X, pady=(0, 2))

ttk.Button(f7_top_1, text="1. 刷新文件夹", command=load_m7_files).pack(side=tk.LEFT, padx=5)
m7_file_cb = ttk.Combobox(f7_top_1, textvariable=m7_file_var, state="readonly", width=25)
m7_file_cb.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
ttk.Button(f7_top_1, text="2. 读取字幕内容", command=load_m7_content).pack(side=tk.LEFT, padx=2)

f7_top_2 = ttk.Frame(f7_top)
f7_top_2.pack(fill=tk.X, pady=(2, 0))

m7_video_dir = tk.StringVar()
ttk.Label(f7_top_2, text="参考视频夹(可选):").pack(side=tk.LEFT, padx=(5, 2))
ttk.Entry(f7_top_2, textvariable=m7_video_dir, width=15).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
ttk.Button(f7_top_2, text="浏览...", command=lambda: ask_dir(m7_video_dir, "选择视频")).pack(side=tk.LEFT, padx=2)
ttk.Button(f7_top_2, text="🎬 视频实时预览调参", command=open_video_preview).pack(side=tk.LEFT, padx=5)
# ==============================================================

f7_tree_f = ttk.Frame(etab_f7)
f7_tree_f.pack(fill=tk.X, pady=5)
m7_tree = ttk.Treeview(f7_tree_f, columns=("idx", "time", "style", "effect", "text"), show="headings", height=5)
m7_tree.heading("idx", text="序号"); m7_tree.column("idx", width=40, anchor="center")
m7_tree.heading("time", text="时间轴"); m7_tree.column("time", width=140)
m7_tree.heading("style", text="样式"); m7_tree.column("style", width=100)
m7_tree.heading("effect", text="特效"); m7_tree.column("effect", width=100)
m7_tree.heading("text", text="内容"); m7_tree.column("text", width=300)
m7_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
sb_m7 = ttk.Scrollbar(f7_tree_f, command=m7_tree.yview)
sb_m7.pack(side=tk.RIGHT, fill=tk.Y)
m7_tree.config(yscrollcommand=sb_m7.set)

f7_bot = ttk.Frame(etab_f7)
f7_bot.pack(fill=tk.BOTH, expand=True)

edit_m7_target_var = tk.StringVar(value="新选中样式")
f7_b_top = ttk.Frame(f7_bot)
f7_b_top.pack(fill=tk.X, pady=5)
ttk.Label(f7_b_top, text="赋予的新样式名称:").pack(side=tk.LEFT, padx=(0,5))
ttk.Entry(f7_b_top, textvariable=edit_m7_target_var, width=15).pack(side=tk.LEFT)

edit_m7_mode = tk.IntVar(value=0)
def update_m7_ui():
    if edit_m7_mode.get() == 0:
        f_m7_custom.pack(fill=tk.BOTH, expand=True); f_m7_ref.pack_forget()
    else:
        f_m7_custom.pack_forget(); f_m7_ref.pack(fill=tk.BOTH, expand=True)

ttk.Radiobutton(f7_b_top, text="使用下方自定义样式赋予", variable=edit_m7_mode, value=0, command=update_m7_ui).pack(side=tk.LEFT, padx=(15, 5))
ttk.Radiobutton(f7_b_top, text="从外部 ASS 偷取样式赋予", variable=edit_m7_mode, value=1, command=update_m7_ui).pack(side=tk.LEFT)

f_m7_custom = ttk.Frame(f7_bot)
e_m7_font, e_m7_size, e_m7_col, e_m7_ocol = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
e_m7_mv, e_m7_mlr, e_m7_outl = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
e_m7_align, e_m7_shad, e_m7_bold, e_m7_ita = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)
e_m7_alpha, e_m7_outalpha = tk.StringVar(value="00"), tk.StringVar(value="00")
cb_m7, c_btn_m7, oc_btn_m7 = build_style_tab(f_m7_custom, e_m7_font, e_m7_size, e_m7_col, e_m7_ocol, e_m7_mv, e_m7_mlr, e_m7_outl, e_m7_align, e_m7_shad, e_m7_bold, e_m7_ita, e_m7_alpha, e_m7_outalpha)

e_m7_resx, e_m7_resy = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_m7_res = ttk.Frame(f_m7_custom)
f_m7_res.grid(row=3, column=0, columnspan=8, sticky="w", pady=(5, 0))
ttk.Label(f_m7_res, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_m7_res, textvariable=e_m7_resx, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_m7_res, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_m7_res, textvariable=e_m7_resy, width=8).pack(side=tk.LEFT, padx=5)

m7_preset_var = tk.StringVar()
if current_presets_ass: m7_preset_var.set(list(current_presets_ass.keys())[0])
f_m7_ps = create_ass_preset_bar(f_m7_custom, [e_m7_font, e_m7_size, e_m7_col, e_m7_ocol, e_m7_mv, e_m7_mlr, e_m7_outl, e_m7_align, e_m7_shad, e_m7_bold, e_m7_ita, e_m7_alpha, e_m7_outalpha], None, [c_btn_m7, oc_btn_m7], m7_preset_var, [e_m7_resx, e_m7_resy])
f_m7_ps.grid(row=4, column=0, columnspan=8, sticky="w", pady=5)

f_m7_ref = ttk.Frame(f7_bot)
m7_ref_path, m7_ref_style = tk.StringVar(), tk.StringVar()
f_m7_top = ttk.Frame(f_m7_ref); f_m7_top.pack(anchor="w", pady=5)
ttk.Label(f_m7_top, text="外部 ASS:").pack(side=tk.LEFT)
ttk.Entry(f_m7_top, textvariable=m7_ref_path, width=25).pack(side=tk.LEFT, padx=5)
ttk.Button(f_m7_top, text="浏览...", command=lambda: ask_file(m7_ref_path, "选择", [("ASS","*.ass")])).pack(side=tk.LEFT, padx=5)
m7_ref_cb = ttk.Combobox(f_m7_top, textvariable=m7_ref_style, width=15)
ttk.Button(f_m7_top, text="扫描样式 ->", command=lambda: scan_ref_for_cb(m7_ref_path.get(), m7_ref_cb, m7_ref_style)).pack(side=tk.LEFT, padx=5)
m7_ref_cb.pack(side=tk.LEFT, padx=5)

def scan_ref_for_cb(path, cb_widget, var_widget):
    s = scan_ass_for_styles(path.strip())
    k = list(s.keys()); cb_widget['values'] = k
    if k: var_widget.set(k[0])

e_m7_font_mode, e_m7_override_font = tk.IntVar(value=0), tk.StringVar(value="SimHei")
f_m7_bot2 = ttk.Frame(f_m7_ref); f_m7_bot2.pack(anchor="w", pady=5)
ttk.Radiobutton(f_m7_bot2, text="保留参考样式原字体", variable=e_m7_font_mode, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_m7_bot2, text="覆盖字体为:", variable=e_m7_font_mode, value=1).pack(side=tk.LEFT, padx=5)
cb_m7_ref_font = ttk.Combobox(f_m7_bot2, textvariable=e_m7_override_font, width=25)
cb_m7_ref_font.pack(side=tk.LEFT, padx=5)
update_m7_ui()

# ------ 功能5 (对应底层的模式 5): 条件定位替换样式 ------
etab_f8 = create_scrollable_tab(edit_nb, "条件定位替换样式")

f_f8_cond, f8_logic_var, f8_c1_var, f8_bracket_var, f8_c2_var, lb_f8_effs, f8_c3_var, lb_f8_styles = build_advanced_condition_ui(etab_f8, edit_in_var, "定位条件 (必须至少启用一个条件，都不勾选则不执行)")
f_f8_cond.pack(fill=tk.X, pady=5)

f8_bot = ttk.Frame(etab_f8)
f8_bot.pack(fill=tk.BOTH, expand=True, pady=5)

edit_m8_target_var = tk.StringVar(value="新选中样式")
f8_b_top = ttk.Frame(f8_bot)
f8_b_top.pack(fill=tk.X, pady=5)
ttk.Label(f8_b_top, text="符合条件则赋予新样式名称:").pack(side=tk.LEFT, padx=(0,5))
ttk.Entry(f8_b_top, textvariable=edit_m8_target_var, width=15).pack(side=tk.LEFT)

edit_m8_mode = tk.IntVar(value=0)
def update_m8_ui():
    if edit_m8_mode.get() == 0:
        f_m8_custom.pack(fill=tk.BOTH, expand=True); f_m8_ref.pack_forget()
    else:
        f_m8_custom.pack_forget(); f_m8_ref.pack(fill=tk.BOTH, expand=True)

ttk.Radiobutton(f8_b_top, text="使用下方自定义样式赋予", variable=edit_m8_mode, value=0, command=update_m8_ui).pack(side=tk.LEFT, padx=(15, 5))
ttk.Radiobutton(f8_b_top, text="从外部 ASS 偷取样式赋予", variable=edit_m8_mode, value=1, command=update_m8_ui).pack(side=tk.LEFT)

f_m8_custom = ttk.Frame(f8_bot)
e_m8_font, e_m8_size, e_m8_col, e_m8_ocol = tk.StringVar(value="SimHei"), tk.StringVar(value="60"), tk.StringVar(value="#FFFFFF"), tk.StringVar(value="#000000")
e_m8_mv, e_m8_mlr, e_m8_outl = tk.StringVar(value="20"), tk.StringVar(value="20"), tk.StringVar(value="2")
e_m8_align, e_m8_shad, e_m8_bold, e_m8_ita = tk.StringVar(value="2"), tk.StringVar(value="0"), tk.IntVar(value=0), tk.IntVar(value=0)
e_m8_alpha, e_m8_outalpha = tk.StringVar(value="00"), tk.StringVar(value="00")
cb_m8, c_btn_m8, oc_btn_m8 = build_style_tab(f_m8_custom, e_m8_font, e_m8_size, e_m8_col, e_m8_ocol, e_m8_mv, e_m8_mlr, e_m8_outl, e_m8_align, e_m8_shad, e_m8_bold, e_m8_ita, e_m8_alpha, e_m8_outalpha)

e_m8_resx, e_m8_resy = tk.StringVar(value="1080"), tk.StringVar(value="1920")
f_m8_res = ttk.Frame(f_m8_custom)
f_m8_res.grid(row=3, column=0, columnspan=8, sticky="w", pady=(5, 0))
ttk.Label(f_m8_res, text="视频分辨率 (宽/X):").pack(side=tk.LEFT)
ttk.Entry(f_m8_res, textvariable=e_m8_resx, width=8).pack(side=tk.LEFT, padx=5)
ttk.Label(f_m8_res, text="(高/Y):").pack(side=tk.LEFT)
ttk.Entry(f_m8_res, textvariable=e_m8_resy, width=8).pack(side=tk.LEFT, padx=5)

m8_preset_var = tk.StringVar()
if current_presets_ass: m8_preset_var.set(list(current_presets_ass.keys())[0])
f_m8_ps = create_ass_preset_bar(f_m8_custom, [e_m8_font, e_m8_size, e_m8_col, e_m8_ocol, e_m8_mv, e_m8_mlr, e_m8_outl, e_m8_align, e_m8_shad, e_m8_bold, e_m8_ita, e_m8_alpha, e_m8_outalpha], None, [c_btn_m8, oc_btn_m8], m8_preset_var, [e_m8_resx, e_m8_resy])
f_m8_ps.grid(row=4, column=0, columnspan=8, sticky="w", pady=5)

f_m8_ref = ttk.Frame(f8_bot)
m8_ref_path, m8_ref_style = tk.StringVar(), tk.StringVar()
f_m8_top = ttk.Frame(f_m8_ref); f_m8_top.pack(anchor="w", pady=5)
ttk.Label(f_m8_top, text="外部 ASS:").pack(side=tk.LEFT)
ttk.Entry(f_m8_top, textvariable=m8_ref_path, width=30).pack(side=tk.LEFT, padx=5)
ttk.Button(f_m8_top, text="浏览...", command=lambda: ask_file(m8_ref_path, "选择", [("ASS","*.ass")])).pack(side=tk.LEFT, padx=5)
m8_ref_cb = ttk.Combobox(f_m8_top, textvariable=m8_ref_style, width=15)
ttk.Button(f_m8_top, text="扫描样式 ->", command=lambda: scan_ref_for_cb(m8_ref_path.get(), m8_ref_cb, m8_ref_style)).pack(side=tk.LEFT, padx=5)
m8_ref_cb.pack(side=tk.LEFT, padx=5)

e_m8_font_mode, e_m8_override_font = tk.IntVar(value=0), tk.StringVar(value="SimHei")
f_m8_bot2 = ttk.Frame(f_m8_ref); f_m8_bot2.pack(anchor="w", pady=5)
ttk.Radiobutton(f_m8_bot2, text="保留参考样式原字体", variable=e_m8_font_mode, value=0).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_m8_bot2, text="覆盖字体为:", variable=e_m8_font_mode, value=1).pack(side=tk.LEFT, padx=5)
cb_m8_ref_font = ttk.Combobox(f_m8_bot2, textvariable=e_m8_override_font, width=25)
cb_m8_ref_font.pack(side=tk.LEFT, padx=5)
update_m8_ui()
# ======================= 这里开始插入 =======================
# ------ 功能6: 批量重设分辨率 ------
etab_res = ttk.Frame(edit_nb, padding=10)
edit_nb.add(etab_res, text="批量重设分辨率")

ttk.Label(etab_res, text="此功能将直接修改 ASS 文件的 [Script Info] 头信息。\n只需填好输入/输出文件夹，配置下方数值后点击下方【执行全部批处理】即可。", foreground="gray").pack(anchor="w", pady=(5, 10))

f_res_inputs = ttk.Frame(etab_res)
f_res_inputs.pack(anchor="w", pady=5)

edit_res_x_var = tk.StringVar(value="1080")
edit_res_y_var = tk.StringVar(value="1920")

ttk.Label(f_res_inputs, text="目标分辨率 宽 (PlayResX):").pack(side=tk.LEFT)
ttk.Entry(f_res_inputs, textvariable=edit_res_x_var, width=10).pack(side=tk.LEFT, padx=(5, 20))

ttk.Label(f_res_inputs, text="高 (PlayResY):").pack(side=tk.LEFT)
ttk.Entry(f_res_inputs, textvariable=edit_res_y_var, width=10).pack(side=tk.LEFT, padx=5)

# 新增：等比缩放样式参数选项
edit_res_scale_var = tk.IntVar(value=1)
ttk.Checkbutton(etab_res, text="同时按比例缩放样式中的字号、左右/上下边距、阴影和描边等参数", variable=edit_res_scale_var).pack(anchor="w", pady=15)
# ======================= 插入结束 =======================

def execute_ass_editor(stage_only=False):
    i_dir, o_dir = edit_in_var.get().strip(), edit_out_var.get().strip()
    if not i_dir or (not stage_only and not o_dir): return messagebox.showwarning("警告", "请填好输入和输出目录！")
    
    if not stage_only: os.makedirs(o_dir, exist_ok=True)
    mode = edit_nb.index(edit_nb.select())

    fmt = "ASS"
    if mode == 3: fmt = f4_format_var.get()
    
    ext = '.srt' if fmt == "SRT" else '.ass'
    files = [f for f in os.listdir(i_dir) if f.lower().endswith(ext)]
    if not files: return messagebox.showwarning("警告", f"输入文件夹中没有 {ext} 文件！")

    # 新增：用于收集正则查找/替换的报告数据
    global_report_data = []

    # ======================= 智能旁路逻辑：放行空操作直接输出暂存 =======================
    is_empty = False
    if mode == 0 and not lb_m0_vals.curselection(): is_empty = True
    elif mode == 1:
        if (m2_c2_var.get() == 1 and not lb_m2_effs.curselection()) or (m2_c3_var.get() == 1 and not lb_m2_styles.curselection()): is_empty = True
    elif mode == 2 and not m3_ref_dir.get().strip(): is_empty = True
    elif mode == 3:
        if (f4_c2_var.get() == 1 and not lb_f4_effs.curselection()) or (f4_c3_var.get() == 1 and not lb_f4_styles.curselection()): is_empty = True
    elif mode == 4 and not m7_tree.selection(): is_empty = True
    elif mode == 5:
        use_c1, use_c2, use_c3 = f8_c1_var.get() == 1, f8_c2_var.get() == 1, f8_c3_var.get() == 1
        if not use_c1 and not use_c2 and not use_c3: is_empty = True
        elif (use_c2 and not lb_f8_effs.curselection()) or (use_c3 and not lb_f8_styles.curselection()): is_empty = True
    elif mode == 6 and (not edit_res_x_var.get().strip() or not edit_res_y_var.get().strip()): is_empty = True

    # 如果当前标签页没操作、当前点击的是“输出”按钮，并且内存里有暂存的数据
    if is_empty and not stage_only and global_ass_memory_cache:
        mode = -1  # 将处理模式变更为旁路模式，绕过下方所有的标签页检查，直接将内存写出到文件
    # ==============================================================================
    
    if mode == 2:
        ref_dir = m3_ref_dir.get().strip()
        if not ref_dir or not os.path.exists(ref_dir): return messagebox.showwarning("警告", "请选择有效的参考 ASS 文件夹！")
        sync_type = m3_sync_type.get()
        keep_font = m3_keep_font.get() == 1
        all_errors = []

    global_ref_resx, global_ref_resy = None, None
    rp = None
    if mode == 0:
        if edit_m0_mode.get() == 1: rp = m0_ref_path.get().strip()
        else:
            if e_m0_resx.get().strip(): global_ref_resx = f"PlayResX: {e_m0_resx.get().strip()}"
            if e_m0_resy.get().strip(): global_ref_resy = f"PlayResY: {e_m0_resy.get().strip()}"
    elif mode == 1:
        if m2_c2_var.get() == 1 and not lb_m2_effs.curselection():
            return messagebox.showwarning("警告", "勾选了特效条件，但未在列表中选中任何特效！")
        if m2_c3_var.get() == 1 and not lb_m2_styles.curselection():
            return messagebox.showwarning("警告", "勾选了样式条件，但未在列表中选中任何样式！")
            
        if edit_m2_mode.get() == 1: rp = m2_ref_path.get().strip()
    elif mode == 4:
        if edit_m7_mode.get() == 1: rp = m7_ref_path.get().strip()
        else:
            if e_m7_resx.get().strip(): global_ref_resx = f"PlayResX: {e_m7_resx.get().strip()}"
            if e_m7_resy.get().strip(): global_ref_resy = f"PlayResY: {e_m7_resy.get().strip()}"
    elif mode == 5:
        if edit_m8_mode.get() == 1: rp = m8_ref_path.get().strip()
        else:
            if e_m8_resx.get().strip(): global_ref_resx = f"PlayResX: {e_m8_resx.get().strip()}"
            if e_m8_resy.get().strip(): global_ref_resy = f"PlayResY: {e_m8_resy.get().strip()}"
            
    
    if rp and os.path.exists(rp):
        with open(rp, 'r', encoding='utf-8-sig') as f:
            for l in f:
                if l.startswith('PlayResX:'): global_ref_resx = l.strip()
                elif l.startswith('PlayResY:'): global_ref_resy = l.strip()

    regex_rules = []
    if mode == 3: raw_text = f4_regex_text.get("1.0", tk.END).split('\n')
    else: raw_text = []
    
    for line in raw_text:
        line_clean = line.strip('\r\n') # 只删换行符，完美保留左右可能需要的空格
        if not line_clean: continue
        
        if '>>>' in line_clean:
            pat, repl = line_clean.split('>>>', 1)
            # 仅精准剥离 '>>>' 旁边紧挨着的一个多余空格
            if pat.endswith(' '): pat = pat[:-1]
            if repl.startswith(' '): repl = repl[1:]
        else:
            # 如果没有 >>>，则把整行视为纯查找模式
            pat = line_clean
            repl = ""
            
        repl_python = re.sub(r'\$(\d+)', r'\\\1', repl)
        
        # 修复 Bug：无论是有 >>> 还是没有 >>>，最后都必须将规则加进列表！
        regex_rules.append((pat, repl_python))

    for file in files:
        # 【修复1】：删除了 mode 4 的强制跳过，让所有文件都能走完最后的保存流程
        in_path, out_path = os.path.join(i_dir, file), os.path.join(o_dir, file)
        
        # ====== 内存优先加载 ======
        if file in global_ass_memory_cache:
            content = global_ass_memory_cache[file]
        else:
            with open(in_path, 'r', encoding='utf-8-sig') as f: content = f.read()
            global_ass_memory_cache[file] = content

        # ====== SRT 格式文件独立处理逻辑 ======
        if fmt == "SRT" and mode == 3:
            blocks = re.split(r'\n\s*\n', content.strip())
            parsed_blocks = []
            for block in blocks:
                lines = block.strip().split('\n')
                if len(lines) >= 3:
                    parsed_blocks.append({'ID': lines[0].strip(), 'Timeline': lines[1].strip(), 'Text': "\n".join(lines[2:]).strip()})
                    
            b = f4_bracket_var.get().strip()
            logic_mode = f4_logic_var.get()
            use_c1 = f4_c1_var.get() == 1
            # 对 SRT 会在底层判定器中自动忽略后两个参数
            sel_effs, sel_styles = [], []

            for block in parsed_blocks:
                p = [block['ID'], block['Timeline'], block['Text']]
                tgt_idx = int(f4_target_col.get().split(':')[0])
                
                # 直接传入正则规则 b
                is_match = evaluate_advanced_condition("SRT", p, logic_mode, use_c1, b, False, sel_effs, False, sel_styles)
                    
                if is_match:
                    orig_tgt = p[tgt_idx]
                    current_val = orig_tgt
                    matched_parts = []
                    find_only = f4_find_only_var.get() == 1
                    gen_report = f4_report_var.get() == 1
                    
                    for pat, repl in regex_rules:
                        # 核心修改：利用 finditer 准确提取匹配到的文本
                        matches = list(re.finditer(pat, current_val))
                        if matches:
                            matched_parts.extend([m.group(0) for m in matches])
                            if not find_only:
                                current_val = re.sub(pat, repl, current_val)
                                
                    # ====== 在上面这段 for 循环下方，直接粘贴插入 ======
                    punct_mode = f4_punct_mode.get()
                    punct_changed = False
                    if not find_only and punct_mode in (1, 2):
                        if tgt_idx == 2: # 如果用户正好选了处理文本列
                            new_txt = safe_punct_convert(current_val, punct_mode)
                            if new_txt != current_val:
                                punct_changed = True; current_val = new_txt
                        else: # 无论用户选了什么列，都在最后单独清洗一遍文本列 (p[2])
                            new_txt = safe_punct_convert(p[2], punct_mode)
                            if new_txt != p[2]:
                                punct_changed = True; p[2] = new_txt
                    # ====================================================

                    # 覆盖原来的 if matched_parts:
                    if matched_parts or punct_changed:
                        if find_only:
                            if gen_report and matched_parts:
                                global_report_data.append([file, p[1], p[2], orig_tgt, " | ".join(matched_parts)])
                        else:
                            if gen_report:
                                global_report_data.append([file, p[1], p[2], orig_tgt, " | ".join(matched_parts) if matched_parts else "[仅执行了标点转换]", current_val])
                            p[tgt_idx] = current_val
                        
                block['ID'], block['Timeline'], block['Text'] = p[0], p[1], p[2]
            
            srt_content = [f"{b['ID']}\n{b['Timeline']}\n{b['Text']}\n" for b in parsed_blocks]
            final_content = "\n".join(srt_content)
            
            # 新增：如果是只查找模式，拦截保存操作（不污染内存，也不输出文件）
            if f4_find_only_var.get() == 0:
                if stage_only: global_ass_memory_cache[file] = final_content
                else: 
                    global_ass_memory_cache[file] = final_content
                    with open(out_path, 'w', encoding='utf-8') as f: f.write(final_content)
            continue
        # ================= ASS 处理 =================
        lines = content.split('\n')
        h_lines, s_lines, ev_lines = [], [], []
        curr = "info"
        tgt_format_line = "Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding"

        for line in lines:
            l = line.strip()
            if l.startswith('[V4+ Styles]'): curr = "styles"
            elif l.startswith('[Events]'): curr = "events"
            
            if curr == "info": h_lines.append(line)
            elif curr == "styles": 
                s_lines.append(line)
                if l.startswith('Format:'): tgt_format_line = l
            elif curr == "events": ev_lines.append(line)
            
        if len(s_lines) <= 1: s_lines = ["[V4+ Styles]", tgt_format_line]

        ref_resx, ref_resy = global_ref_resx, global_ref_resy
        if mode == 2 and sync_type == 0:
            ref_path = os.path.join(ref_dir, file)
            if os.path.exists(ref_path):
                with open(ref_path, 'r', encoding='utf-8-sig') as f:
                    for l in f:
                        if l.startswith('PlayResX:'): ref_resx = l.strip()
                        elif l.startswith('PlayResY:'): ref_resy = l.strip()
        
        if ref_resx and ref_resy:
            has_x = has_y = False
            for i, l in enumerate(h_lines):
                if l.startswith('PlayResX:'): h_lines[i] = ref_resx; has_x = True
                elif l.startswith('PlayResY:'): h_lines[i] = ref_resy; has_y = True
            if not has_x: h_lines.append(ref_resx)
            if not has_y: h_lines.append(ref_resy)

        # ====== 功能0 ======
        if mode == 0:
            sel_idxs = lb_m0_vals.curselection()
            if not sel_idxs: return messagebox.showwarning("警告", "请在列表中选中至少一项要修改的数据！")
            sel_vals = set(lb_m0_vals.get(i) for i in sel_idxs)
            col_idx = int(m0_col_var.get().split(':')[0])
            
            new_style_name = m0_target_style_var.get().strip()
            if not new_style_name: return messagebox.showwarning("警告", "请输入赋予的新样式名称！")
            
            new_line = ""
            if edit_m0_mode.get() == 0:
                new_line = build_ass_style_line(new_style_name, e_m0_font.get(), e_m0_size.get(), e_m0_col.get(), e_m0_ocol.get(), e_m0_mv.get(), e_m0_mlr.get(), e_m0_outl.get(), e_m0_align.get(), e_m0_shad.get(), e_m0_bold.get(), e_m0_ita.get(), e_m0_alpha.get(), e_m0_outalpha.get())
            else:
                rp, rs = m0_ref_path.get(), m0_ref_style.get()
                if not os.path.exists(rp) or not rs: return messagebox.showwarning("警告", "请正确提供参考文件和样式！")
                ref_dict = scan_all_styles_from_ass(rp)
                if rs not in ref_dict: return messagebox.showwarning("错误", "参考中没找到该样式")
                new_line = rename_style_line(ref_dict[rs], new_style_name)
                if e_m0_font_mode.get() == 1:
                    new_line = replace_font_in_style(new_line, e_m0_override_font.get())

            matched_any = False
            new_ev = []
            for ev in ev_lines:
                if ev.startswith('Dialogue:'):
                    p = ev.split(',', 9)
                    if len(p) > col_idx and p[col_idx].strip() in sel_vals:
                        p[3] = new_style_name  # 统一将 Style 引用修改为新样式
                        matched_any = True
                        new_ev.append(",".join(p))
                    else: new_ev.append(ev)
                else: new_ev.append(ev)
            ev_lines = new_ev

            if matched_any:
                rep = False
                for i, sl in enumerate(s_lines):
                    if sl.startswith('Style:') and sl.split('Style:')[1].split(',')[0].strip() == new_style_name:
                        s_lines[i] = new_line; rep = True
                if not rep: s_lines.append(new_line)
                
        # ====== 功能1 ======
        # ====== 功能1 ======
        elif mode == 1:
            b = edit_m2_bracket.get().strip()
            logic_mode = m2_logic_var.get()
            use_c1 = m2_c1_var.get() == 1
            use_c2 = m2_c2_var.get() == 1
            use_c3 = m2_c3_var.get() == 1
            sel_effs = [lb_m2_effs.get(i) for i in lb_m2_effs.curselection()]
            sel_styles = [lb_m2_styles.get(i) for i in lb_m2_styles.curselection()]

            if use_c2 and not sel_effs: return messagebox.showwarning("警告", "功能2：勾选了特效条件，但未在列表中选中任何特效！")
            if use_c3 and not sel_styles: return messagebox.showwarning("警告", "功能2：勾选了样式条件，但未在列表中选中任何样式！")

            n_line, s_line = "", ""
            if edit_m2_mode.get() == 0:
                n_line = build_ass_style_line("对白字幕", e_m2n_font.get(), e_m2n_size.get(), e_m2n_col.get(), e_m2n_ocol.get(), e_m2n_mv.get(), e_m2n_mlr.get(), e_m2n_outl.get(), e_m2n_align.get(), e_m2n_shad.get(), e_m2n_bold.get(), e_m2n_ita.get(), e_m2n_alpha.get(), e_m2n_outalpha.get())
                s_line = build_ass_style_line("画面字", e_m2s_font.get(), e_m2s_size.get(), e_m2s_col.get(), e_m2s_ocol.get(), e_m2s_mv.get(), e_m2s_mlr.get(), e_m2s_outl.get(), e_m2s_align.get(), e_m2s_shad.get(), e_m2s_bold.get(), e_m2s_ita.get(), e_m2s_alpha.get(), e_m2s_outalpha.get())
                n_name, s_name = "对白字幕", "画面字"
            else:
                rp, rn, rs = m2_ref_path.get(), m2_ref_n.get(), m2_ref_s.get()
                if not os.path.exists(rp) or not rn or not rs: return messagebox.showwarning("警告", "请正确提供参考文件和两类样式！")
                ref_dict = scan_all_styles_from_ass(rp)
                if rn not in ref_dict or rs not in ref_dict: return messagebox.showwarning("错误", "参考文件中未找到指定的对白或画面字样式！")
                n_line = rename_style_line(ref_dict[rn], rn)
                s_line = rename_style_line(ref_dict[rs], rs)
                if e_m2_font_mode.get() == 1:
                    n_line = replace_font_in_style(n_line, e_m2_override_font.get())
                    s_line = replace_font_in_style(s_line, e_m2_override_font.get())
                n_name, s_name = rn, rs

            for nl in [n_line, s_line]:
                nm = nl.split('Style:')[1].split(',')[0].strip()
                rep = False
                for i, l in enumerate(s_lines):
                    if l.strip().startswith('Style:') and l.split('Style:')[1].split(',')[0].strip() == nm:
                        s_lines[i] = nl; rep = True
                if not rep: s_lines.append(nl)

            new_ev = []
            for ev in ev_lines:
                parts = ev.split(',', 9)
                if len(parts) >= 10 and ev.strip().startswith('Dialogue:'):
                    # 直接传入正则规则 b
                    is_screen = evaluate_advanced_condition("ASS", parts, logic_mode, use_c1, b, use_c2, sel_effs, use_c3, sel_styles)
                    if is_screen: parts[3] = s_name
                    else: parts[3] = n_name
                    new_ev.append(",".join(parts))
                else: new_ev.append(ev)
            ev_lines = new_ev

        # ====== 功能2 ======
        elif mode == 2:
            ref_path = os.path.join(ref_dir, file)
            if not os.path.exists(ref_path):
                all_errors.append({'文件名': file, '时间轴': 'N/A', '文本': 'N/A', '错误': '参考文件夹中无同名文件'})
                continue
            
            ref_styles = scan_all_styles_from_ass(ref_path)
            ref_events = {}
            with open(ref_path, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    if line.startswith('Dialogue:'):
                        p = line.strip().split(',', 9)
                        if len(p) >= 10: ref_events[(p[1].strip(), p[2].strip())] = p
            
            tgt_styles = scan_all_styles_from_ass(in_path)
            new_ev = []
            styles_to_add = {}

            for ev in ev_lines:
                if ev.startswith('Dialogue:'):
                    p = ev.split(',', 9)
                    if len(p) >= 10:
                        start, end = p[1].strip(), p[2].strip()
                        if (start, end) in ref_events:
                            rp = ref_events[(start, end)]
                            if sync_type == 1: p[8] = rp[8]
                            else:
                                new_style_name = rp[3]
                                orig_style_name = p[3]
                                p[3] = new_style_name
                                if new_style_name in ref_styles:
                                    st_line = ref_styles[new_style_name]
                                    if keep_font:
                                        orig_font = "Arial"
                                        if orig_style_name in tgt_styles:
                                            orig_font = tgt_styles[orig_style_name].split('Style:')[1].split(',')[1].strip()
                                        st_parts = st_line.split('Style:')[1].split(',')
                                        if len(st_parts) > 1: st_parts[1] = orig_font
                                        st_line = "Style:" + ",".join(st_parts)
                                    styles_to_add[new_style_name] = st_line
                        else:
                            all_errors.append({'文件名': file, '时间轴': f"{start} --> {end}", '文本': p[9], '错误': '在参考文件中未找到对应时间轴'})
                    new_ev.append(",".join(p))
                else: new_ev.append(ev)
            ev_lines = new_ev

            if sync_type == 0:
                for n_name, n_line in styles_to_add.items():
                    rep = False
                    for i, sl in enumerate(s_lines):
                        if sl.startswith('Style:') and sl.split('Style:')[1].split(',')[0].strip() == n_name:
                            s_lines[i] = n_line; rep = True
                    if not rep: s_lines.append(n_line)

        # ====== 功能3 (批量/条件正则替换 - ASS处理) ======
        elif mode == 3:
            b = f4_bracket_var.get().strip()
            logic_mode = f4_logic_var.get()
            use_c1 = f4_c1_var.get() == 1
            use_c2 = f4_c2_var.get() == 1
            use_c3 = f4_c3_var.get() == 1
            sel_effs = [lb_f4_effs.get(i) for i in lb_f4_effs.curselection()]
            sel_styles = [lb_f4_styles.get(i) for i in lb_f4_styles.curselection()]

            if use_c2 and not sel_effs: return messagebox.showwarning("警告", "正则替换：勾选了特效条件，但未选中任何特效！")
            if use_c3 and not sel_styles: return messagebox.showwarning("警告", "正则替换：勾选了样式条件，但未选中任何样式！")

            tgt_idx = int(f4_target_col.get().split(':')[0])
            
            new_ev = []
            for ev in ev_lines:
                if ev.startswith('Dialogue:'):
                    p = ev.split(',', 9)
                    if len(p) >= 10:
                        # 直接传入正则规则 b
                        is_match = evaluate_advanced_condition("ASS", p, logic_mode, use_c1, b, use_c2, sel_effs, use_c3, sel_styles)
                        
                        if is_match:
                            orig_tgt = p[tgt_idx]
                            current_val = orig_tgt
                            matched_parts = []
                            find_only = f4_find_only_var.get() == 1
                            gen_report = f4_report_var.get() == 1
                            
                            for pat, repl in regex_rules:
                                matches = list(re.finditer(pat, current_val))
                                if matches:
                                    matched_parts.extend([m.group(0) for m in matches])
                                    if not find_only:
                                        current_val = re.sub(pat, repl, current_val)
                                        
                            # ====== 在上面这段 for 循环下方，直接粘贴插入 ======
                            punct_mode = f4_punct_mode.get()
                            punct_changed = False
                            if not find_only and punct_mode in (1, 2):
                                if tgt_idx == 9: # 如果用户选的是文本列
                                    new_txt = safe_punct_convert(current_val, punct_mode)
                                    if new_txt != current_val:
                                        punct_changed = True; current_val = new_txt
                                else: # 单独清洗 ASS 的文本列 (p[9])
                                    new_txt = safe_punct_convert(p[9], punct_mode)
                                    if new_txt != p[9]:
                                        punct_changed = True; p[9] = new_txt
                            # ====================================================

                            # 覆盖原来的 if matched_parts:
                            if matched_parts or punct_changed:
                                timeline = f"{p[1]} --> {p[2]}"
                                if find_only:
                                    if gen_report and matched_parts:
                                        global_report_data.append([file, timeline, p[9], orig_tgt, " | ".join(matched_parts)])
                                else:
                                    if gen_report:
                                        global_report_data.append([file, timeline, p[9], orig_tgt, " | ".join(matched_parts) if matched_parts else "[仅执行了标点转换]", current_val])
                                    p[tgt_idx] = current_val
                        
                        # ====== 致命 Bug 修复：把处理完（或未处理）的字幕行塞回列表中 ======
                        new_ev.append(",".join(p))
                    else: new_ev.append(ev)
                else: new_ev.append(ev)
            ev_lines = new_ev
        # ====== 功能4 ======
        elif mode == 4:
            # 【修复2】：只对当前在下拉框里选中的文件，应用树状图里的选中修改
            if file == m7_file_var.get().strip():
                sel_items = m7_tree.selection()
                if not sel_items: return messagebox.showwarning("警告", "请在下方列表中点击选择你要修改的字幕行！")
                sel_indices = [int(m7_tree.item(item, 'values')[0]) for item in sel_items]
                
                new_style_name = edit_m7_target_var.get().strip()
                if not new_style_name: return messagebox.showwarning("警告", "请输入赋予的新样式名称！")
                
                new_line = ""
                if edit_m7_mode.get() == 0:
                    new_line = build_ass_style_line(new_style_name, e_m7_font.get(), e_m7_size.get(), e_m7_col.get(), e_m7_ocol.get(), e_m7_mv.get(), e_m7_mlr.get(), e_m7_outl.get(), e_m7_align.get(), e_m7_shad.get(), e_m7_bold.get(), e_m7_ita.get(), e_m7_alpha.get(), e_m7_outalpha.get())
                else:
                    rp, rs = m7_ref_path.get(), m7_ref_style.get()
                    if not os.path.exists(rp) or not rs: return messagebox.showwarning("警告", "请正确提供参考文件和样式！")
                    ref_dict = scan_all_styles_from_ass(rp)
                    if rs not in ref_dict: return messagebox.showwarning("错误", "参考中没找到该样式")
                    new_line = rename_style_line(ref_dict[rs], new_style_name)
                    if e_m7_font_mode.get() == 1: new_line = replace_font_in_style(new_line, e_m7_override_font.get())
                    
                rep = False
                for i, sl in enumerate(s_lines):
                    if sl.startswith('Style:') and sl.split('Style:')[1].split(',')[0].strip() == new_style_name:
                        s_lines[i] = new_line; rep = True
                if not rep: s_lines.append(new_line)
                
                new_ev = []
                for i, ev in enumerate(ev_lines):
                    if i in sel_indices and ev.startswith('Dialogue:'):
                        p = ev.split(',', 9)
                        if len(p) >= 10:
                            p[3] = new_style_name
                            new_ev.append(",".join(p))
                        else: new_ev.append(ev)
                    else: new_ev.append(ev)
                ev_lines = new_ev
            else:
                # 【核心】：如果是暂存过的其他文件，什么都不做，带着它原本的内存数据直接进入保存环节！
                pass

        # ====== 功能5: 条件定位替换样式 ======
        elif mode == 5:
            b = f8_bracket_var.get().strip()
            logic_mode = f8_logic_var.get()
            use_c1 = f8_c1_var.get() == 1
            use_c2 = f8_c2_var.get() == 1
            use_c3 = f8_c3_var.get() == 1
            sel_effs = [lb_f8_effs.get(i) for i in lb_f8_effs.curselection()]
            sel_styles = [lb_f8_styles.get(i) for i in lb_f8_styles.curselection()]

            if not use_c1 and not use_c2 and not use_c3:
                return messagebox.showwarning("警告", "请至少启用一个定位条件！")
            if use_c2 and not sel_effs: return messagebox.showwarning("警告", "条件替换：勾选了特效条件，但未选中特效！")
            if use_c3 and not sel_styles: return messagebox.showwarning("警告", "条件替换：勾选了样式条件，但未选中样式！")
            
            new_style_name = edit_m8_target_var.get().strip()
            if not new_style_name: return messagebox.showwarning("警告", "请输入赋予的新样式名称！")
            
            new_line = ""
            if edit_m8_mode.get() == 0:
                new_line = build_ass_style_line(new_style_name, e_m8_font.get(), e_m8_size.get(), e_m8_col.get(), e_m8_ocol.get(), e_m8_mv.get(), e_m8_mlr.get(), e_m8_outl.get(), e_m8_align.get(), e_m8_shad.get(), e_m8_bold.get(), e_m8_ita.get(), e_m8_alpha.get(), e_m8_outalpha.get())
            else:
                rp, rs = m8_ref_path.get(), m8_ref_style.get()
                if not os.path.exists(rp) or not rs: return messagebox.showwarning("警告", "请正确提供参考文件和样式！")
                ref_dict = scan_all_styles_from_ass(rp)
                if rs not in ref_dict: return messagebox.showwarning("错误", "参考中没找到该样式")
                new_line = rename_style_line(ref_dict[rs], new_style_name)
                if e_m8_font_mode.get() == 1: new_line = replace_font_in_style(new_line, e_m8_override_font.get())
                
            matched_any = False
            new_ev = []
            for ev in ev_lines:
                if ev.startswith('Dialogue:'):
                    p = ev.split(',', 9)
                    if len(p) >= 10:
                        # 直接传入正则规则 b
                        is_match = evaluate_advanced_condition("ASS", p, logic_mode, use_c1, b, use_c2, sel_effs, use_c3, sel_styles)
                            
                        if is_match:
                            p[3] = new_style_name
                            matched_any = True
                        new_ev.append(",".join(p))
                    else: new_ev.append(ev)
                else: new_ev.append(ev)
            ev_lines = new_ev

            if matched_any:
                rep = False
                for i, sl in enumerate(s_lines):
                    if sl.startswith('Style:') and sl.split('Style:')[1].split(',')[0].strip() == new_style_name:
                        s_lines[i] = new_line; rep = True
                if not rep: s_lines.append(new_line)
        # ====== 功能6: 批量重设分辨率及等比缩放 ======
        elif mode == 6:
            try:
                target_rx = float(edit_res_x_var.get().strip())
                target_ry = float(edit_res_y_var.get().strip())
            except:
                return messagebox.showwarning("警告", "分辨率必须是有效的数字！")
                
            do_scale = edit_res_scale_var.get() == 1

            orig_rx, orig_ry = 384.0, 288.0 # ASS默认基准兜底
            rx_idx, ry_idx = -1, -1

            # 1. 扫描当前文件原有的分辨率
            for idx, l in enumerate(h_lines):
                if l.startswith('PlayResX:'):
                    try: orig_rx = float(l.split(':')[1].strip())
                    except: pass
                    rx_idx = idx
                elif l.startswith('PlayResY:'):
                    try: orig_ry = float(l.split(':')[1].strip())
                    except: pass
                    ry_idx = idx
            
            # 2. 覆盖头文件分辨率信息
            if rx_idx != -1: h_lines[rx_idx] = f"PlayResX: {int(target_rx)}"
            else: h_lines.append(f"PlayResX: {int(target_rx)}")
            
            if ry_idx != -1: h_lines[ry_idx] = f"PlayResY: {int(target_ry)}"
            else: h_lines.append(f"PlayResY: {int(target_ry)}")

            # 3. 按动态比例换算所有样式的排版参数
            if do_scale and orig_rx > 0 and orig_ry > 0:
                scale_x = target_rx / orig_rx
                scale_y = target_ry / orig_ry
                scale_min = min(scale_x, scale_y) # 描边阴影采用最保守缩放比例

                for idx, sl in enumerate(s_lines):
                    if sl.startswith('Style:'):
                        prefix, body = sl.split(':', 1)
                        pts = body.split(',')
                        if len(pts) >= 22:
                            try:
                                # Fontsize (2) -> 按 Y轴 缩放
                                pts[2] = f"{float(pts[2]) * scale_y:.2f}".rstrip('0').rstrip('.')
                                # Outline (16), Shadow (17) -> 按 最小比例缩放，防止粗细变形
                                pts[16] = f"{float(pts[16]) * scale_min:.2f}".rstrip('0').rstrip('.')
                                pts[17] = f"{float(pts[17]) * scale_min:.2f}".rstrip('0').rstrip('.')
                                # MarginL (19), MarginR (20) -> 按 X轴 缩放 (强制整数)
                                pts[19] = str(int(round(float(pts[19]) * scale_x)))
                                pts[20] = str(int(round(float(pts[20]) * scale_x)))
                                # MarginV (21) -> 按 Y轴 缩放 (强制整数)
                                pts[21] = str(int(round(float(pts[21]) * scale_y)))
                                
                                s_lines[idx] = prefix + ":" + ",".join(pts)
                            except:
                                pass

        final_content = "\n".join(h_lines) + "\n\n" + "\n".join(s_lines) + "\n\n" + "\n".join(ev_lines)
        
        # --- 根据按钮模式，分流输出目标 ---
        # 新增：如果当前处于正则功能且开启了只查找模式，拦截一切写入行为
        if mode == 3 and f4_find_only_var.get() == 1:
            pass
        else:
            if stage_only:
                global_ass_memory_cache[file] = final_content
            else:
                global_ass_memory_cache[file] = final_content
                with open(out_path, 'w', encoding='utf-8') as f:
                    f.write(final_content)

    # ======================= 这里开始插入：正则报告输出 =======================
    if mode == 3 and f4_report_var.get() == 1 and global_report_data:
        import csv
        out_d = edit_out_var.get().strip()
        if not out_d:
            # 兼容暂存模式下未填输出目录的情况，降级保存在输入目录
            out_d = edit_in_var.get().strip() 
        
        report_path = os.path.join(out_d, "Regex_Report.csv")
        try:
            # 使用 utf-8-sig 以兼容 Excel 直接打开不乱码
            with open(report_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                if f4_find_only_var.get() == 1:
                    writer.writerow(["文件名", "时间轴", "原始字幕内容", "原始对应列内容", "查找匹配到的内容"])
                else:
                    writer.writerow(["文件名", "时间轴", "原始字幕内容", "原始对应列内容", "查找匹配到的内容", "替换后的对应列内容"])
                writer.writerows(global_report_data)
            messagebox.showinfo("报告已生成", f"正则查找/替换报告已生成至：\n{report_path}")
        except Exception as e:
            messagebox.showerror("生成报告失败", f"无法写入报告文件：\n{str(e)}")

    # 新增：针对只查找模式的专属弹窗，避免出现误导情况
    if mode == 3 and f4_find_only_var.get() == 1:
        messagebox.showinfo("扫描完毕", f"查找任务结束！共扫描了 {len(files)} 个文件。\n\n由于您开启了【只查找不替换】模式，本次操作作为纯粹的数据检索，并未对任何字幕文件或内存产生修改/覆盖。")
    elif stage_only:
        messagebox.showinfo("暂存成功", "【处理完毕】更改已无缝注入底层内存引擎！\n\n你可以随时切到其他标签页加载、刷新以进行二次、三次叠加操作。\n等所有流水线走完，只需点击底部【批量输出保存】按钮即可一次性落盘。")
    else:
        if mode == 2 and m3_err_rep.get().strip() and all_errors:
            pd.DataFrame(all_errors).to_excel(m3_err_rep.get().strip(), index=False)
            messagebox.showinfo("完成", f"导出完成！但有 {len(all_errors)} 处时间轴不匹配，已导出报错报告。")
        else:
            messagebox.showinfo("完成", f"完成！所有流式处理已同步到输出文件夹中的 {len(files)} 个文件！")

# ================= TAB 7: 批量复用指定列(SRT/ASS) =================
tab_eff = ttk.Frame(nb_ass, padding=20)
nb_ass.add(tab_eff, text=" 批量复用指定列(SRT/ASS) ")
tab_eff.columnconfigure(1, weight=1)

eff_fmt_var = tk.StringVar(value="ASS")
eff_col_var = tk.StringVar(value=ASS_COLS[8])

def update_eff_cols(*args):
    if eff_fmt_var.get() == "ASS":
        cb_eff_col['values'] = ASS_COLS + ["Header: 文件头(包含样式/信息等)"]
        if eff_col_var.get() not in cb_eff_col['values']:
            eff_col_var.set(ASS_COLS[8])
    else:
        cb_eff_col['values'] = SRT_COLS
        if eff_col_var.get() not in SRT_COLS:
            eff_col_var.set(SRT_COLS[2])
            
    if eff_col_var.get().startswith("Header"):
        f_eff_header.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 5))
    else:
        f_eff_header.grid_remove()

def scan_eff_headers():
    d = eff_src_var.get().strip()
    if not d or not os.path.exists(d): return messagebox.showwarning("提示", "请先在下方选择【提供数据的源文件夹】！")
    files = [f for f in os.listdir(d) if f.lower().endswith('.ass')]
    if not files: return messagebox.showwarning("提示", "源文件夹中没有 .ass 文件！")
    
    headers = set()
    for f in files:
        try:
            with open(os.path.join(d, f), 'r', encoding='utf-8-sig') as fp:
                for line in fp:
                    l = line.strip()
                    if l.startswith('[') and l.endswith(']') and l != '[Events]':
                        headers.add(l)
        except: continue
    lb_eff_headers.delete(0, tk.END)
    for h in sorted(list(headers)): lb_eff_headers.insert(tk.END, h)
    messagebox.showinfo("成功", f"扫描完毕！共发现 {len(headers)} 种文件头区块。")

def run_scan_ext():
    f = ext_file_var.get().strip()
    if not os.path.exists(f): return messagebox.showwarning("错误", "文件不存在")
    s = scan_ass_for_styles(f)
    rx, ry = get_ass_resolution(f)
    ext_res_cache["x"] = rx
    ext_res_cache["y"] = ry
    
    ext_styles_cache.clear()
    ext_styles_cache.update(s)
    k = list(s.keys())
    cb_ext_style['values'] = k
    if k: ext_style_var.set(k[0])
    messagebox.showinfo("成功", f"扫描到 {len(k)} 个样式\n自动获取到视频分辨率: {rx} x {ry}")

eff_col_var.trace_add("write", update_eff_cols)

f_eff_top = ttk.Frame(tab_eff)
f_eff_top.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
ttk.Radiobutton(f_eff_top, text="处理 ASS 格式", variable=eff_fmt_var, value="ASS", command=update_eff_cols).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(f_eff_top, text="处理 SRT 格式", variable=eff_fmt_var, value="SRT", command=update_eff_cols).pack(side=tk.LEFT, padx=5)

ttk.Label(f_eff_top, text="需要复用同步的项:").pack(side=tk.LEFT, padx=(20, 5))
cb_eff_col = ttk.Combobox(f_eff_top, textvariable=eff_col_var, values=ASS_COLS, width=25, state="readonly")
cb_eff_col.pack(side=tk.LEFT)

# --- 新增：文件头替换子选项面板 ---
f_eff_header = ttk.LabelFrame(tab_eff, text="文件头同步子选项 (提取前需扫描源文件夹)", padding=10)

ttk.Button(f_eff_header, text="🔍 扫描源文件夹文件头", command=scan_eff_headers).pack(anchor="w", pady=(0, 5))
ttk.Label(f_eff_header, text="选择需要覆盖同步的文件头区块 (支持按住Ctrl多选，都不选则默认全量覆盖整个文件头):").pack(anchor="w")

f_eff_header_lb = ttk.Frame(f_eff_header)
f_eff_header_lb.pack(fill=tk.X, expand=True, pady=5)
lb_eff_headers = tk.Listbox(f_eff_header_lb, selectmode=tk.MULTIPLE, height=4, exportselection=False)
lb_eff_headers.pack(side=tk.LEFT, fill=tk.X, expand=True)
sb_eff_h = ttk.Scrollbar(f_eff_header_lb, command=lb_eff_headers.yview)
sb_eff_h.pack(side=tk.LEFT, fill=tk.Y)
lb_eff_headers.config(yscrollcommand=sb_eff_h.set)
# -------------------------------

eff_src_var, eff_tgt_var = tk.StringVar(), tk.StringVar()
eff_out_var, eff_err_var = tk.StringVar(), tk.StringVar()

ttk.Label(tab_eff, text="提供数据的源文件夹:").grid(row=2, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_eff, textvariable=eff_src_var).grid(row=2, column=1, sticky="ew", padx=5)
ttk.Button(tab_eff, text="浏览...", command=lambda: ask_dir(eff_src_var, "选择源文件夹")).grid(row=2, column=2, padx=5)

ttk.Label(tab_eff, text="待接收数据的目标文件夹:").grid(row=3, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_eff, textvariable=eff_tgt_var).grid(row=3, column=1, sticky="ew", padx=5)
ttk.Button(tab_eff, text="浏览...", command=lambda: ask_dir(eff_tgt_var, "选择目标文件夹")).grid(row=3, column=2, padx=5)

ttk.Label(tab_eff, text="合成后的新文件输出至:").grid(row=4, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_eff, textvariable=eff_out_var).grid(row=4, column=1, sticky="ew", padx=5)
ttk.Button(tab_eff, text="浏览...", command=lambda: ask_dir(eff_out_var, "选择输出文件夹")).grid(row=4, column=2, padx=5)

ttk.Label(tab_eff, text="报告保存至:").grid(row=5, column=0, sticky="e", pady=10, padx=(0,10))
ttk.Entry(tab_eff, textvariable=eff_err_var).grid(row=5, column=1, sticky="ew", padx=5)
ttk.Button(tab_eff, text="浏览...", command=lambda: ask_save_file(eff_err_var, "保存报错报告", [("Excel", "*.xlsx")], ".xlsx")).grid(row=5, column=2, padx=5)

ttk.Label(tab_eff, text="* 注：将基于【文件同名】提取。指定列模式按行数复用，文件头模式按区块名复用\n如果复用时间轴，时间轴报错请忽略，按行数复用", foreground="gray").grid(row=6, column=0, columnspan=3, pady=(0,10))
ttk.Button(tab_eff, text="执行批量同步", command=run_column_copy, style='TButton').grid(row=7, column=0, columnspan=3, pady=10, ipadx=20, ipady=5)

update_eff_cols() # 初始化面板状态

# ================= TAB 10: ASS 拆分 (画面字/普通字) =================
tab_ass_split = ttk.Frame(nb_ass, padding=10)
nb_ass.add(tab_ass_split, text=" ASS拆分/转SRT ")
tab_ass_split.columnconfigure(1, weight=1)

split_ass_in_var = tk.StringVar()
split_ass_out_scr_var, split_ass_out_norm_var = tk.StringVar(), tk.StringVar()

ttk.Label(tab_ass_split, text="ASS 输入文件夹:").grid(row=0, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ass_split, textvariable=split_ass_in_var).grid(row=0, column=1, sticky="ew", padx=5)
f_split_in_btns = ttk.Frame(tab_ass_split)
f_split_in_btns.grid(row=0, column=2, sticky="w", padx=5)
ttk.Button(f_split_in_btns, text="浏览...", command=lambda: ask_dir(split_ass_in_var, "选择目录")).pack(side=tk.LEFT)

ttk.Label(tab_ass_split, text="画面字 ASS 存至:").grid(row=1, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ass_split, textvariable=split_ass_out_scr_var).grid(row=1, column=1, sticky="ew", padx=5)
ttk.Button(tab_ass_split, text="浏览...", command=lambda: ask_dir(split_ass_out_scr_var, "选择目录")).grid(row=1, column=2, sticky="w", padx=5)

ttk.Label(tab_ass_split, text="普通字 ASS 存至:").grid(row=2, column=0, sticky="e", pady=5, padx=5)
ttk.Entry(tab_ass_split, textvariable=split_ass_out_norm_var).grid(row=2, column=1, sticky="ew", padx=5)
ttk.Button(tab_ass_split, text="浏览...", command=lambda: ask_dir(split_ass_out_norm_var, "选择目录")).grid(row=2, column=2, sticky="w", padx=5)

# 完美复用高级判定组件
f_split_cond, split_ass_logic_var, split_ass_c1_var, split_ass_bracket_var, split_ass_c2_var, lb_split_effs, split_ass_c3_var, lb_split_styles = build_advanced_condition_ui(tab_ass_split, split_ass_in_var, "拆分判定条件 (组合判定为画面字，其余为普通字)")
f_split_cond.grid(row=3, column=0, columnspan=3, sticky="ew", pady=10, padx=5)

# 新增：保存为 SRT 的勾选选项
split_ass_to_srt_var = tk.IntVar(value=0)

ttk.Checkbutton(tab_ass_split, text="拆分后自动剥离特效与样式，直接转为标准 SRT 格式保存", variable=split_ass_to_srt_var).grid(row=4, column=0, columnspan=3, sticky="w", padx=10, pady=(5, 0))

ttk.Button(tab_ass_split, text="▶ 开始拆分 ASS", command=run_ass_split, style='TButton').grid(row=5, column=0, columnspan=3, pady=10, ipadx=20, ipady=5)

# ================= TAB 11: ASS 样式预设提取 =================
tab_ext = ttk.Frame(nb_ass, padding=20)
nb_ass.add(tab_ext, text=" ASS 样式预设提取 ")

ext_file_var = tk.StringVar()
ext_style_var = tk.StringVar()
ext_preset_name = tk.StringVar()
ext_styles_cache = {}
ext_res_cache = {"x": "1080", "y": "1920"}

ttk.Label(tab_ext, text="选择用于提取的 ASS 文件:").grid(row=0, column=0, sticky="e", pady=10)
ttk.Entry(tab_ext, textvariable=ext_file_var, width=40).grid(row=0, column=1, sticky="w", padx=5)
ttk.Button(tab_ext, text="浏览...", command=lambda: ask_file(ext_file_var, "选择ASS", [("ASS","*.ass")])).grid(row=0, column=2, padx=5)

ttk.Button(tab_ext, text="🔍 扫描样式", command=run_scan_ext).grid(row=0, column=3, padx=5)

ttk.Label(tab_ext, text="选择要提取的样式:").grid(row=1, column=0, sticky="e", pady=10)
cb_ext_style = ttk.Combobox(tab_ext, textvariable=ext_style_var, state="readonly", width=30)
cb_ext_style.grid(row=1, column=1, sticky="w", padx=5)

ttk.Label(tab_ext, text="保存为预设名称:").grid(row=2, column=0, sticky="e", pady=10)
ttk.Entry(tab_ext, textvariable=ext_preset_name, width=30).grid(row=2, column=1, sticky="w", padx=5)

def parse_ass_color(ass_str):
    """解析 ASS 颜色代码，返回 (HEX颜色, HEX透明度)"""
    try:
        s = ass_str.strip().upper().replace('&H', '')
        if len(s) >= 8:
            a, b, g, r = s[-8:-6], s[-6:-4], s[-4:-2], s[-2:]
        elif len(s) == 6:
            a, b, g, r = "00", s[-6:-4], s[-4:-2], s[-2:]
        else:
            s = s.zfill(6)
            a, b, g, r = "00", s[-6:-4], s[-4:-2], s[-2:]
        return f"#{r}{g}{b}", a
    except: 
        return "#FFFFFF", "00"

def parse_style_to_dict(style_line):
    parts = style_line.split('Style:')[1].split(',')
    c_hex, c_a = parse_ass_color(parts[3].strip())
    oc_hex, oc_a = parse_ass_color(parts[5].strip())
    return {
        "font": parts[1].strip(), "size": parts[2].strip(),
        "color": c_hex, "alpha": c_a, "out_color": oc_hex, "out_alpha": oc_a,
        "margin_v": parts[21].strip(), "margin_lr": parts[19].strip(),
        "outline": parts[16].strip(), "align": parts[18].strip(),
        "shadow": parts[17].strip(), "bold": 1 if parts[7].strip() == "-1" else 0,
        "italic": 1 if parts[8].strip() == "-1" else 0
    }

def save_ext_preset():
    name = ext_preset_name.get().strip()
    if not name: return messagebox.showwarning("警告", "请输入预设名称")
    target_style = ext_style_var.get()
    if target_style not in ext_styles_cache: return messagebox.showwarning("警告", "请先扫描并选择有效样式")
    
    d = DEFAULT_PRESETS_ASS["默认样式"].copy()
    pd_dict = parse_style_to_dict(ext_styles_cache[target_style])
    
    # 将提取到的样式同时应用给预设底层的双字段，确保在其他页面的面板加载时完美映射
    d.update({
        "play_res_x": ext_res_cache["x"], "play_res_y": ext_res_cache["y"],
        "n_font": pd_dict["font"], "n_size": pd_dict["size"], "n_color": pd_dict["color"], "n_alpha": pd_dict["alpha"], "n_out_color": pd_dict["out_color"], "n_out_alpha": pd_dict["out_alpha"], "n_margin_v": pd_dict["margin_v"], "n_margin_lr": pd_dict["margin_lr"], "n_outline": pd_dict["outline"], "n_align": pd_dict["align"], "n_shadow": pd_dict["shadow"], "n_bold": pd_dict["bold"], "n_italic": pd_dict["italic"],
        "s_font": pd_dict["font"], "s_size": pd_dict["size"], "s_color": pd_dict["color"], "s_alpha": pd_dict["alpha"], "s_out_color": pd_dict["out_color"], "s_out_alpha": pd_dict["out_alpha"], "s_margin_v": pd_dict["margin_v"], "s_margin_lr": pd_dict["margin_lr"], "s_outline": pd_dict["outline"], "s_align": pd_dict["align"], "s_shadow": pd_dict["shadow"], "s_bold": pd_dict["bold"], "s_italic": pd_dict["italic"]
    })

    current_presets_ass[name] = d

    save_presets_to_file(PRESET_FILE_ASS, current_presets_ass)
    update_all_ass_preset_cbs()
    messagebox.showinfo("成功", f"预设 [{name}] 已提取并保存！\n\n该预设已同步至所有编辑界面，且包含了完整的透明度、阴影等 13 项样式参数。")

ttk.Button(tab_ext, text="💾 提取并保存为全局 ASS 预设", command=save_ext_preset, style='TButton').grid(row=3, column=0, columnspan=3, pady=20, ipadx=20, ipady=5)

def run_xlsx_merge():
    in_dir = ext_merge_in_var.get().strip()
    out_file = ext_merge_out_var.get().strip()
    if not in_dir or not out_file:
        return messagebox.showwarning("警告", "请完整选择输入文件夹和输出文件路径！")
    try:
        import openpyxl
        from copy import copy
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        files = [f for f in os.listdir(in_dir) if f.lower().endswith('.xlsx') and not f.startswith('~')]
        if not files:
            return messagebox.showwarning("警告", "输入文件夹中没有找到有效的 .xlsx 文件！")
        
        files.sort()
        
        # 创建新的母表
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        current_out_row = 1
        
        for i, f in enumerate(files):
            fp = os.path.join(in_dir, f)
            
            try:
                wb_in = openpyxl.load_workbook(fp, rich_text=True)
            except TypeError:
                wb_in = openpyxl.load_workbook(fp)
            
            ws_in = wb_in.active
            
            # 继承列宽，整体向右偏移一列
            if i == 0:
                # 给第一列（From列）设置一个默认的合适宽度
                ws_out.column_dimensions['A'].width = 25
                for col_letter, col_dim in ws_in.column_dimensions.items():
                    try:
                        old_idx = column_index_from_string(col_letter)
                        new_letter = get_column_letter(old_idx + 1)
                        ws_out.column_dimensions[new_letter].width = col_dim.width
                    except: pass
            
            start_row = 1 if i == 0 else 2
            max_row = ws_in.max_row
            max_col = ws_in.max_column
            
            if max_row < start_row: continue 
            
            for row_idx in range(start_row, max_row + 1):
                
                # ====== 新增：在第一列写入 From 来源信息 ======
                if i == 0 and row_idx == 1:
                    # 第一个文件的第一行，写表头 "From"
                    ws_out.cell(row=current_out_row, column=1, value="From")
                else:
                    # 其余所有数据行，写入当前对应的文件名
                    ws_out.cell(row=current_out_row, column=1, value=f)
                # ===============================================

                # 继承行高
                source_row_dim = ws_in.row_dimensions[row_idx]
                if source_row_dim.height is not None:
                    ws_out.row_dimensions[current_out_row].height = source_row_dim.height

                for col_idx in range(1, max_col + 1):
                    source_cell = ws_in.cell(row=row_idx, column=col_idx)
                    
                    # 核心改动：目标单元格的 column 需要加 1（向右偏移）
                    target_cell = ws_out.cell(row=current_out_row, column=col_idx + 1, value=source_cell.value)
                    
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
                    
                    if hasattr(source_cell, 'hyperlink') and source_cell.hyperlink:
                        target_cell.hyperlink = copy(source_cell.hyperlink)
                        # 修复超链接坐标引发 Excel 报错的机制
                        target_cell.hyperlink.ref = target_cell.coordinate
                
                current_out_row += 1
                
        wb_out.save(out_file)
        messagebox.showinfo("完成", f"合并成功！\n共完美拼接了 {len(files)} 个 XLSX 文件。\n\n💡 已在第一列成功附加【From】文件名来源列！")
    except Exception as e:
        messagebox.showerror("错误", f"合并失败:\n{str(e)}")
        
root.update_idletasks()
try:
    fonts = list(tkfont.families())
    n_cb['values'] = fonts
    s_cb['values'] = fonts
    cb_m0['values'] = fonts     # 新增
    cb_m2n['values'] = fonts
    cb_m2s['values'] = fonts
    cb_msn['values'] = fonts
    cb_mss['values'] = fonts
    
    cb_ref_font_5['values'] = fonts
    cb_ref_font_9['values'] = fonts
    cb_m0_ref_font['values'] = fonts  # 新增
    cb_m2_ref_font['values'] = fonts
    
    cb_m7['values'] = fonts
    cb_m7_ref_font['values'] = fonts
    
    cb_m8['values'] = fonts
    cb_m8_ref_font['values'] = fonts
except: pass

def run_term_check():
    s_dir = tc_src_dir.get().strip()
    t_dir = tc_tgt_dir.get().strip()
    tb_file = tc_tb_path.get().strip()
    out_file = tc_out_path.get().strip()
    scol, tcol = tc_src_col.get().strip(), tc_tgt_col.get().strip()
    
    if not all([s_dir, t_dir, tb_file, out_file, scol, tcol]):
        return messagebox.showwarning("警告", "请填写完整的输入、输出路径及列名！")
        
    try:
        # 1. 加载术语表
        if tb_file.lower().endswith('.csv'): tb_df = pd.read_csv(tb_file)
        else: tb_df = pd.read_excel(tb_file)
        
        use_opt3 = tc_partial_match.get() == 1
        col_cat = tc_cat_col.get().strip()
        sel_cats = [lb_tc_cats.get(i) for i in lb_tc_cats.curselection()]

        if scol not in tb_df.columns or tcol not in tb_df.columns:
            return messagebox.showerror("错误", f"术语表中找不到指定的列名：'{scol}' 或 '{tcol}'")
        if use_opt3 and col_cat and col_cat not in tb_df.columns:
            return messagebox.showerror("错误", f"术语表中找不到分类列：'{col_cat}'\n请确认列名是否正确！")
            
        tb_df[scol] = tb_df[scol].astype(object).fillna('')
        tb_df[tcol] = tb_df[tcol].astype(object).fillna('')
        if use_opt3 and col_cat: tb_df[col_cat] = tb_df[col_cat].astype(object).fillna('')
        
        # 数据结构升级：存储目标翻译的同时，记录该术语拥有的所有分类
        term_dict = {}
        for _, row in tb_df.iterrows():
            st, tt = str(row[scol]).strip(), str(row[tcol]).strip()
            cat = str(row[col_cat]).strip() if use_opt3 and col_cat else ""
            if st and tt and st.lower() not in ['nan', 'none'] and tt.lower() not in ['nan', 'none']:
                if st not in term_dict: term_dict[st] = {'targets': set(), 'cats': set()}
                term_dict[st]['targets'].add(tt)
                if cat: term_dict[st]['cats'].add(cat)
                
        if not term_dict: return messagebox.showwarning("警告", "术语表为空或未提取到有效术语！")

        # 参数预备
        mode = tc_match_mode.get()
        c_range = tc_ctx_range.get()
        strict_mode = tc_strict_ctx.get() == 1
        ctx_regex_pat = tc_strict_syms.get().strip()
        ign_case = tc_ign_case.get() == 1
        ign_count = tc_ign_count.get() == 1
        flags = re.IGNORECASE if ign_case else 0
        
        files = [f for f in os.listdir(s_dir) if f.lower().endswith('.srt')]
        if not files: return messagebox.showwarning("警告", "源语言文件夹中没有 .srt 文件！")
        
        # 准备富文本生成器
        red_font = InlineFont(color='FFFF0000') if RICH_TEXT_SUPPORTED else None
        black_font = InlineFont(color='FF000000') if RICH_TEXT_SUPPORTED else None # 新增：黑色默认字体阻断溢出
        
        def build_rich_text(text, terms_to_highlight):
            if not RICH_TEXT_SUPPORTED or not text or not terms_to_highlight: return text
            sorted_terms = sorted(list(terms_to_highlight), key=len, reverse=True)
            pattern = '|'.join([re.escape(t) for t in sorted_terms])
            rich_elements = []
            last_idx = 0
            for m in re.finditer(pattern, text, flags):
                start, end = m.span()
                if start > last_idx: 
                    # 修复颜色溢出：给普通文本强制加上黑色装甲
                    rich_elements.append(TextBlock(black_font, text[last_idx:start]))
                rich_elements.append(TextBlock(red_font, text[start:end]))
                last_idx = end
                
            if last_idx < len(text): 
                rich_elements.append(TextBlock(black_font, text[last_idx:]))
            
            if not rich_elements: return text
            return CellRichText(*rich_elements)

        if not RICH_TEXT_SUPPORTED: messagebox.showinfo("提示", "当前环境中未检测到支持富文本的 openpyxl 版本，报告将以纯文本形式输出。")

        wb = openpyxl.Workbook() if RICH_TEXT_SUPPORTED else None
        ws = wb.active if wb else None
        if ws: ws.append(["文件名", "字幕id", "时间轴", "源语言字幕内容", "目标语言字幕内容", "该行目标语言字幕的上下文内容", "报错信息"])
        report_data_fallback = [] 

        error_total = 0
        
        for file in files:
            sf_path = os.path.join(s_dir, file)
            tf_path = os.path.join(t_dir, file)
            
            if not os.path.exists(tf_path):
                row = [file, "N/A", "N/A", "N/A", "N/A", "N/A", "目标语言文件夹中缺失同名文件"]
                if ws: ws.append(row)
                else: report_data_fallback.append(row)
                error_total += 1
                continue
                
            src_blocks = parse_srt_file(sf_path)
            tgt_blocks = parse_srt_file(tf_path)
            
            for i in range(max(len(src_blocks), len(tgt_blocks))):
                if i >= len(src_blocks): continue
                sb = src_blocks[i]
                if i >= len(tgt_blocks):
                    row = [file, sb['ID'], sb['Timeline'], sb['Text'], "N/A", "N/A", "目标语言字幕缺失对应行"]
                    if ws: ws.append(row)
                    else: report_data_fallback.append(row)
                    error_total += 1
                    continue
                    
                tb = tgt_blocks[i]
                
                # ====== 新增：彻底清洗非法 XML 控制字符的消毒函数 ======
                def clean_xml(txt):
                    # 过滤掉会导致 Excel 损坏的低位控制字符，但安全保留换行(\n)、制表符(\t)、回车(\r)
                    if not isinstance(txt, str): return ""
                    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', txt)
                
                # 提前提取并清洗文本，杜绝污染
                s_txt_clean = clean_xml(sb['Text'])
                t_txt_clean = clean_xml(tb['Text'])
                # =======================================================

                # 1. 基础匹配验证 (增强清洗：强制剥离可能存在的 \ufeff 隐形BOM字符和多余首尾空格)
                match_errs = []
                s_id_clean = sb['ID'].strip('\ufeff \t\n\r')
                t_id_clean = tb['ID'].strip('\ufeff \t\n\r')
                
                if s_id_clean != t_id_clean: match_errs.append("ID不一致")
                if mode == 1 and sb['Timeline'].strip() != tb['Timeline'].strip(): match_errs.append("时间轴不一致")
                
                if match_errs:
                    row = [file, s_id_clean, sb['Timeline'], s_txt_clean, t_txt_clean, "N/A", "匹配失败: " + " & ".join(match_errs)]
                    if ws: ws.append(row)
                    else: report_data_fallback.append(row)
                    error_total += 1
                    continue
                
                # 2. 构建上下文
                start_idx, end_idx = i, i
                if c_range > 0:
                    if strict_mode and ctx_regex_pat:
                        # 向上追溯：检查上一句的末尾是否符合连贯正则
                        for k in range(i-1, max(-1, i-c_range-1), -1):
                            prev_txt = clean_xml(tgt_blocks[k]['Text']).strip()
                            try:
                                if re.search(ctx_regex_pat, prev_txt): start_idx = k
                                else: break
                            except: break
                        # 向下追溯：检查本句(及顺延句)的末尾是否符合连贯正则
                        for k in range(i, min(len(tgt_blocks)-1, i+c_range)):
                            curr_txt = clean_xml(tgt_blocks[k]['Text']).strip()
                            try:
                                if re.search(ctx_regex_pat, curr_txt): end_idx = k + 1
                                else: break
                            except: break
                    else:
                        start_idx = max(0, i - c_range)
                        end_idx = min(len(tgt_blocks) - 1, i + c_range)
                        
                ctx_lines = [clean_xml(tgt_blocks[k]['Text']) for k in range(start_idx, end_idx + 1)]
                tgt_context = "\n".join(ctx_lines)
                
                # 3. 终极加强版：术语检测 + 长短词屏蔽
                s_txt, t_txt = s_txt_clean, t_txt_clean
                found_src_terms, found_tgt_terms = set(), set()
                line_errors = []
                
                # 新增：建立原文的“屏蔽遮罩”，解决长短词包含误报（例如“齐城主”优先匹配，“城主”便不再截胡）
                s_txt_mask = s_txt
                
                # 按原文术语长度从长到短排序，优先让长词匹配并“消耗”原文
                sorted_src_terms = sorted(term_dict.keys(), key=len, reverse=True)
                
                for src_term in sorted_src_terms:
                    t_info = term_dict[src_term]
                    
                    # 在遮罩文本中查找，避免被长词消耗过的部分重复触发
                    s_count = len(re.findall(re.escape(src_term), s_txt_mask, flags))
                    
                    if s_count > 0:
                        # 命中后，将遮罩中该词对应的位置“涂黑”（替换为等长空格），防止短词重复匹配
                        s_txt_mask = re.sub(re.escape(src_term), ' ' * len(src_term), s_txt_mask, flags=flags)
                        
                        found_src_terms.add(src_term)
                        t_count = 0
                        
                        is_partial = use_opt3 and any(c in sel_cats for c in t_info['cats'])

                        for tgt_term in t_info['targets']:
                            if is_partial:
                                parts = [p.strip() for p in tgt_term.split() if p.strip()]
                                if not parts: parts = [tgt_term]
                                parts.sort(key=len, reverse=True) 
                                
                                pat = "|".join(re.escape(p) for p in parts)
                                matches = re.findall(pat, tgt_context, flags)
                                if matches:
                                    t_count += len(matches)
                                    found_tgt_terms.update(matches)
                            else:
                                matches = re.findall(re.escape(tgt_term), tgt_context, flags)
                                if matches:
                                    t_count += len(matches)
                                    found_tgt_terms.add(tgt_term)

                        if t_count == 0:
                            tgt_str = " 或 ".join(t_info['targets'])
                            line_errors.append(f"缺失翻译: [{src_term}] (应为: {tgt_str})")
                        elif not ign_count and t_count != s_count:
                            tgt_str = " 或 ".join(t_info['targets'])
                            line_errors.append(f"数量不符: [{src_term}] (应为: {tgt_str}) 源{s_count}个/译{t_count}个")
                # 4. 如果有报错，生成带有富文本的行写入
                if line_errors:
                    err_str = " | ".join(line_errors)
                    s_rich = build_rich_text(s_txt, found_src_terms)
                    t_rich = build_rich_text(t_txt, found_tgt_terms)
                    ctx_rich = build_rich_text(tgt_context, found_tgt_terms)
                    
                    row = [file, sb['ID'], sb['Timeline'], s_rich, t_rich, ctx_rich, err_str]
                    if ws:
                        ws.append([""] * 7) # 占位
                        row_idx = ws.max_row
                        for col_idx, val in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if isinstance(val, CellRichText): cell.value = val
                            else: cell.value = val
                    else:
                        report_data_fallback.append([file, sb['ID'], sb['Timeline'], s_txt, t_txt, tgt_context, err_str])
                    error_total += 1

        if ws: wb.save(out_file)
        else: pd.DataFrame(report_data_fallback, columns=["文件名", "字幕id", "时间轴", "源语言字幕内容", "目标语言字幕内容", "该行目标语言字幕的上下文内容", "报错信息"]).to_excel(out_file, index=False)
        
        if error_total == 0: messagebox.showinfo("完成", "检查完毕！未发现任何术语遗漏或匹配错误。")
        else: messagebox.showwarning("检查完成", f"检查完毕！共发现 {error_total} 处异常。\n报告已导出至：\n{out_file}")
    
    except Exception as e:
        messagebox.showerror("错误", f"运行中发生错误:\n{str(e)}")

# --- 选项3专属的分类扫描及选框 ---
def scan_tc_cats():
    tb_file = tc_tb_path.get().strip()
    col_cat = tc_cat_col.get().strip()
    if not tb_file or not os.path.exists(tb_file): return messagebox.showwarning("提示", "请先在上方输入有效的术语表文件路径！")
    if not col_cat: return messagebox.showwarning("提示", "请输入分类列名！")
    try:
        if tb_file.lower().endswith('.csv'): tb_df = pd.read_csv(tb_file)
        else: tb_df = pd.read_excel(tb_file)
        if col_cat not in tb_df.columns: return messagebox.showerror("错误", f"术语表中找不到列名：'{col_cat}'")
        cats = set(str(x).strip() for x in tb_df[col_cat] if pd.notna(x) and str(x).strip() and str(x).lower() not in ['nan', 'none'])
        lb_tc_cats.delete(0, tk.END)
        for c in sorted(list(cats)): lb_tc_cats.insert(tk.END, c)
        messagebox.showinfo("成功", f"扫描完毕！共发现 {len(cats)} 种分类。")
    except Exception as e:
        messagebox.showerror("错误", f"读取失败:\n{str(e)}")

# ================= TAB 13: SRT字幕术语批量检查 =================
tab_term_check = create_scrollable_tab(nb_other, " SRT术语检查 ", padding=20)
tab_term_check.columnconfigure(1, weight=1)

tc_src_dir = tk.StringVar()
tc_tgt_dir = tk.StringVar()
tc_tb_path = tk.StringVar()
tc_out_path = tk.StringVar()

tc_src_col = tk.StringVar(value="zh_CN")
tc_tgt_col = tk.StringVar(value="id_ID")
tc_match_mode = tk.IntVar(value=1) # 0: ID匹配, 1: ID+时间轴匹配
tc_ctx_range = tk.IntVar(value=1)
tc_strict_ctx = tk.IntVar(value=1)
tc_strict_syms = tk.StringVar(value=r"([,\-]$|\.\.\.$|[^.。!?！？”\"';；]$)")
tc_ign_case = tk.IntVar(value=1)
tc_ign_count = tk.IntVar(value=0)

tc_ign_count = tk.IntVar(value=0)

# ====== 新增：选项3 专属控制变量 ======
tc_partial_match = tk.IntVar(value=0)
tc_cat_col = tk.StringVar(value="Type")
# ==================================

# --- UI 布局 ---
f_tc_inputs = ttk.Frame(tab_term_check)
f_tc_inputs.pack(fill=tk.X, pady=10, padx=5)
f_tc_inputs.columnconfigure(1, weight=1) # 让中间的输入框自动拉伸填满空间

# 第一行：源语言文件夹
ttk.Label(f_tc_inputs, text="1. 源语言字幕文件夹:").grid(row=0, column=0, sticky="e", pady=5)
ttk.Entry(f_tc_inputs, textvariable=tc_src_dir).grid(row=0, column=1, sticky="ew", padx=10, pady=5)
ttk.Button(f_tc_inputs, text="浏览...", command=lambda: ask_dir(tc_src_dir, "选择源语言字幕文件夹")).grid(row=0, column=2, padx=5, pady=5)

# 第二行：目标语言文件夹
ttk.Label(f_tc_inputs, text="2. 目标语言字幕文件夹:").grid(row=1, column=0, sticky="e", pady=5)
ttk.Entry(f_tc_inputs, textvariable=tc_tgt_dir).grid(row=1, column=1, sticky="ew", padx=10, pady=5)
ttk.Button(f_tc_inputs, text="浏览...", command=lambda: ask_dir(tc_tgt_dir, "选择目标语言字幕文件夹")).grid(row=1, column=2, padx=5, pady=5)

# 第三行：术语表文件
ttk.Label(f_tc_inputs, text="3. 术语表文件 (Excel/CSV):").grid(row=2, column=0, sticky="e", pady=5)
ttk.Entry(f_tc_inputs, textvariable=tc_tb_path).grid(row=2, column=1, sticky="ew", padx=10, pady=5)
ttk.Button(f_tc_inputs, text="浏览...", command=lambda: tc_tb_path.set(filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.csv")]))).grid(row=2, column=2, padx=5, pady=5)

# 第四行：读取列名设置
ttk.Label(f_tc_inputs, text="   读取列名 ->").grid(row=3, column=0, sticky="e", pady=5)
f_tc_cols = ttk.Frame(f_tc_inputs)
f_tc_cols.grid(row=3, column=1, columnspan=2, sticky="w", padx=10, pady=5)
ttk.Label(f_tc_cols, text="源语言列:").pack(side=tk.LEFT)
ttk.Entry(f_tc_cols, textvariable=tc_src_col, width=15).pack(side=tk.LEFT, padx=(5, 20))
ttk.Label(f_tc_cols, text="目标语言列:").pack(side=tk.LEFT)
ttk.Entry(f_tc_cols, textvariable=tc_tgt_col, width=15).pack(side=tk.LEFT, padx=5)

f_tc_3 = ttk.LabelFrame(tab_term_check, text="匹配与上下文设置", padding=10)
f_tc_3.pack(fill=tk.X, pady=10)
ttk.Radiobutton(f_tc_3, text="基础匹配: 仅校验 ID 一致", variable=tc_match_mode, value=0).grid(row=0, column=0, sticky="w")
ttk.Radiobutton(f_tc_3, text="严格匹配: ID 及 时间轴 均一致", variable=tc_match_mode, value=1).grid(row=0, column=1, sticky="w", padx=20)

f_tc_ctx = ttk.Frame(f_tc_3)
f_tc_ctx.grid(row=1, column=0, columnspan=3, sticky="w", pady=(10,0))
ttk.Label(f_tc_ctx, text="上下文查找范围 (单侧扩展行数):").pack(side=tk.LEFT)
tk.Spinbox(f_tc_ctx, from_=0, to=10, textvariable=tc_ctx_range, width=5).pack(side=tk.LEFT, padx=5)

ttk.Checkbutton(f_tc_ctx, text="启用严格连贯模式 | 连贯条件(满足正则即连入下文):", variable=tc_strict_ctx).pack(side=tk.LEFT)
ttk.Entry(f_tc_ctx, textvariable=tc_strict_syms, width=28).pack(side=tk.LEFT, padx=5)
ttk.Label(f_tc_ctx, text="(只要正则匹配成功，句子即算作连贯。支持复杂组合逻辑)", foreground="gray").pack(side=tk.LEFT)

f_tc_4 = ttk.LabelFrame(tab_term_check, text="术语校验规则", padding=10)
f_tc_4.pack(fill=tk.X, pady=5)

f_tc_4_top = ttk.Frame(f_tc_4)
f_tc_4_top.pack(fill=tk.X, anchor="w")
ttk.Checkbutton(f_tc_4_top, text="选项1: 忽略目标翻译大小写 (例如 ABC 等同 abc)", variable=tc_ign_case).pack(side=tk.LEFT, padx=10)
ttk.Checkbutton(f_tc_4_top, text="选项2: 忽略术语出现数量 (只要出现即不报错)", variable=tc_ign_count).pack(side=tk.LEFT, padx=30)

f_tc_4_mid = ttk.Frame(f_tc_4)
f_tc_4_mid.pack(fill=tk.X, pady=(10, 5), anchor="w")
ttk.Checkbutton(f_tc_4_mid, text="选项3: 部分匹配 (目标翻译术语以空格分隔，任意单词命中即算作翻译成功)", variable=tc_partial_match).pack(side=tk.LEFT, padx=10)


f_tc_4_bot = ttk.Frame(f_tc_4)
f_tc_4_bot.pack(fill=tk.X, anchor="w", padx=30)
ttk.Label(f_tc_4_bot, text="指定应用部分匹配的列名:").pack(side=tk.LEFT)
ttk.Entry(f_tc_4_bot, textvariable=tc_cat_col, width=12).pack(side=tk.LEFT, padx=5)
ttk.Button(f_tc_4_bot, text="🔍 扫描列内容", command=scan_tc_cats).pack(side=tk.LEFT, padx=5)
ttk.Label(f_tc_4_bot, text="选择允许部分匹配的分类 (支持多选):").pack(side=tk.LEFT, padx=(10, 5))

f_tc_lb = ttk.Frame(f_tc_4_bot)
f_tc_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
lb_tc_cats = tk.Listbox(f_tc_lb, selectmode=tk.MULTIPLE, height=3, exportselection=False)
lb_tc_cats.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
sb_tc_cats = ttk.Scrollbar(f_tc_lb, command=lb_tc_cats.yview)
sb_tc_cats.pack(side=tk.LEFT, fill=tk.Y)
lb_tc_cats.config(yscrollcommand=sb_tc_cats.set)

f_tc_5 = ttk.Frame(tab_term_check)
f_tc_5.pack(fill=tk.X, pady=15)
ttk.Label(f_tc_5, text="4. 检查报告输出至 (Excel):").pack(side=tk.LEFT)
ttk.Entry(f_tc_5, textvariable=tc_out_path, width=60).pack(side=tk.LEFT, padx=5)
ttk.Button(f_tc_5, text="浏览", command=lambda: ask_save_file(tc_out_path, "保存术语检查报告", [("Excel", "*.xlsx")], ".xlsx")).pack(side=tk.LEFT)

ttk.Button(tab_tc:=ttk.Frame(tab_term_check), text="🚀 开始执行批量术语检查", command=run_term_check, style='TButton').pack(pady=10, ipadx=30, ipady=5)
tab_tc.pack(fill=tk.X)
# ===============================================================

# ====== 新增：XLSX 批量合并功能 ======
ext_merge_in_var = tk.StringVar()
ext_merge_out_var = tk.StringVar()
# ================= TAB 14: XLSX 批量合并 =================
tab_xlsx_merge = ttk.Frame(nb_other, padding=30)
# 直接将它作为一个全新的标签页，挂载到全局大容器 nb_ass 中
nb_other.add(tab_xlsx_merge, text=" 📊 XLSX批量合并 ")
tab_xlsx_merge.columnconfigure(1, weight=1)

# --- 全新独立标签页的 UI 布局 (间距更宽敞、更美观) ---
ttk.Label(tab_xlsx_merge, text="1. 包含待合并 XLSX 文件的输入文件夹:").grid(row=0, column=0, sticky="e", pady=20)
ttk.Entry(tab_xlsx_merge, textvariable=ext_merge_in_var, width=50).grid(row=0, column=1, sticky="ew", padx=15, pady=20)
ttk.Button(tab_xlsx_merge, text="浏览...", command=lambda: ask_dir(ext_merge_in_var, "选择需要合并的XLSX文件夹")).grid(row=0, column=2, padx=5, pady=20)

ttk.Label(tab_xlsx_merge, text="2. 合并后的新文件保存至 (Excel):").grid(row=1, column=0, sticky="e", pady=20)
ttk.Entry(tab_xlsx_merge, textvariable=ext_merge_out_var, width=50).grid(row=1, column=1, sticky="ew", padx=15, pady=20)
ttk.Button(tab_xlsx_merge, text="浏览...", command=lambda: ask_save_file(ext_merge_out_var, "保存合并后的文件", [("Excel", "*.xlsx")], ".xlsx")).grid(row=1, column=2, padx=5, pady=20)

ttk.Button(tab_xlsx_merge, text="⚡ 开始执行批量合并", command=run_xlsx_merge, style='TButton').grid(row=2, column=0, columnspan=3, pady=40, ipadx=30, ipady=10)

ttk.Label(tab_xlsx_merge, text="* 提示：本功能会以文件名排序合并文件。第一个文件将保留第一行作为表头，其余所有文件会自动剔除第一行再拼接。", foreground="gray").grid(row=3, column=0, columnspan=3)
# ========================================================

# ================= TAB 15: 配音物料处理与 DeepL 翻译 =================
tab_dubbing = create_scrollable_tab(nb_other, " 配音物料处理 ", padding=20)
tab_dubbing.columnconfigure(1, weight=1)

# --- 变量区 ---
dub_info_dir = tk.StringVar()
dub_char_dir = tk.StringVar()
dub_tpl_file = tk.StringVar() # 新增：模板文件路径
dub_out_dir = tk.StringVar()

trans_in_file = tk.StringVar()
trans_out_file = tk.StringVar()
# 改为直接使用完整 URL 变量
deepl_api_url = tk.StringVar(value="")
deepl_tgt_lang = tk.StringVar(value="EN") # DeepLX 通常使用 EN 而不是 EN-US

# === 新增一个用于记录是否使用网页版请求的变量 ===
trans_service_mode = tk.StringVar(value="自定义 API (需填下方地址)")

# 新增：用于自定义底部的固定文案
dub_note_var = tk.StringVar(value="- 音色匹配：主要角色需试音确认\n- 字幕同步：字幕与人声匹配，严格对齐时间轴（需提供修改后的字幕文件）\n- 情感表达：情色戏有感染力，高潮戏有爆发力，日常对话自然\n- 反应声：喘息声、哭泣、笑声、喊叫声等需真实到位")
dub_sample_var = tk.StringVar(value="无需试音，提供样音")

# === 新增：控制翻译运行状态的全局变量 ===
trans_is_running = tk.BooleanVar(value=False)

def stop_xlsx_translation():
    if trans_is_running.get():
        trans_is_running.set(False)
        btn_trans_stop.config(state=tk.DISABLED, text="正在终止...")
# ======================================

# --- 核心函数 2：DeepL/DeepLX XLSX 深度翻译 (多线程防卡死版) ---
def run_xlsx_translation():
    in_file = trans_in_file.get().strip()
    out_file = trans_out_file.get().strip()
    api_url = deepl_api_url.get().strip()
    t_lang = deepl_tgt_lang.get().strip()
    
    service_mode = trans_service_mode.get()
    use_custom_api = (service_mode == "自定义 API (需填下方地址)")
    
    if not all([in_file, out_file, t_lang]):
        return messagebox.showwarning("警告", "请完整填写输入/输出文件路径及目标语言！")
    if use_custom_api and not api_url:
        return messagebox.showwarning("警告", "请填写API请求地址，或在上方选择免配置的网页翻译引擎！")
        
    # 定义一个后台打工函数，所有的耗时操作都在这里面进行
    def translation_worker():
        try:
            import requests
            import openpyxl
            import time
            
            # 加载 Excel
            wb = openpyxl.load_workbook(in_file)
            unique_texts = set()
            
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and val.strip() and not val.startswith('='):
                            unique_texts.add(val.strip())
                            
            if not unique_texts: 
                # 使用 root.after 跨线程安全地调用界面弹窗
                root.after(0, lambda: messagebox.showinfo("提示", "表格中没有检测到需要翻译的文本！"))
                return
            
            text_list = list(unique_texts)
            trans_dict = {}
            total = len(text_list)
            
            # ================= 新增：如果选中 ChatGPT，走 LQA 批量复用逻辑 =================
            if "ChatGPT" in service_mode:
                lqa_instance = next((obj for obj in globals().values() if type(obj).__name__ == "LQA_App"), None)
                if not lqa_instance:
                    root.after(0, lambda: messagebox.showerror("错误", "未能找到 LQA 引擎实例，无法复用大模型功能。"))
                    return

                api_endpoint_val = lqa_instance.api_endpoint.get().strip()
                api_key_val = lqa_instance.api_key.get().strip()
                if not api_endpoint_val or not api_key_val:
                    root.after(0, lambda: messagebox.showerror("错误", "请先在上方配置全局 API 或在 LQA 页面填写接口信息！"))
                    return

                from openai import AzureOpenAI
                client = AzureOpenAI(
                    azure_endpoint=api_endpoint_val,
                    api_key=api_key_val,
                    api_version=DEFAULT_API_VERSION
                )
                
                ui_model_name = dub_model_box.get()
                actual_deployment_name = ENGINES_MAP.get(ui_model_name, ui_model_name)
                t_limit = int(dub_token_limit.get())
                target_lang_code = LANGUAGES_MAP.get(t_lang, t_lang)

                sys_prompt = lqa_instance.build_prompt(
                    target_lang_name=t_lang,
                    target_lang_code=target_lang_code,
                    additional_context=dub_context_text.get(),
                    with_src=False,
                    task_mode="纯翻译 (Translation)", 
                    system_role=dub_role_box.get()
                )

                # ================= 【核心黑科技】：动态拦截 LQA 日志 =================
                original_lqa_log = lqa_instance.log # 暂存原来的日志函数
                
                def hijacked_log(msg):
                    dub_log(msg) # 将底层的所有日志强制重定向到配音的日志框中
                    
                # 【新增修复】：手动初始化 Token 计费器，防止未运行过 LQA 导致找不到该属性
                lqa_instance.total_tokens_used = 0
                # =================================================================

                try:
                    dub_log("\n====== 🚀 开始执行配音物料 ChatGPT 批量翻译 ======")
                    dub_log(f"参数 | 目标语言: {t_lang} | 模型: {ui_model_name} | Token上限: {t_limit}/次")
                    dub_log(f"任务 | 共提取到 {total} 条唯一的待翻译文本，正在分批打包...")
                    
                    current_batch = []
                    current_tokens = 0
                    
                    for i, text in enumerate(text_list):
                        if not trans_is_running.get():
                            if os.path.exists(out_file):
                                try: os.remove(out_file)
                                except: pass
                            root.after(0, lambda: messagebox.showwarning("已终止", "翻译已被手动停止！\\n后续流程已中断，并撤销了输出文件。"))
                            return
                            
                        # 不再仅仅更新按钮，也在日志框中输出一下打包进度
                        root.after(0, lambda curr=i: btn_trans.config(text=f"打包 ChatGPT 请求中... ( {curr} / {total} )"))
                        
                        item = {"i": str(i), "s": str(text)}
                        item_json_str = json.dumps(item, ensure_ascii=False)
                        item_tokens = lqa_instance.estimate_tokens(item_json_str)

                        if current_tokens + item_tokens > t_limit and current_batch:
                            dub_log(f"\\n> 📦 第 {len(trans_dict) + len(current_batch)}/{total} 条进度触发阈值，准备发送 {len(current_batch)} 条文本...")
                            root.after(0, lambda: btn_trans.config(text=f"请求 ChatGPT 翻译批次... (包含 {len(current_batch)} 条文本)"))
                            
                            # --- 核心复用与计费计算 ---
                            tokens_before = lqa_instance.total_tokens_used
                            res = lqa_instance._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                            tokens_after = lqa_instance.total_tokens_used
                            
                            cost = tokens_after - tokens_before
                            dub_log(f"  ✅ 批次翻译成功！本次消耗: {cost} tokens | 当前任务累计消耗: {tokens_after} tokens")
                            
                            for r in res:
                                idx_str = r.get("i", "")
                                if idx_str.isdigit():
                                    trans_dict[text_list[int(idx_str)]] = r.get("r", "")
                                    
                            current_batch = []
                            current_tokens = 0

                        current_batch.append(item)
                        current_tokens += item_tokens

                    if current_batch:
                        dub_log(f"\n> 📦 正在发送最后一批次 {len(current_batch)} 条文本...")
                        root.after(0, lambda: btn_trans.config(text=f"请求 ChatGPT 最后一批次... (包含 {len(current_batch)} 条文本)"))
                        
                        # --- 核心复用与计费计算 ---
                        tokens_before = lqa_instance.total_tokens_used
                        res = lqa_instance._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                        tokens_after = lqa_instance.total_tokens_used
                        
                        cost = tokens_after - tokens_before
                        dub_log(f"  ✅ 批次翻译成功！本次消耗: {cost} tokens | 当前任务累计消耗: {tokens_after} tokens")
                        
                        for r in res:
                            idx_str = r.get("i", "")
                            if idx_str.isdigit():
                                trans_dict[text_list[int(idx_str)]] = r.get("r", "")
                                
                    dub_log("\\n====== ✨ 所有网络翻译请求已完成！正在写入表格... ======")
                    
                except Exception as e:
                    root.after(0, lambda err=str(e): messagebox.showerror("请求失败", f"ChatGPT API 错误:\\n{err}"))
                    dub_log(f"\\n❌ 发生致命错误: {str(e)}")
                    return
                finally:
                    # ================= 无论成功报错，物归原主 =================
                    # 执行完毕后，把 LQA 原来的日志函数还回去，绝不破坏 LQA 原有功能
                    lqa_instance.log = original_lqa_log

                if current_batch:
                    root.after(0, lambda: btn_trans.config(text=f"请求 ChatGPT 最后一批次... (包含 {len(current_batch)} 条文本)"))
                    try:
                        res = lqa_instance._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                        for r in res:
                            idx_str = r.get("i", "")
                            if idx_str.isdigit():
                                trans_dict[text_list[int(idx_str)]] = r.get("r", "")
                    except Exception as e:
                        root.after(0, lambda err=str(e): messagebox.showerror("请求失败", f"ChatGPT API 错误:\n{err}"))
                        return
                        
            else:
                # ================= 原有的逐条请求逻辑 (Google/Bing/自定义API) =================
                # 逐条发送请求
                for i, text in enumerate(text_list):
                # ====== 新增：中断检测与终止输出 ======
                    if not trans_is_running.get():
                        import os
                        # 发现被按下停止键，立即清理刚复制出来的中间文件，拒绝输出残缺文件
                        if os.path.exists(out_file):
                            try: os.remove(out_file)
                            except: pass
                        root.after(0, lambda: messagebox.showwarning("已终止", "翻译已被手动停止！\n后续流程已中断，并撤销了输出文件。"))
                        return # 直接一刀切断后续的所有执行，跳出线程
                    # ====================================

                    root.after(0, lambda curr=i: btn_trans.config(text=f"翻译中... ( {curr} / {total} )，请勿关闭软件"))

                    print(f"\n[{i+1}/{total}] 准备翻译原文: {text}")
                    
                    if service_mode == "Google 原生网页爬虫 (无需API)":
                        # ============ 模式 B：纯 HTML 网页爬虫 ============
                        t_lang_g = t_lang.lower()
                        if t_lang_g == "zh": t_lang_g = "zh-CN"
                        
                        print(f" -> [纯网页爬虫] 目标语言: {t_lang_g}")
                        
                        # 访问真正的 Web 页面地址，而非任何 API 后端
                        url = "https://translate.google.com/m"
                        params = {
                            "sl": "auto",
                            "tl": t_lang_g,
                            "q": text
                        }
                        headers = {
                            "User-Agent": "Mozilla/5.0 (Linux; Android 10; K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36"
                        }
                        
                        res = requests.get(url, params=params, headers=headers, timeout=15)
                        print(f" -> [纯网页响应] 状态码: {res.status_code}")
                        
                        if res.status_code != 200:
                            raise Exception(f"网页请求被拦截 (状态码 {res.status_code})")
                        
                        try:
                            import re
                            import html as html_lib
                            
                            # 纯正爬虫逻辑：在 HTML 源码中寻找存放翻译结果的 <div> 容器
                            html_text = res.text
                            match = re.search(r'<div[^>]*class="[^"]*result-container[^"]*"[^>]*>(.*?)</div>', html_text, re.IGNORECASE | re.DOTALL)
                            
                            if match:
                                raw_result = match.group(1)
                                # 还原网页中的换行标签
                                raw_result = re.sub(r'<br\s*/?>', '\n', raw_result, flags=re.IGNORECASE)
                                # 还原 HTML 实体符号（例如把 &#39; 还原成单引号）
                                translated_text = html_lib.unescape(raw_result)
                                
                                trans_dict[text] = translated_text
                                print(f" -> [HTML抠取结果]: {trans_dict[text][:50]}...")
                            else:
                                print(f" -> [解析失败] 未在网页中找到 result-container，可能网页结构已改变。")
                                trans_dict[text] = text
                        except Exception as e:
                            print(f" -> [爬虫崩溃] 正则解析异常: {str(e)}")
                            trans_dict[text] = text # 兜底保留原文
                            
                        # 真正的爬虫必须伪装人类浏览速度，加入1秒延迟防止触发验证码
                        time.sleep(1)
                    elif service_mode == "Bing 翻译接口 (translators库)":
                        # ============ 模式 C：基于 translators 库的纯净调用 ============
                        t_lang_bing = t_lang.lower()
                        if t_lang_bing == "zh": t_lang_bing = "zh-Hans"
                        
                        print(f" -> [Bing Translators 模式] 目标语言: {t_lang_bing}")
                        
                        try:
                            import translators as ts
                            
                            # 直接调用库封装好的极简翻译接口，摒弃其他冗余内容处理
                            result = ts.translate_text(text, translator='bing', to_language=t_lang_bing)
                            
                            if result:
                                trans_dict[text] = str(result)
                                print(f" -> [Bing 结果]: {trans_dict[text][:50]}...")
                            else:
                                print(" -> [Bing 警告] 翻译返回为空，保留原文")
                                trans_dict[text] = text
                                
                        except ImportError:
                            raise Exception("缺少核心依赖库！\n请先在系统命令行终端运行: pip install translators")
                        except Exception as e:
                            print(f" -> [Bing 翻译崩溃] 异常详情: {str(e)}")
                            trans_dict[text] = text # 解析失败时保留原文兜底
                            
                        # 依然保留适当的延迟，防止高频触发 Bing 的临时 IP 封锁
                        time.sleep(1)
                    elif service_mode == "Google 翻译接口 (translators库)":
                        # ============ 新增模式：基于 translators 库的 Google 调用 ============
                        # Google 的中文代码通常要求是 zh-CN 或 zh-TW
                        t_lang_google = t_lang.lower()
                        if t_lang_google == "zh": t_lang_google = "zh-CN"
                        
                        print(f" -> [Google Translators 模式] 目标语言: {t_lang_google}")
                        
                        try:
                            import translators as ts
                            
                            # 调用 translators 库封装好的 Google 引擎
                            result = ts.translate_text(text, translator='google', to_language=t_lang_google)
                            
                            if result:
                                trans_dict[text] = str(result)
                                print(f" -> [Google 结果]: {trans_dict[text][:50]}...")
                            else:
                                print(" -> [Google 警告] 翻译返回为空，保留原文")
                                trans_dict[text] = text
                                
                        except ImportError:
                            raise Exception("缺少核心依赖库！\n请先在系统命令行终端运行: pip install translators")
                        except Exception as e:
                            print(f" -> [Google 翻译崩溃] 异常详情: {str(e)}")
                            trans_dict[text] = text # 解析失败时保留原文兜底
                            
                        # Google 相对宽容，但依然保留 1 秒延迟防止长期高频调用的临时风控
                        time.sleep(1)
                
                    else:
                        # ============ 模式 C：自定义 API ============
                        print(f" -> [自定义API 模式] 目标语言: {t_lang.upper()}")
                        payload = {'text': text, 'source_lang': 'auto', 'target_lang': t_lang.upper()}
                        res = requests.post(api_url, json=payload, timeout=15)
                        
                        print(f" -> [自定义API 响应] 状态码: {res.status_code}")
                        if res.status_code != 200:
                            raise Exception(f"API接口报错 (状态码 {res.status_code}):\n{res.text}")
                        
                        data = res.json()
                        if 'translations' in data: trans_dict[text] = data['translations'][0]['text']
                        elif 'data' in data: trans_dict[text] = str(data['data'])
                        elif 'text' in data: trans_dict[text] = str(data['text'])
                        else: trans_dict[text] = text
                        
                        print(f" -> [自定义API 结果]: {trans_dict[text]}")
                        time.sleep(0.1)
                    
            # 进度：正在写入文件
            root.after(0, lambda: btn_trans.config(text="正在生成带格式的 Excel 文件..."))
            
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and val.strip() and not val.startswith('='):
                            cell.value = trans_dict.get(val.strip(), val)
                            
            wb.save(out_file)
            root.after(0, lambda: messagebox.showinfo("完成", f"翻译完毕！\n共深度翻译了 {total} 条唯一文本，源文件格式及插入图片已完美保留。"))
            
        except Exception as e:
            root.after(0, lambda err=str(e): messagebox.showerror("翻译失败", f"错误详情:\n{err}"))
        finally:
            # 无论成功、报错还是中途被强行终止，最后都要将按钮恢复到初始状态
            root.after(0, lambda: btn_trans.config(text="🌐 开始执行 XLSX 深度翻译", state=tk.NORMAL))
            root.after(0, lambda: btn_trans_stop.config(text="⏹ 停止翻译", state=tk.DISABLED))
            trans_is_running.set(False)

    # ================= 主线程干的活 =================
    trans_is_running.set(True) # 开启红绿灯状态
    
    btn_trans.config(text="正在读取并分析表格...", state=tk.DISABLED)
    btn_trans_stop.config(state=tk.NORMAL) # 激活停止按钮
    root.update()

    threading.Thread(target=translation_worker, daemon=True).start()
    
# --- 核心函数 1：配音物料表批量合并（严格对应纯净版模板格式） ---
def run_dubbing_merge():
    i_dir = dub_info_dir.get().strip()
    c_dir = dub_char_dir.get().strip()
    t_file = dub_tpl_file.get().strip()
    o_dir = dub_out_dir.get().strip()
    
    if not all([i_dir, c_dir, t_file, o_dir]):
        return messagebox.showwarning("警告", "请完整选择信息表、角色表、模板文件和输出文件夹！")
        
    try:
        import openpyxl
        from copy import copy
        import io
        from openpyxl.drawing.image import Image as xlImage
        from openpyxl.styles import Border, Side, Font, Alignment
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # 定义宋体12号的基础样式与边框
        ft_normal = Font(name='宋体', size=12)
        ft_bold = Font(name='宋体', size=12, bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', wrap_text=True)
        
        # 定义用于富文本局部加粗的内联字体
        inline_bold = InlineFont(rFont='宋体', sz=12, b=True)
        inline_norm = InlineFont(rFont='宋体', sz=12, b=False)
        
        info_files = {f: os.path.join(i_dir, f) for f in os.listdir(i_dir) if f.lower().endswith('.xlsx') and not f.startswith('~')}
        char_files = {f: os.path.join(c_dir, f) for f in os.listdir(c_dir) if f.lower().endswith('.xlsx') and not f.startswith('~')}
        common_files = set(info_files.keys()).intersection(set(char_files.keys()))
        
        if not common_files: return messagebox.showwarning("警告", "两个文件夹中没有找到同名的 XLSX 文件！")
        os.makedirs(o_dir, exist_ok=True)
        processed_count = 0
        
        for file in common_files:
            # ================= 1. 提取信息表 =================
            wb_info = openpyxl.load_workbook(info_files[file], data_only=True)
            ws_info = wb_info.active
            info_dict = {}
            for r in range(1, ws_info.max_row + 1):
                k = str(ws_info.cell(r, 1).value or "").strip()
                v = str(ws_info.cell(r, 2).value or "").strip()
                if k: info_dict[k] = v
            
            # ================= 2. 提取角色表图片映射 =================
            wb_char = openpyxl.load_workbook(char_files[file], data_only=True)
            ws_char = wb_char.active
            
            img_map = {}
            for img in getattr(ws_char, '_images', []):
                try:
                    r_idx = img.anchor._from.row + 1 
                    img_map[r_idx] = img
                except: pass

            # ================= 3. 载入模板并注入数据 =================
            wb_tpl = openpyxl.load_workbook(t_file)
            ws_tpl = wb_tpl.active
            
            # 填入顶部信息 (A列，拼接内容，利用富文本完美保留标题加粗)
            ws_tpl.cell(2, 1).value = CellRichText(TextBlock(inline_bold, "剧集名称："), TextBlock(inline_norm, info_dict.get('剧集名称', '')))
            ws_tpl.cell(3, 1).value = CellRichText(TextBlock(inline_bold, "总集数："), TextBlock(inline_norm, info_dict.get('总集数', '')))
            ws_tpl.cell(4, 1).value = CellRichText(TextBlock(inline_bold, "配音语言："), TextBlock(inline_norm, info_dict.get('配音语言', '')))
            ws_tpl.cell(6, 1).value = CellRichText(TextBlock(inline_bold, "剧集梗概："), TextBlock(inline_norm, info_dict.get('简介', '')))
            
            start_r = 7
            
            # 解除 start_r 行以下所有的合并单元格，防止清理时报错
            merged_ranges = list(ws_tpl.merged_cells.ranges)
            for m_range in merged_ranges:
                if m_range.min_row >= start_r:
                    ws_tpl.unmerge_cells(str(m_range))
            
            # 清理旧数据
            for row in ws_tpl.iter_rows(min_row=start_r, max_row=ws_tpl.max_row):
                for cell in row: 
                    cell.value = None
                    cell.border = Border()
            
            current_out_row = start_r
            
            # ================= 4. 遍历并写入角色表 =================
            for r in range(2, ws_char.max_row + 1):
                name = str(ws_char.cell(r, 1).value or "").strip()
                if not name: continue
                
                identity = str(ws_char.cell(r, 3).value or "").strip()
                gender = str(ws_char.cell(r, 4).value or "").strip()
                age = str(ws_char.cell(r, 5).value or "").strip()
                desc = str(ws_char.cell(r, 6).value or "").strip()
                voice = str(ws_char.cell(r, 7).value or "").strip()
                
                # 构造富文本对象 (局部加粗的宋体)
                text_blocks = [
                    TextBlock(inline_bold, name),
                    TextBlock(inline_norm, f"\n\n年龄：{age}\n身份：{identity}\n性格：{desc}\n声线：{voice}")
                ]
                
                # 写入基本框架 (列A至G，即1至7)
                ws_tpl.cell(current_out_row, 1, "") # 第1列：留空
                ws_tpl.cell(current_out_row, 2, "") # 第2列：留空
                c_info = ws_tpl.cell(current_out_row, 3) # 第3列：角色信息富文本
                c_info.value = CellRichText(*text_blocks)
                
                # 给第1至7列全部刷上宋体、居中和边框
                for col_idx in range(1, 8):
                    c = ws_tpl.cell(current_out_row, col_idx)
                    # c.border = thin_border
                    c.alignment = align_center
                    if col_idx != 3: c.font = ft_normal
                
                # 规则：角色信息占4个单元格横向合并 (C, D, E, F)
                ws_tpl.merge_cells(start_row=current_out_row, start_column=3, end_row=current_out_row, end_column=6)
                
                # 图片安全提取并放置在第7列 (G列)
                if r in img_map:
                    try:
                        old_img = img_map[r]
                        img_bytes = old_img._data() if hasattr(old_img, '_data') else old_img.ref.getvalue()
                        new_img = xlImage(io.BytesIO(img_bytes))
                        
                        # ====== 核心修复：智能等比例缩放图片，完美限制在单元格内部 ======
                        orig_w, orig_h = old_img.width, old_img.height
                        # 设定单元格边界的最大容纳像素 (留出一点边距防贴边)
                        max_w, max_h = 100, 135 
                        if orig_w > 0 and orig_h > 0:
                            # 计算缩放比例，以最长的一边为准进行等比缩小
                            ratio = min(max_w / orig_w, max_h / orig_h)
                            new_img.width = int(orig_w * ratio)
                            new_img.height = int(orig_h * ratio)
                        # ==================================================================
                        
                        ws_tpl.add_image(new_img, f"G{current_out_row}")
                    except: pass
                    
                # 根据文本行数动态撑开行高
                line_count = 6 + (len(desc)//20) 
                # 确保行高至少有 110 磅 (约146像素)，足以将 135 像素高的图片完美包裹其中
                ws_tpl.row_dimensions[current_out_row].height = max(110, line_count * 18)
                current_out_row += 1
                
            # ================= 5. 写入底部注意事项 =================
            # 规则：占5个单元格横向合并
            ws_tpl.cell(current_out_row, 1, "配音注意事项").font = ft_bold  # 修复：恢复 A 列标题加粗
            ws_tpl.cell(current_out_row, 2, dub_note_var.get().strip()).font = ft_normal
            for col_idx in range(1, 7):
                c = ws_tpl.cell(current_out_row, col_idx)
                # c.border = thin_border
                c.alignment = align_center
            ws_tpl.merge_cells(start_row=current_out_row, start_column=2, end_row=current_out_row, end_column=6)
            ws_tpl.row_dimensions[current_out_row].height = 85
            
            ws_tpl.cell(current_out_row + 1, 1, "试样集数").font = ft_bold  # 修复：恢复 A 列标题加粗
            ws_tpl.cell(current_out_row + 1, 2, dub_sample_var.get().strip()).font = ft_normal
            for col_idx in range(1, 7):
                c = ws_tpl.cell(current_out_row + 1, col_idx)
                # c.border = thin_border
                c.alignment = align_center
            ws_tpl.merge_cells(start_row=current_out_row + 1, start_column=2, end_row=current_out_row + 1, end_column=6)
            ws_tpl.row_dimensions[current_out_row + 1].height = 30
            
            # 保存
            out_name = f"{os.path.splitext(file)[0]}_前期物料表.xlsx"
            wb_tpl.save(os.path.join(o_dir, out_name))
            processed_count += 1
            
        messagebox.showinfo("完成", f"批量合并完成！\n共生成了 {processed_count} 份纯净版物料表。\n宋体12号、加粗样式及图片均已完美贴合模板格式！")
    except Exception as e:
        messagebox.showerror("错误", f"处理过程中发生错误:\n{str(e)}")



# --- UI 布局区 ---
# 模块 1：配音物料合并
f_dub = ttk.LabelFrame(tab_dubbing, text=" 📂 第一步：批量合并 [配音信息表] 与 [配音角色表] ", padding=15)
f_dub.pack(fill=tk.X, pady=10)
f_dub.columnconfigure(1, weight=1)

ttk.Label(f_dub, text="配音信息表 文件夹:").grid(row=0, column=0, sticky="e", pady=5)
ttk.Entry(f_dub, textvariable=dub_info_dir).grid(row=0, column=1, sticky="ew", padx=10)
ttk.Button(f_dub, text="浏览...", command=lambda: ask_dir(dub_info_dir, "选择配音信息表文件夹")).grid(row=0, column=2, padx=5)

ttk.Label(f_dub, text="配音角色表 文件夹:").grid(row=1, column=0, sticky="e", pady=5)
ttk.Entry(f_dub, textvariable=dub_char_dir).grid(row=1, column=1, sticky="ew", padx=10)
ttk.Button(f_dub, text="浏览...", command=lambda: ask_dir(dub_char_dir, "选择配音角色表文件夹")).grid(row=1, column=2, padx=5)

# === 新增：选择模板文件 ===
ttk.Label(f_dub, text="严格对齐排版的 模板文件:").grid(row=2, column=0, sticky="e", pady=5)
ttk.Entry(f_dub, textvariable=dub_tpl_file).grid(row=2, column=1, sticky="ew", padx=10)
ttk.Button(f_dub, text="浏览...", command=lambda: dub_tpl_file.set(filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))).grid(row=2, column=2, padx=5)

ttk.Label(f_dub, text="最终生成的物料表 输出至:").grid(row=3, column=0, sticky="e", pady=5)
ttk.Entry(f_dub, textvariable=dub_out_dir).grid(row=3, column=1, sticky="ew", padx=10)
ttk.Button(f_dub, text="浏览...", command=lambda: ask_dir(dub_out_dir, "选择输出文件夹")).grid(row=3, column=2, padx=5)

# === 新增：自定义配音要求 ===
ttk.Label(f_dub, text="配音注意事项:").grid(row=4, column=0, sticky="ne", pady=5)
tk.Text(f_dub, width=65, height=4).grid(row=4, column=1, columnspan=2, sticky="ew", padx=10, pady=5)
f_dub.children[list(f_dub.children.keys())[-1]].insert("1.0", dub_note_var.get())
f_dub.children[list(f_dub.children.keys())[-1]].bind("<KeyRelease>", lambda e: dub_note_var.set(e.widget.get("1.0", tk.END)))

ttk.Label(f_dub, text="试样集数 要求:").grid(row=5, column=0, sticky="e", pady=5)
ttk.Entry(f_dub, textvariable=dub_sample_var).grid(row=5, column=1, columnspan=2, sticky="ew", padx=10, pady=5)

ttk.Button(f_dub, text="⚡ 按照模板严格合并生成物料表", command=run_dubbing_merge, style='TButton').grid(row=6, column=0, columnspan=3, pady=15, ipadx=20)

# 模块 2：DeepL 无损翻译
f_trans = ttk.LabelFrame(tab_dubbing, text=" 🌐 第二步：DeepL API 无损排版翻译 (支持任意 Excel) ", padding=15)
f_trans.pack(fill=tk.X, pady=10)
f_trans.columnconfigure(1, weight=1)

ttk.Label(f_trans, text="需翻译的原始文件 (Excel):").grid(row=0, column=0, sticky="e", pady=5)
ttk.Entry(f_trans, textvariable=trans_in_file).grid(row=0, column=1, sticky="ew", padx=10)
ttk.Button(f_trans, text="浏览...", command=lambda: trans_in_file.set(filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))).grid(row=0, column=2, padx=5)

ttk.Label(f_trans, text="翻译后另存为 (Excel):").grid(row=1, column=0, sticky="e", pady=5)
ttk.Entry(f_trans, textvariable=trans_out_file).grid(row=1, column=1, sticky="ew", padx=10)
ttk.Button(f_trans, text="浏览...", command=lambda: ask_save_file(trans_out_file, "保存翻译后文件", [("Excel", "*.xlsx")], ".xlsx")).grid(row=1, column=2, padx=5)

# --- 新增：翻译服务引擎选择 ---
f_service = ttk.Frame(f_trans)
f_service.grid(row=2, column=0, columnspan=3, sticky="w", pady=(15, 5))
ttk.Label(f_service, text="翻译服务引擎:").pack(side=tk.LEFT, padx=(0, 5))
service_opts = [
    "自定义 API (需填下方地址)", 
    "Google 原生网页爬虫 (无需API)",
    "Google 翻译接口 (translators库)",
    "Bing 翻译接口 (translators库)",
    "ChatGPT API (复用 LQA 引擎)"
]
ttk.Combobox(f_service, textvariable=trans_service_mode, values=service_opts, width=32, state="readonly").pack(side=tk.LEFT)

f_api = ttk.Frame(f_trans)
f_api.grid(row=3, column=0, columnspan=3, sticky="w", pady=5)
ttk.Label(f_api, text="自定义 API 请求地址:").pack(side=tk.LEFT, padx=(0, 5))
ttk.Entry(f_api, textvariable=deepl_api_url, width=65).pack(side=tk.LEFT)

f_lang = ttk.Frame(f_trans)
f_lang.grid(row=4, column=0, columnspan=3, sticky="w", pady=5)
ttk.Label(f_lang, text="目标语言 (如 EN, ZH, ID, 可手动输入):").pack(side=tk.LEFT, padx=(0, 5))
ttk.Combobox(f_lang, textvariable=deepl_tgt_lang, values=["EN", "ZH", "JA", "KO", "ID", "ES", "RU", "FR", "DE"], width=15).pack(side=tk.LEFT)

# ================= 新增：配音翻译的专属 ChatGPT 配置区 =================
f_dub_chatgpt = ttk.Frame(f_trans)

# 1. 角色与模型
ttk.Label(f_dub_chatgpt, text="系统角色:").grid(row=0, column=0, sticky="w", pady=2)
dub_role_box = ttk.Combobox(f_dub_chatgpt, values=[
    "You are a professional native translator with extensive experience in script and dubbing translation.",
    "You are an expert in subtitle localization and voice-over adaptation.",
    "You are a colloquial dialogue translator, making translations sound natural and spoken.",
    "You are a professional native translator with extensive experience in Localized Translation Field"
], width=45)
dub_role_box.current(0)
dub_role_box.grid(row=0, column=1, sticky="w", padx=5)

ttk.Label(f_dub_chatgpt, text="选择模型:").grid(row=0, column=2, sticky="w", padx=(10,0))
dub_model_box = ttk.Combobox(f_dub_chatgpt, values=list(ENGINES_MAP.keys()), width=15)
dub_model_box.current(0)
dub_model_box.grid(row=0, column=3, sticky="w", padx=5)

# 2. 背景要求与 Token
ttk.Label(f_dub_chatgpt, text="要求/背景:").grid(row=1, column=0, sticky="w", pady=2)
dub_context_text = ttk.Entry(f_dub_chatgpt, width=47)
dub_context_text.insert(0, "准确翻译配音文案，保留语气和专有名词，输出符合本地化习惯的自然对白。")
dub_context_text.grid(row=1, column=1, sticky="w", padx=5)

ttk.Label(f_dub_chatgpt, text="Token/次:").grid(row=1, column=2, sticky="w", padx=(10,0))
dub_token_limit = ttk.Entry(f_dub_chatgpt, width=15)
dub_token_limit.insert(0, "2000")
dub_token_limit.grid(row=1, column=3, sticky="w", padx=5)
# ====================================================================

# --- 底部操作按钮组 ---
btn_frame = ttk.Frame(f_trans)
btn_frame.grid(row=6, column=0, columnspan=3, pady=15) # 注意：这里自动下移到了 row=6

btn_trans = ttk.Button(btn_frame, text="🌐 开始执行 XLSX 深度翻译", command=run_xlsx_translation, style='TButton')
btn_trans.pack(side=tk.LEFT, padx=10, ipadx=20)

btn_trans_stop = ttk.Button(btn_frame, text="⏹ 停止翻译", command=stop_xlsx_translation, style='TButton', state=tk.DISABLED)
btn_trans_stop.pack(side=tk.LEFT, padx=10, ipadx=20)
# ===============================================================
# ================= 新增：配音翻译的专属日志显示框 =================
f_dub_log = ttk.Frame(f_trans)
dub_log_text = scrolledtext.ScrolledText(f_dub_log, height=10, width=80, state=tk.DISABLED, bg="#f9f9f9")
dub_log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

# 3. 动态显示/隐藏联动
def update_dub_trans_ui(*args):
    mode = trans_service_mode.get()
    if "ChatGPT" in mode:
        f_dub_chatgpt.grid(row=5, column=0, columnspan=4, sticky="w", pady=(5, 10))
        # 只要选了 ChatGPT 就显示日志框 (放在第7行)
        f_dub_log.grid(row=7, column=0, columnspan=4, sticky="we", pady=(0, 5)) 
        f_api.grid_remove() 
    else:
        f_dub_chatgpt.grid_remove()
        f_dub_log.grid_remove() # 隐藏日志框
        if "自定义 API" in mode: f_api.grid()
        else: f_api.grid_remove()

def dub_log(message):
    """跨线程安全的日志打印函数"""
    def _append():
        dub_log_text.config(state=tk.NORMAL)
        dub_log_text.insert(tk.END, str(message) + "\n")
        dub_log_text.see(tk.END)
        dub_log_text.config(state=tk.DISABLED)
    root.after(0, _append)
# ===============================================================
trans_service_mode.trace_add("write", update_dub_trans_ui)
update_dub_trans_ui() # 初始化执行

update_m0_ui()
update_m8_ui()
update_m2_ui()
update_ass_style_mode_5()
update_ms_style_mode_9()

root.mainloop()